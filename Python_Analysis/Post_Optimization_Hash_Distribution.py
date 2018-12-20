import math
import numpy as np
from openpyxl import load_workbook
import random
import re


def separate_string(input_string):
    items = []
    match = re.match(r"([a-z]+)([0-9]+)", input_string, re.I)
    if match:
        items = match.groups()
    return items


def column_row_index(input_string, column_dist):
    items = separate_string(input_string)
    column_index = items[0]
    row_index = int(items[1]) + column_dist
    return column_index + str(row_index)


def compute_weights(data_list_):
    data_list_ = np.asarray(data_list_)
    contribution_list = []
    weight_list_ = []
    data_list_cumsum = np.cumsum(data_list_)
    for i in range(top_m):
        c_h = data_list_[i]
        temp_contribution = 0
        for j in range(i, top_m):
            denominator = data_list_cumsum[j]
            temp_contribution = 1.0 * temp_contribution + (1.0 * c_h/denominator)
        contribution_list.append(temp_contribution)
    sum_contribution = sum(contribution_list)
    for i in range(top_m):
        temp_weight = (1.0 * contribution_list[i])/sum_contribution
        weight_list_.append(temp_weight)
    return weight_list_


def post_optimization_opt(collision_probility_, weight_list_, total_error_, data_list_, K_List_, L_List_, hash_used_, hash_budget_):
    smallest = min(data_list_)
    smallest_index = data_list_.index(min(data_list_))

    # total_error = 0.412502654
    total_error_gain = 0
    flag = False
    while hash_used_ + smallest <= hash_budget_:
        # the condition of improvement is to check if the total error rate drops

        # loop through, keep track of hash-relocation and error rate drop
        # find the biggest error gain, while keep hash-resources constraints
        # update hash_used, LList
        delta_error_list = []
        for i in range(len(data_list_)):
            cur_k = K_List_[i]
            cur_l = L_List_[i]
            c_old = math.pow((1 - math.pow(collision_probility_, cur_k)), cur_l)

            # each time increment hash layer by 1
            c_new = math.pow((1 - math.pow(collision_probility_, cur_k)), (cur_l+1))
            delta_error_list.append((c_old - c_new))

        # sort and check each delta_error
        delta_error_list = np.asarray(delta_error_list)
        # sort in descending order
        sorted_index = delta_error_list.argsort()[::-1][:len(delta_error_list)]
        sorted_pivot = 0
        while sorted_pivot < len(sorted_index):
            cur_index = sorted_index[sorted_pivot]

            # each time increment hash layer by 1
            # temp_hash_used = hash_used + (LList[cur_index] + 1) * data_list[cur_index]
            temp_hash_used = hash_used_ + data_list_[cur_index]
            if temp_hash_used < hash_budget_:
                L_List_[cur_index] = L_List_[cur_index] + 1
                hash_used_ = hash_used_ + data_list_[cur_index]
                break
            sorted_pivot = sorted_pivot + 1
    total_error = 0
    total_hash_used = 0
    for i in range(len(L_List_)):
        cur_k = K_List_[i]
        cur_l = L_List_[i]
        total_error = total_error + weight_list_[i] * math.pow((1 - math.pow(collision_probility_, cur_k)), cur_l)
        total_hash_used = total_hash_used + data_list_[i] * cur_l
    # print("Updated total error: " + str(total_error))
    # print("total hash used: " + str(total_hash_used))
    # print("Optimized approach done")
    return L_List_


####################################################################################
# for uniformly distributed approach
# uniformly pick one that fits the current hash budget instead of the one minimizing total error rate
def post_optimization_uni(data_list_, L_List_, hash_used_, hash_budget_):
    flag = False
    smallest = min(data_list_)
    data_list_ = np.asarray(data_list_)
    while hash_used_ + smallest <= hash_budget_:
        # sort in descending order
        sorted_index = data_list_.argsort()[::-1][:len(data_list_)]

        temp_pivot_list = []
        # first find all the allow current hash re-allocation
        for i in range(len(sorted_index)):
            temp_pivot_index = sorted_index[i]
            if hash_used_ + data_list_[temp_pivot_index] <= hash_budget_:
                temp_pivot_list.append(temp_pivot_index)
        # randomly pick one from those
        cur_pivot = random.choice(temp_pivot_list)
        # update L_Uni_List, hash_used
        L_List_[cur_pivot] = L_List_[cur_pivot] + 1
        hash_used_ = hash_used_ + data_list_[cur_pivot]
    return L_List_


####################################################################################
# types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
# Data_Types = ['anti_correlated', 'correlated', 'random']
types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
Data_Types = ['anti_correlated']
top_m_cell = 'E1'


# need to update budget cell per data type
budget_cell_anti = 'B4'
budget_cell_corr = 'S4'
budget_cell_rand = 'AI4'


cardinality_cell = 'B2'
top_m_cardinality_anti = 0
top_m_cardinality_corr = 0
top_m_cardinality_random = 0

data_anti_list_10 = ['J6',  'J15', 'J21', 'J30', 'J37', 'J46', 'J51', 'J60', 'J68', 'J77']
k_ranges_anti_10 = ['E6',  'E15', 'E21', 'E30', 'E37', 'E46', 'E51', 'E60', 'E68', 'E77']
l_ranges_opt_anti_10 = ['F6', 'F15', 'F21', 'F30', 'F37', 'F46', 'F51', 'F60', 'F68', 'F77']
l_ranges_max_anti_10 = ['G6', 'G15', 'G21', 'G30', 'G37', 'G46', 'G51', 'G60', 'G68', 'G77']
l_ranges_uni_anti_10 = ['H6', 'H15', 'H21', 'H30', 'H37', 'H46', 'H51', 'H60', 'H68', 'H77']
hash_used_anti_opt_cells_10 = ['I16', 'I31', 'I47', 'I61', 'I78']
hash_used_anti_uni_cells_10 = ['O16', 'O31', 'O47', 'O61', 'O78']

data_corr_list_10 = ['AA6',  'AA15', 'AA21', 'AA30', 'AA37', 'AA46', 'AA51', 'AA60', 'AA68', 'AA77']
k_ranges_corr_10 = ['V6', 'V15', 'V21', 'V30', 'V37', 'V46', 'V51', 'V60', 'V68', 'V77']
l_ranges_opt_corr_10 = ['W6', 'W15', 'W21', 'W30', 'W37', 'W46', 'W51', 'W60', 'W68', 'W77']
l_ranges_max_corr_10 = ['X6', 'X15', 'X21', 'X30', 'X37', 'X46', 'X51', 'X60', 'X68', 'X77']
l_ranges_uni_corr_10 = ['Y6', 'Y15', 'Y21', 'Y30', 'Y37', 'Y46', 'Y51', 'Y60', 'Y68', 'Y77']
hash_used_corr_opt_cells_10 = ['Z16', 'Z31', 'Z47', 'Z61', 'Z78']
hash_used_corr_uni_cells_10 = ['AF16', 'AF31', 'AF47', 'AF61', 'AF78']


data_random_list_10 = ['AQ6',  'AQ15', 'AQ21', 'AQ30', 'AQ37', 'AQ46', 'AQ51', 'AQ60', 'AQ68', 'AQ77']
k_ranges_random_10 = ['AL6', 'AL15', 'AL21', 'AL30', 'AL37', 'AL46', 'AL51', 'AL60', 'AL68', 'AL77']
l_ranges_opt_random_10 = ['AM6', 'AM15', 'AM21', 'AM30', 'AM37', 'AM46', 'AM51', 'AM60', 'AM68', 'AM77']
l_ranges_max_random_10 = ['AN6', 'AN15', 'AN21', 'AN30', 'AN37', 'AN46', 'AN51', 'AN60', 'AN68', 'AN77']
l_ranges_uni_random_10 = ['AO6', 'AO15', 'AO21', 'AO30', 'AO37', 'AO46', 'AO51', 'AO60', 'AO68', 'AO77']
hash_used_rand_opt_cells_10 = ['AP6', 'AP31', 'AP47', 'AP61', 'AP78']
hash_used_rand_uni_cells_10 = ['AV16', 'AV31', 'AV47', 'AV61', 'AV78']

data_anti_list_25 = ['J6', 'J30', 'J38', 'J62', 'J69', 'J93', 'J100', 'J124', 'J131', 'J155']
k_ranges_anti_25 = ['E6', 'E30', 'E38', 'E62', 'E69', 'E93', 'E100', 'E124', 'E131', 'E155']
l_ranges_opt_anti_25 = ['F6', 'F30', 'F38', 'F62', 'F69', 'F93', 'F100', 'F124', 'F131', 'F155']
l_ranges_max_anti_25 = ['G6', 'G30', 'G38', 'G62', 'G69', 'G93', 'G100', 'G124', 'G131', 'G155']
l_ranges_uni_anti_25 = ['H6', 'H30', 'H38', 'H62', 'H69', 'H93', 'H100', 'H124', 'H131', 'H155']
hash_used_anti_opt_cells_25 = ['I31', 'I63', 'I94', 'I125', 'I156']
hash_used_anti_uni_cells_25 = ['O31', 'O63', 'O94', 'O125', 'O156']

data_corr_list_25 = ['AA6', 'AA30', 'AA38', 'AA62', 'AA69', 'AA93', 'AA100', 'AA124', 'AA131', 'AA155']
k_ranges_corr_25 = ['V6', 'V30', 'V38', 'V62', 'V69', 'V93', 'V100', 'V124', 'V131', 'V155']
l_ranges_opt_corr_25 = ['W6', 'W30', 'W38', 'W62', 'W69', 'W93', 'W100', 'W124', 'W131', 'W155']
l_ranges_max_corr_25 = ['X6', 'X30', 'X38', 'X62', 'X69', 'X93', 'X100', 'X124', 'X131', 'X155']
l_ranges_uni_corr_25 = ['Y6', 'Y30', 'Y38', 'Y62', 'Y69', 'Y93', 'Y100', 'Y124', 'Y131', 'Y155']
hash_used_corr_opt_cells_25 = ['Z31', 'Z63', 'Z94', 'Z125', 'Z156']
hash_used_corr_uni_cells_25 = ['AF31', 'AF63', 'AF94', 'AF125', 'AF156']


data_random_list_25 = ['AQ6', 'AQ30', 'AQ38', 'AQ62', 'AQ69', 'AQ93', 'AQ100', 'AQ124', 'AQ131', 'AQ155']
k_ranges_random_25 = ['AL6', 'AL30', 'AL38', 'AL62', 'AL69', 'AL93', 'AL100', 'AL124', 'AL131', 'AL155']
l_ranges_opt_random_25 = ['AM6', 'AM30', 'AM38', 'AM62', 'AM69', 'AM93', 'AM100', 'AM124', 'AM131', 'AM155']
l_ranges_max_random_25 = ['AN6', 'AN30', 'AN38', 'AN62', 'AN69', 'AN93', 'AN100', 'AN124', 'AN131', 'AN155']
l_ranges_uni_random_25 = ['AO6', 'AO30', 'AO38', 'AO62', 'AO69', 'AO93', 'AO100', 'AO124', 'AO131', 'AO155']
hash_used_rand_opt_cells_25 = ['AP31', 'AP63', 'AP94', 'AP125', 'AP156']
hash_used_rand_uni_cells_25 = ['AV31', 'AV63', 'AV94', 'AV125', 'AV156']


data_anti_list_50 = ['J6', 'J55', 'J63', 'J112', 'J120', 'J169', 'J177', 'J226', 'J234', 'J283']
k_ranges_anti_50 = ['E6', 'E55', 'E63', 'E112', 'E120', 'E169', 'E177', 'E226', 'E234', 'E283']
l_ranges_opt_anti_50 = ['F6', 'F55', 'F63', 'F112', 'F120', 'F169', 'F177', 'F226', 'F234', 'F283']
l_ranges_max_anti_50 = ['G6', 'G55', 'G63', 'G112', 'G120', 'G169', 'G177', 'G226', 'G234', 'G283']
l_ranges_uni_anti_50 = ['H6', 'H55', 'H63', 'H112', 'H120', 'H169', 'H177', 'H226', 'H234', 'H283']
hash_used_anti_opt_cells_50 = ['I56', 'I113', 'I170', 'I227', 'I284']
hash_used_anti_uni_cells_50 = ['O56', 'O113', 'O170', 'O227', 'O284']


data_corr_list_50 = ['AA6', 'AA55', 'AA63', 'AA112', 'AA120', 'AA169', 'AA177', 'AA226', 'AA234', 'AA283']
k_ranges_corr_50 = ['V6', 'V55', 'V63', 'V112', 'V120', 'V169', 'V177', 'V226', 'V234', 'V283']
l_ranges_opt_corr_50 = ['W6', 'W55', 'W63', 'W112', 'W120', 'W169', 'W177', 'W226', 'W234', 'W283']
l_ranges_max_corr_50 = ['X6', 'X55', 'X63', 'X112', 'X120', 'X169', 'X177', 'X226', 'X234', 'X283']
l_ranges_uni_corr_50 = ['Y6', 'Y55', 'Y63', 'Y112', 'Y120', 'Y169', 'Y177', 'Y226', 'Y234', 'Y283']
hash_used_corr_opt_cells_50 = ['Z56', 'Z113', 'Z170', 'Z227', 'Z284']
hash_used_corr_uni_cells_50 = ['AF56', 'AF113', 'AF170', 'AF227', 'AF284']


data_random_list_50 = ['AQ6', 'AQ55', 'AQ63', 'AQ112', 'AQ120', 'AQ169', 'AQ177', 'AQ226', 'AQ234', 'AQ283']
k_ranges_random_50 = ['AL6', 'AL55', 'AL63', 'AL112', 'AL120', 'AL169', 'AL177', 'AL226', 'AL234', 'AL283']
l_ranges_opt_random_50 = ['AM6', 'AM55', 'AM63', 'AM112', 'AM120', 'AM169', 'AM177', 'AM226', 'AM234', 'AM283']
l_ranges_max_random_50 = ['AN6', 'AN55', 'AN63', 'AN112', 'AN120', 'AN169', 'AN177', 'AN226', 'AN234', 'AN283']
l_ranges_uni_random_50 = ['AO6', 'AO55', 'AO63', 'AO112', 'AO120', 'AO169', 'AO177', 'AO226', 'AO234', 'AO283']
hash_used_rand_opt_cells_50 = ['AP56', 'AP113', 'AP170', 'AP227', 'AP284']
hash_used_rand_uni_cells_50 = ['AV56', 'AV113', 'AV170', 'AV227', 'AV284']


collision_probility = 0.9
total_error = 0
# K_log_list_start_cell = ['E6', 'V6', 'AL6']
# K_log_minus_list_start_cell = ['']
# K_log_plus_list_start_cell = []
# K_log_plus_plus_list_start_cell = []
# K_log_uni_list_start_cell = []

####################################################################################

# dimensions = [2, 3, 4, 5, 6, 7]
dimensions = [4]
excel_file_dir = './'

# for each excel file
for i in range(len(dimensions)):
    cur_d = dimensions[i]
    # excel_file_name = excel_file_dir + 'Checkpoint_Result_Nov_26_' + str(cur_d) + 'D_test.xlsx'
    excel_file_name = excel_file_dir + str(cur_d) + 'D.xlsx'
    wb = load_workbook(filename=excel_file_name, data_only=True)
    wb1 = load_workbook(filename=excel_file_name)
    wss = wb.get_sheet_names()

    for wwss in wss:
        print(wwss)
        ws = wb.get_sheet_by_name(wwss)
        ws1 = wb1.get_sheet_by_name(wwss)
        top_m = ws[top_m_cell].value
        # ws = wb.get_sheet_by_name(wss[0])
        # ws1 = wb1.get_sheet_by_name(wss[0])
        # top_m = ws[top_m_cell].value
        # print("Sheet name: " + str(ws) + " top-m : " + str(top_m))
        hash_budget_anti = ws[budget_cell_anti].value
        hash_budget_corr = ws[budget_cell_corr].value
        hash_budget_rand = ws[budget_cell_rand].value
        total_cardinality = ws[cardinality_cell].value
        if top_m == 10:
            top_m_cardinality_anti_cell = 'J16'
            top_m_cardinality_anti = ws[top_m_cardinality_anti_cell].value
            # hash_used_anti_opt_cell = 'I16'
            # hash_used_anti_uni_cell = 'O16'
            # hash_used_anti_opt = ws[hash_used_anti_opt_cell].value
            # hash_used_anti_uni = ws[hash_used_anti_uni_cell].value
            hash_used_anti_opt_cells = hash_used_anti_opt_cells_10
            hash_used_anti_uni_cells = hash_used_anti_uni_cells_10
            data_anti_list = data_anti_list_10
            k_ranges_anti = k_ranges_anti_10
            l_ranges_opt_anti = l_ranges_opt_anti_10
            l_ranges_max_anti = l_ranges_max_anti_10
            l_ranges_uni_anti = l_ranges_uni_anti_10

            top_m_cardinality_corr_cell = 'AA16'
            top_m_cardinality_corr = ws[top_m_cardinality_corr_cell].value
            # hash_used_corr_opt_cell = 'Z16'
            # hash_used_corr_uni_cell = 'AF16'
            # hash_used_corr_opt = ws[hash_used_corr_opt_cell].value
            # hash_used_corr_opt = ws[hash_used_corr_uni_cell].value
            hash_used_corr_opt_cells = hash_used_corr_opt_cells_10
            hash_used_corr_uni_cells = hash_used_corr_uni_cells_10
            data_corr_list = data_corr_list_10
            k_ranges_corr = k_ranges_corr_10
            l_ranges_opt_corr = l_ranges_opt_corr_10
            l_ranges_max_corr = l_ranges_max_corr_10
            l_ranges_uni_corr = l_ranges_uni_corr_10

            top_m_cardinality_random_cell = 'AQ16'
            top_m_cardinality_random = ws[top_m_cardinality_random_cell].value
            # hash_used_random_opt_cell = 'AP16'
            # hash_used_random_uni_cell = 'AV16'
            # hash_used_random_opt = ws[hash_used_random_opt_cell].value
            # hash_used_random_uni = ws[hash_used_random_uni_cell].value
            hash_used_rand_opt_cells = hash_used_rand_opt_cells_10
            hash_used_rand_uni_cells = hash_used_rand_uni_cells_10
            data_random_list = data_random_list_10
            k_ranges_random = k_ranges_random_10
            l_ranges_opt_random = l_ranges_opt_random_10
            l_ranges_max_random = l_ranges_max_random_10
            l_ranges_uni_random = l_ranges_uni_random_10
        elif top_m == 25:
            top_m_cardinality_anti_cell = 'J31'
            top_m_cardinality_anti = ws[top_m_cardinality_anti_cell].value
            # hash_used_anti_opt_cell = 'I31'
            # hash_used_anti_uni_cell = 'O31'
            # hash_used_anti_opt = ws[hash_used_anti_opt_cell].value
            # hash_used_anti_uni = ws[hash_used_anti_uni_cell].value
            hash_used_anti_opt_cells = hash_used_anti_opt_cells_25
            hash_used_anti_uni_cells = hash_used_anti_uni_cells_25
            data_anti_list = data_anti_list_25
            k_ranges_anti = k_ranges_anti_25
            l_ranges_opt_anti = l_ranges_opt_anti_25
            l_ranges_max_anti = l_ranges_max_anti_25
            l_ranges_uni_anti = l_ranges_uni_anti_25

            top_m_cardinality_corr_cell = 'AA31'
            top_m_cardinality_corr = ws[top_m_cardinality_corr_cell].value
            # hash_used_corr_opt_cell = 'Z31'
            # hash_used_corr_uni_cell = 'AF31'
            # hash_used_corr_opt = ws[hash_used_corr_opt_cell].value
            # hash_used_corr_uni = ws[hash_used_corr_uni_cell].value
            hash_used_corr_opt_cells = hash_used_corr_opt_cells_25
            hash_used_corr_uni_cells = hash_used_corr_uni_cells_25
            data_corr_list = data_corr_list_25
            k_ranges_corr = k_ranges_corr_25
            l_ranges_opt_corr = l_ranges_opt_corr_25
            l_ranges_max_corr = l_ranges_max_corr_25
            l_ranges_uni_corr = l_ranges_uni_corr_25

            top_m_cardinality_random_cell = 'AQ31'
            top_m_cardinality_random = ws[top_m_cardinality_random_cell].value
            # hash_used_random_opt_cell = 'AP31'
            # hash_used_random_uni_cell = 'AV31'
            # hash_used_random_opt = ws[hash_used_random_opt_cell].value
            # hash_used_random_uni = ws[hash_used_random_uni_cell].value
            hash_used_rand_opt_cells = hash_used_rand_opt_cells_25
            hash_used_rand_uni_cells = hash_used_rand_uni_cells_25
            data_random_list = data_random_list_25
            k_ranges_random = k_ranges_random_25
            l_ranges_opt_random = l_ranges_opt_random_25
            l_ranges_max_random = l_ranges_max_random_25
            l_ranges_uni_random = l_ranges_uni_random_25
        else:
            top_m_cardinality_anti_cell = 'J56'
            top_m_cardinality_anti = ws[top_m_cardinality_anti_cell].value
            # hash_used_anti_opt_cell = 'I56'
            # hash_used_anti_uni_cell = 'O56'
            # hash_used_anti_opt = ws[hash_used_anti_opt_cell].value
            # hash_used_anti_uni = ws[hash_used_anti_uni_cell].value
            hash_used_anti_opt_cells = hash_used_anti_opt_cells_50
            hash_used_anti_uni_cells = hash_used_anti_uni_cells_50
            data_anti_list = data_anti_list_50
            k_ranges_anti = k_ranges_anti_50
            l_ranges_opt_anti = l_ranges_opt_anti_50
            l_ranges_max_anti = l_ranges_max_anti_50
            l_ranges_uni_anti = l_ranges_uni_anti_50

            top_m_cardinality_corr_cell = 'AA56'
            top_m_cardinality_corr = ws[top_m_cardinality_corr_cell].value
            # hash_used_corr_opt_cell = 'Z56'
            # hash_used_corr_uni_cell = 'AF56'
            # hash_used_corr_opt = ws[hash_used_corr_opt_cell].value
            # hash_used_corr_uni = ws[hash_used_corr_uni_cell].value
            hash_used_corr_opt_cells = hash_used_corr_opt_cells_50
            hash_used_corr_uni_cells = hash_used_corr_uni_cells_50
            data_corr_list = data_corr_list_50
            k_ranges_corr = k_ranges_corr_50
            l_ranges_opt_corr = l_ranges_opt_corr_50
            l_ranges_max_corr = l_ranges_max_corr_50
            l_ranges_uni_corr = l_ranges_uni_corr_50

            top_m_cardinality_random_cell = 'AQ56'
            top_m_cardinality_random = ws[top_m_cardinality_random_cell].value
            # hash_used_random_opt_cell = 'AP56'
            # hash_used_random_uni_cell = 'AV56'
            # hash_used_random_opt = ws[hash_used_random_opt_cell].value
            # hash_used_random_uni = ws[hash_used_random_uni_cell].value
            hash_used_rand_opt_cells = hash_used_rand_opt_cells_50
            hash_used_rand_uni_cells = hash_used_rand_uni_cells_50
            data_random_list = data_random_list_50
            k_ranges_random = k_ranges_random_50
            l_ranges_opt_random = l_ranges_opt_random_50
            l_ranges_max_random = l_ranges_max_random_50
            l_ranges_uni_random = l_ranges_uni_random_50

        data_anti = []
        data_corr = []
        data_random = []
        # within each excel file, for each data type
        for j in range(len(Data_Types)):
            # read data list
            data_anti_list_start = data_anti_list[0]
            data_anti_list_end = data_anti_list[1]

            data_corr_list_start = data_corr_list[0]
            data_corr_list_end = data_corr_list[1]

            data_random_list_start = data_random_list[0]
            data_random_list_end = data_random_list[1]

            for columns in ws[data_anti_list_start: data_anti_list_end]:
                for cell in columns:
                    data_anti.append(cell.value)

            for columns in ws[data_corr_list_start: data_corr_list_end]:
                for cell in columns:
                    data_corr.append(cell.value)

            for columns in ws[data_random_list_start: data_random_list_end]:
                for cell in columns:
                    data_random.append(cell.value)

            weight_anti = compute_weights(data_anti)
            weight_corr = compute_weights(data_corr)
            weight_random = compute_weights(data_random)

            # for each type, log, log_minus, log_plus, etc
            for jj in range(types.__len__()):
                # read k and l
                type_name = types[jj]
                if jj == 1:
                    print(type_name)
                start = 2 * jj
                end = 2 * jj + 1
                k_anti = []
                for columns in ws[k_ranges_anti[start]: k_ranges_anti[end]]:
                    for cell in columns:
                        k_anti.append(cell.value)

                l_anti_opt = []
                for columns in ws[l_ranges_opt_anti[start]: l_ranges_opt_anti[end]]:
                    for cell in columns:
                        l_anti_opt.append(cell.value)

                l_anti_max = []
                for columns in ws[l_ranges_max_anti[start]: l_ranges_max_anti[end]]:
                    for cell in columns:
                        l_anti_max.append(cell.value)

                l_anti_uni = []
                for columns in ws[l_ranges_uni_anti[start]: l_ranges_uni_anti[end]]:
                    for cell in columns:
                        l_anti_uni.append(cell.value)

                # read data type correlated
                k_corr = []
                for columns in ws[k_ranges_corr[start]: k_ranges_corr[end]]:
                    for cell in columns:
                        k_corr.append(cell.value)

                l_corr_opt = []
                for columns in ws[l_ranges_opt_corr[start]: l_ranges_opt_corr[end]]:
                    for cell in columns:
                        l_corr_opt.append(cell.value)

                l_corr_max = []
                for columns in ws[l_ranges_max_corr[start]: l_ranges_max_corr[end]]:
                    for cell in columns:
                        l_corr_max.append(cell.value)

                l_corr_uni = []
                for columns in ws[l_ranges_uni_corr[start]: l_ranges_uni_corr[end]]:
                    for cell in columns:
                        l_corr_uni.append(cell.value)

                # read data type random
                k_random = []
                for columns in ws[k_ranges_random[start]: k_ranges_random[end]]:
                    for cell in columns:
                        k_random.append(cell.value)

                l_random_opt = []
                for columns in ws[l_ranges_opt_random[start]: l_ranges_opt_random[end]]:
                    for cell in columns:
                        l_random_opt.append(cell.value)

                l_random_max = []
                for columns in ws[l_ranges_max_random[start]: l_ranges_max_random[end]]:
                    for cell in columns:
                        l_random_max.append(cell.value)

                l_random_uni = []
                for columns in ws[l_ranges_uni_random[start]: l_ranges_uni_random[end]]:
                    for cell in columns:
                        l_random_uni.append(cell.value)

                hash_used_anti_opt_cell = hash_used_anti_opt_cells[jj]
                hash_used_corr_opt_cell = hash_used_corr_opt_cells[jj]
                hash_used_random_opt_cell = hash_used_rand_opt_cells[jj]

                hash_used_anti_opt = ws[hash_used_anti_opt_cell].value
                hash_used_corr_opt = ws[hash_used_corr_opt_cell].value
                hash_used_random_opt = ws[hash_used_random_opt_cell].value

                # update LList
                l_anti_opt = post_optimization_opt(collision_probility, weight_anti, total_error, data_anti, k_anti, l_anti_opt,
                                               hash_used_anti_opt, hash_budget_anti)
                l_corr_opt = post_optimization_opt(collision_probility, weight_anti, total_error, data_corr, k_corr, l_corr_opt,
                                               hash_used_corr_opt, hash_budget_corr)
                l_random_opt = post_optimization_opt(collision_probility, weight_anti, total_error, data_random, k_anti, l_random_opt,
                                               hash_used_random_opt, hash_budget_rand)

                hash_used_anti_uni_cell = hash_used_anti_uni_cells[jj]
                hash_used_corr_uni_cell = hash_used_corr_uni_cells[jj]
                hash_used_random_uni_cell = hash_used_rand_uni_cells[jj]

                hash_used_anti_uni = ws[hash_used_anti_uni_cell].value
                hash_used_corr_uni = ws[hash_used_corr_uni_cell].value
                hash_used_random_uni = ws[hash_used_random_uni_cell].value

                l_anti_uni = post_optimization_uni(data_anti, l_anti_uni, hash_used_anti_uni, hash_budget_anti)
                l_corr_uni = post_optimization_uni(data_corr, l_corr_uni, hash_used_corr_uni, hash_budget_corr)
                l_random_uni = post_optimization_uni(data_random, l_random_uni, hash_used_random_uni, hash_budget_rand)

                # write udpate LList back to excel file
                for kk in range(len(l_anti_opt)):
                    cur_cell_anti_opt = column_row_index(l_ranges_opt_anti[start], kk)
                    cur_cell_corr_opt = column_row_index(l_ranges_opt_corr[start], kk)
                    cur_cell_random_opt = column_row_index(l_ranges_opt_random[start], kk)

                    cur_cell_anti_uni = column_row_index(l_ranges_uni_anti[start], kk)
                    cur_cell_corr_uni = column_row_index(l_ranges_uni_corr[start], kk)
                    cur_cell_random_uni = column_row_index(l_ranges_uni_random[start], kk)

                    ws1[cur_cell_anti_opt] = l_anti_opt[kk]
                    ws1[cur_cell_corr_opt] = l_corr_opt[kk]
                    ws1[cur_cell_random_opt] = l_random_opt[kk]

                    ws1[cur_cell_anti_uni] = l_anti_uni[kk]
                    ws1[cur_cell_corr_uni] = l_corr_uni[kk]
                    ws1[cur_cell_random_uni] = l_random_uni[kk]
        wb1.save(excel_file_name)

print("All done")
# values below all read from excel sheet
####################################################################################

# for optimized approach
anti_weight_list = []
corr_weight_list = []
random_weight_list = []

top_m = 25
top_m_cardinality = 275400
hash_budget = top_m_cardinality * 2
hash_used = 534961

data_list = [1519, 2902, 4201, 5435, 6366, 7395, 8147, 8984, 9622, 10307, 11053, 11551, 11982, 12620, 13014, 13552,
             13894, 14338, 14754, 14810, 15288, 15585, 15876, 16060, 16145]
weight_list = compute_weights(data_list)
KList_Log = [11, 12, 13, 13, 13, 13, 13, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]
L_Opt_List = [26, 14, 10, 7, 6, 5, 4, 3, 2, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
L_Uni_List = [15, 8, 5, 4, 3, 3, 3, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 1, 1, 1, 1, 1, 1, 1]
# weight = [0.081600306, 0.079476028, 0.077041773, 0.074457424, 0.069096951, 0.065782109, 0.060756796, 0.057006843,
#           0.052492599, 0.048674672, 0.04538301, 0.041342773, 0.037406774, 0.034323352, 0.030750563, 0.027688552,
#           0.024378937, 0.021398294, 0.018482749, 0.015291434, 0.012672056, 0.009973976, 0.0073666, 0.004810476,
#           0.002344953]

####################################################################################
# import random
#
# flag = False
# hash_used = 524822
# data_list = np.asarray(data_list)
# while hash_used + smallest <= hash_budget:
#     print("still got hash space left")
#     # sort in descending order
#
#     sorted_index = data_list.argsort()[::-1][:len(data_list)]
#
#     temp_pivot_list = []
#     # first find all the allow current hash re-allocation
#     for i in range(len(sorted_index)):
#         temp_pivot_index = sorted_index[i]
#         if hash_used + data_list[temp_pivot_index] <= hash_budget:
#             temp_pivot_list.append(temp_pivot_index)
#     # randomly pick one from those
#     cur_pivot = random.choice(temp_pivot_list)
#     # update L_Uni_List, hash_used
#     L_Uni_List[cur_pivot] = L_Uni_List[cur_pivot] + 1
#     hash_used = hash_used + data_list[cur_pivot]
#
# total_error = 0
# total_hash_used = 0
# for i in range(len(L_Uni_List)):
#     cur_k = KList_Log[i]
#     cur_l = L_Uni_List[i]
#     total_error = total_error + weight_list[i] * math.pow((1 - math.pow(collision_probility, cur_k)), cur_l)
#     total_hash_used = total_hash_used + data_list[i] * cur_l
# print(L_Uni_List)
# print("Updated total error: " + str(total_error))
# print("total hash used: " + str(total_hash_used))
# print("Uniform approach done")
