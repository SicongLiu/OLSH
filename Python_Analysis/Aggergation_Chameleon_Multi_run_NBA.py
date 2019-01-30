import math
import os
from openpyxl import load_workbook


def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def compute_cell_index(column_index_start_, kk_, row_index_):
    column_index_num = column_index_start_ + kk_
    cs = colnum_string(column_index_num)
    cell_index_ = cs + str(row_index_)
    return cell_index_


row_index_start = [5, 15, 25, 35, 45]
column_index_start = 2
column_counts = 45
row_counts = 6

types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]

repeateNBA_run = 5
excel_folders = '/Users/sicongliu/Desktop/Chameleon/Results_075_Redundant_2_NBA_Excel/'

aggregateNBA_file_name_without_opt = excel_folders + 'NBA_all_without_opt.xlsx'
aggregateNBA_hash_hit_file_name_without_opt = excel_folders + 'NBA_all_hash_hits_without_opt.xlsx'

wb_11 = load_workbook(filename=aggregateNBA_file_name_without_opt)
wb_hash_hits_11 = load_workbook(filename=aggregateNBA_hash_hit_file_name_without_opt)

aggregateNBA_file_name_with_opt = excel_folders + 'NBA_all_with_opt.xlsx'
aggregateNBA_hash_hit_file_name_with_opt = excel_folders + 'NBA_all_hash_hits_with_opt.xlsx'

wb_22 = load_workbook(filename=aggregateNBA_file_name_with_opt)
wb_hash_hits_22 = load_workbook(filename=aggregateNBA_hash_hit_file_name_with_opt)

for rr in range(0, repeateNBA_run):
    file_name_without_opt = excel_folders + 'Aggregation_' + 'NBA_without_opt_' + str(rr) + '.xlsx'
    hash_hit_file_name_without_opt = excel_folders + 'Aggregation_' + 'NBA_without_opt_' + str(rr) + '_hash_hits.xlsx'

    wb_1 = load_workbook(filename=file_name_without_opt)
    wb_hash_hits_1 = load_workbook(filename=hash_hit_file_name_without_opt)

    file_name_with_opt = excel_folders + 'Aggregation_' + 'NBA_with_opt_' + str(rr) + '.xlsx'
    hash_hit_file_name_with_opt = excel_folders + 'Aggregation_' + 'NBA_with_opt_' + str(rr) + '_hash_hits.xlsx'

    wb_2 = load_workbook(filename=file_name_with_opt)
    wb_hash_hits_2 = load_workbook(filename=hash_hit_file_name_with_opt)

    # get sheet name from wb_1, same sheet name will be in wb_11, wb_2, wb_22
    wss = wb_1.get_sheet_names()
    for wwss in wss:
        print(wwss)
        ws1 = wb_1.get_sheet_by_name(wwss)
        wsh1 = wb_hash_hits_1.get_sheet_by_name(wwss)
        ws2 = wb_2.get_sheet_by_name(wwss)
        wsh2 = wb_hash_hits_2.get_sheet_by_name(wwss)

        ws11 = wb_11.get_sheet_by_name(wwss)
        wsh11 = wb_hash_hits_11.get_sheet_by_name(wwss)
        ws22 = wb_22.get_sheet_by_name(wwss)
        wsh22 = wb_hash_hits_22.get_sheet_by_name(wwss)

        for ii in range(types.__len__()):
            cell_row_start = row_index_start[ii]

            for jj in range(0, row_counts):
                row_index = cell_row_start + jj
                for kk in range(0, column_counts):
                    cell_index = compute_cell_index(column_index_start, kk, row_index)
                    if ws1[cell_index].value is not None:
                        if rr == 0:
                            ws11[cell_index].value = 0
                            # wsh11[cell_index].value = 0

                            ws22[cell_index].value = 0
                            # wsh22[cell_index].value = 0

                        ws11[cell_index].value = ws11[cell_index].value + ws1[cell_index].value
                        # wsh11[cell_index].value = wsh11[cell_index].value + wsh1[cell_index].value

                        ws22[cell_index].value = ws22[cell_index].value + ws2[cell_index].value
                        # wsh22[cell_index].value = wsh22[cell_index].value + wsh2[cell_index].value

                        if rr == repeateNBA_run - 1:
                            ws11[cell_index].value = ws11[cell_index].value / repeateNBA_run
                            ws22[cell_index].value = ws22[cell_index].value / repeateNBA_run

wb_11.save(aggregateNBA_file_name_without_opt)
# wb_hash_hits_11.save(aggregateNBA_hash_hit_file_name_without_opt)

wb_22.save(aggregateNBA_file_name_with_opt)
# wb_hash_hits_22.save(aggregateNBA_hash_hit_file_name_with_opt)


