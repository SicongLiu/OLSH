import math
import numpy as np
from openpyxl import load_workbook
import random
import re
import string
import os


def separate_string(input_string):
    items = []
    match = re.match(r"([a-z]+)([0-9]+)", input_string, re.I)
    if match:
        items = match.groups()
    return items


def column_row_index(input_string, column_dist):
    items = separate_string(input_string)
    column_index = items[0]
    row_index = int(items[1]) + column_dist
    return column_index + str(row_index)


# for resource re-distribution
def load_weight(weight_file_, bin_count_):
    lines = np.loadtxt(weight_file_)
    assert(lines.__len__() == bin_count_)
    return lines


def compute_collision_prob(dimension_, data_list_):
    prob_list_ = []
    for ii in data_list_:
        theta = pow((2 * math.pi/ii), (1/(dimension_ - 1)))
        collision = 1 - theta / math.pi
        prob_list_.append(0.75)
    return prob_list_


def post_optimization_opt_revised(collision_probilities_, weight_list_, data_list_, K_List_, L_List_, hash_used_, hash_budget_):
    smallest = min(data_list_)
    while hash_used_ + smallest <= hash_budget_:
        delta_error_list = []
        for i in range(len(data_list_)):
            cur_k = K_List_[i]
            cur_l = L_List_[i]
            c_old = weight_list_[i] * math.pow((1 - math.pow(collision_probilities_[i], cur_k)), cur_l)

            # each time increment hash layer by 1
            c_new = weight_list_[i] * math.pow((1 - math.pow(collision_probilities_[i], cur_k)), (cur_l + 1))
            delta_error_list.append((c_old - c_new))

        # sort and check each delta_error
        delta_error_list = np.asarray(delta_error_list)
        # sort in descending order
        sorted_index = delta_error_list.argsort()[::-1][:len(delta_error_list)]
        sorted_pivot = 0
        while sorted_pivot < len(sorted_index):
            cur_index = sorted_index[sorted_pivot]

            # each time increment hash layer by 1
            # temp_hash_used = hash_used + (LList[cur_index] + 1) * data_list[cur_index]
            temp_hash_used = hash_used_ + data_list_[cur_index]
            if temp_hash_used <= hash_budget_:
                L_List_[cur_index] = L_List_[cur_index] + 1
                hash_used_ = hash_used_ + data_list_[cur_index]
                break
            sorted_pivot = sorted_pivot + 1
    total_error = 0
    total_hash_used = 0
    for i in range(len(L_List_)):
        cur_k = K_List_[i]
        cur_l = L_List_[i]
        total_error = total_error + weight_list_[i] * math.pow((1 - math.pow(collision_probilities_[i], cur_k)), cur_l)
        total_hash_used = total_hash_used + data_list_[i] * cur_l
    return L_List_


####################################################################################
# for uniformly distributed approach
# uniformly pick one that fits the current hash budget instead of the one minimizing total error rate
def post_optimization_uni(data_list_, L_List_, hash_used_, hash_budget_):
    flag = False
    smallest = min(data_list_)
    data_list_ = np.asarray(data_list_)
    while hash_used_ + smallest <= hash_budget_:
        # sort in descending order
        sorted_index = data_list_.argsort()[::-1][:len(data_list_)]

        temp_pivot_list = []
        # first find all the allow current hash re-allocation
        for i in range(len(sorted_index)):
            temp_pivot_index = sorted_index[i]
            if hash_used_ + data_list_[temp_pivot_index] <= hash_budget_:
                temp_pivot_list.append(temp_pivot_index)
        # randomly pick one from those
        cur_pivot = random.choice(temp_pivot_list)
        # update L_Uni_List, hash_used
        L_List_[cur_pivot] = L_List_[cur_pivot] + 1
        hash_used_ = hash_used_ + data_list_[cur_pivot]
    return L_List_


def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


# J, AA AQ
def compute_data_list_start_end(data_type_index_, data_list, column_dist_):
    start_items_ = separate_string(data_list[0])
    start_letter_ = start_items_[0]
    start_num_ = start_items_[1]

    end_items_ = separate_string(data_list[1])
    end_letter_ = end_items_[0]
    end_num_ = end_items_[1]

    start_letter_integer_ = int(col2num(start_letter_))
    start_letter_integer_ = start_letter_integer_ + column_dist_ * data_type_index_
    start_letter_ = colnum_string(start_letter_integer_)
    data_list_start = start_letter_ + start_num_
    data_list_end = start_letter_ + end_num_

    return data_list_start, data_list_end


# both letter and integer row index change
def compute_ranges_start_end(column_shift, row_shift, ranges, row_dist_, column_dist_):
    range_start, range_end = compute_data_list_start_end(column_shift, ranges, column_dist_)
    range_column = separate_string(range_start)[0]
    range_row_start = separate_string(range_start)[1] # 6
    range_row_end = separate_string(range_end)[1] # 45
    row_span_ = int(range_row_end) - int(range_row_start)

    target_row_start = int(range_row_end) * row_shift + row_dist_ # 51
    target_row_end = target_row_start + row_span_
    k_ranges_temp_start = range_column + str(target_row_start)
    k_ranges_temp_end = range_column + str(target_row_end)

    return k_ranges_temp_start, k_ranges_temp_end


def compute_hash_cell(l_ranges_opt_temp_end, temp_row_dist, temp_column_dist):
    temp_column = separate_string(l_ranges_opt_temp_end)[0]
    temp_row = int(separate_string(l_ranges_opt_temp_end)[1])
    final_column = colnum_string(col2num(temp_column) + temp_column_dist)
    final_row = temp_row + temp_row_dist
    hash_used_cell = final_column + str(final_row)
    return hash_used_cell


# types_: log, log_minus, etc
def write_script(data_type_, dimension_, bin_count_, top_k_, type_name, budget_, card_excel_, cardinality_, query_count, ratio, with_without_opt_, pot_,  K_List, L_Opt_List, L_Max_List, L_Uni_List, data_list_):
    total_bash_file = SCRIPT_FOLDER + "run_bash_set_cur_" + str(dimension_) + 'D_' + str(card_excel_) + \
                      '_' + str(with_without_opt_) + '.sh'
    f10 = open(total_bash_file, 'a+')
    f10.write("#!/bin/bash \n")

    BASH_FILE_FOLDER = SCRIPT_FOLDER + "bash_set_" + str(dimension_) + "D_top" + str(
        top_k_) + "_budget_" + str(budget_) + "_" + type_name + "_" + str(
        cardinality_) + "/"
    TEMPORAL_RESULT = "../H2_ALSH/temp_result_" + str(dimension_) + "D_top" + str(
        top_k_) + "_budget_" + str(budget_) + "_" + type_name + "_" + str(cardinality_) + "/"
    TEMPORAL_RESULT_FOR_BASH = "./temp_result_" + str(dimension_) + "D_top" + str(
        top_k_) + "_budget_" + str(budget_) + "_" + type_name + "_" + str(cardinality_) + "/"

    if not os.path.exists(BASH_FILE_FOLDER):
        os.makedirs(BASH_FILE_FOLDER)

    if not os.path.exists(TEMPORAL_RESULT):
        os.makedirs(TEMPORAL_RESULT)
    qhull_data_count = []

    for mm in range(len(K_List)):
        # qhull_file = BASE_FOLDER + data_type_ + "_" + str(dimension_) + "_" + str(
        #     cardinality_) \
        #              + "_" + "qhull_layer_" + str(mm)
        # f = open(qhull_file, 'r')
        # lines = f.readlines()
        # second_line = lines[1]
        # cur_data_count = int(second_line.split('\n')[0])
        qhull_data_count.append(data_list_[mm])
        # f.close()

    # opt cumsum_hashsize
    obj_cumsum = []
    hashsize_cumsum = []
    obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_opt_" + data_type_ + "_" + str(
        dimension_) + "_" + str(cardinality_) + '_' + str(
        with_without_opt_) + ".txt"
    f4 = open(obj_hashsize_file, 'w')
    for mm in range(len(L_Opt_List)):
        if obj_cumsum.__len__() == 0:
            obj_cumsum.append(qhull_data_count[mm])
            hashsize_cumsum.append(qhull_data_count[mm] * L_Opt_List[mm])
        else:
            obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
            hashsize_cumsum.append(
                hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] *
                L_Opt_List[mm])
    f4.write(','.join(map(repr, obj_cumsum)))
    f4.write("\n")
    f4.write(','.join(map(repr, hashsize_cumsum)))
    f4.close()

    # max cumsum_hashsize
    obj_cumsum = []
    hashsize_cumsum = []
    obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_max_" + data_type_ + "_" + str(
        dimension_) + "_" + str(cardinality_) + '_' + str(
        with_without_opt_) + ".txt"
    f4 = open(obj_hashsize_file, 'w')
    for mm in range(len(L_Max_List)):
        if obj_cumsum.__len__() == 0:
            obj_cumsum.append(qhull_data_count[mm])
            hashsize_cumsum.append(qhull_data_count[mm] * L_Max_List[mm])
        else:
            obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
            hashsize_cumsum.append(
                hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] *
                L_Max_List[mm])
    f4.write(','.join(map(repr, obj_cumsum)))
    f4.write("\n")
    f4.write(','.join(map(repr, hashsize_cumsum)))
    f4.close()

    # uni cumsum hashsize
    obj_cumsum = []
    hashsize_cumsum = []
    obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_uni_" + data_type_ + "_" + str(
        dimension_) + "_" + str(cardinality_) + '_' + str(
        with_without_opt_) + ".txt"
    f4 = open(obj_hashsize_file, 'w')
    for mm in range(len(L_Uni_List)):
        if obj_cumsum.__len__() == 0:
            obj_cumsum.append(qhull_data_count[mm])
            hashsize_cumsum.append(qhull_data_count[mm] * L_Uni_List[mm])
        else:
            obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
            hashsize_cumsum.append(
                hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] * L_Uni_List[mm])
    f4.write(','.join(map(repr, obj_cumsum)))
    f4.write("\n")
    f4.write(','.join(map(repr, hashsize_cumsum)))
    f4.close()

    if data_type_.__contains__('anti_'):
        ts_data_type_ = 'anti_correlated'
    elif data_type_.__contains__('correlated'):
        ts_data_type_ = 'correlated'
    else:
        ts_data_type_ = 'random'
    sim_threshold = 0.75  # deprecated


    cur_data_type = data_type_
    cur_cardinality = cardinality_
    cur_dimension = dimension_
    # ts_data_type_ = data_type_.split('_')[0]
    f10.write("ts_datatype=" + ts_data_type_ + "\n")
    f10.write("datatype=" + cur_data_type + "\n")
    f10.write("cardinality=" + str(cur_cardinality) + "\n")
    f10.write("d=" + str(cur_dimension) + "\n")
    f10.write("qn=" + str(query_count) + "\n")
    f10.write("c0=" + str(ratio) + "\n")
    f10.write("pot=" + str(pot_) + "\n")

    temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_max"
    overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_max"
    sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_max"
    f10.write("temporalResult=" + temporalResult + "\n")
    f10.write("overallResult=" + overallResult + "\n")
    f10.write("sim_angle=" + sim_angle + "\n")
    f10.write("S=" + str(sim_threshold) + "\n")
    f10.write("num_layer=" + str(len(K_List)) + "\n")
    f10.write("top_k=" + str(top_k_) + "\n")
    f10.write(
        "# ------------------------------------------------------------------------------ \n")
    f10.write("#     Ground-Truth \n")
    f10.write(
        "# ------------------------------------------------------------------------------ \n")
    f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
    f10.write(
        "tsPath=./result/result_${ts_datatype}_${d}D_${cardinality} # path for the ground truth \n")
    f10.write("qPath=./query/query_${d}D.txt \n")
    f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
    f10.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
              "${oFolder}.mip \n")
    f10.write("\n \n \n")
    f10.write(
        "# ------------------------------------------------------------------------------ \n")
    f10.write("#     Layer-Performance \n")
    f10.write(
        "# ------------------------------------------------------------------------------ \n")
    for kk in range(len(K_List)):
        f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
        f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
        f10.write("L" + str(kk) + "=" + str(L_Max_List[kk]) + "\n")

        f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                      "${cardinality}_qhull_layer_" + str(kk) + "\n")

        f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                        "${cardinality}_max/result_${d}D" + str(
            kk) + "_${K" + str(kk)
                  + "}_${L" + str(kk) + "}" + "\n")
        f10.write("temp_hash" + str(
            kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_max_" + str(kk) + "\n")

        f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                  "} -L ${L" + str(kk) + "} -LI " + str(
            kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                  + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
            kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
        f10.write("\n")

    # append overall accuracy computation here
    f10.write(
        "# ------------------------------------------------------------------------------ \n")
    f10.write("#     Overall-Performance \n")
    f10.write(
        "# ------------------------------------------------------------------------------ \n")
    f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
              "${tsPath}.mip -of ${overallResult} \n \n \n")
    f10.write("sleep 2 \n")

    if data_type_.__contains__('prob'):
        return 0

    cur_data_type = data_type_
    cur_cardinality = cardinality_
    cur_dimension = dimension_
    # ts_data_type_ = data_type_.split('_')[0]
    f10.write("ts_datatype=" + ts_data_type_ + "\n")
    f10.write("datatype=" + cur_data_type + "\n")
    f10.write("cardinality=" + str(cur_cardinality) + "\n")
    f10.write("d=" + str(cur_dimension) + "\n")
    f10.write("qn=" + str(query_count) + "\n")
    f10.write("c0=" + str(ratio) + "\n")
    f10.write("pot=" + str(pot_) + "\n")

    temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_opt"
    overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_opt"
    sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_opt"
    f10.write("temporalResult=" + temporalResult + "\n")
    f10.write("overallResult=" + overallResult + "\n")
    f10.write("sim_angle=" + sim_angle + "\n")

    f10.write("S=" + str(sim_threshold) + "\n")
    f10.write("num_layer=" + str(len(K_List)) + "\n")
    f10.write("top_k=" + str(top_k_) + "\n")
    f10.write("# ------------------------------------------------------------------------------ \n")
    f10.write("#     Ground-Truth \n")
    f10.write("# ------------------------------------------------------------------------------ \n")
    f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
    f10.write(
        "tsPath=./result/result_${ts_datatype}_${d}D_${cardinality} # path for the ground truth \n")
    f10.write("qPath=./query/query_${d}D.txt \n")
    f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")

    f10.write(" # ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
              "${oFolder}.mip \n")
    f10.write(" # sleep 1 \n")

    f10.write("\n \n \n")
    f10.write("# ------------------------------------------------------------------------------ \n")
    f10.write("#     Layer-Performance \n")
    f10.write("# ------------------------------------------------------------------------------ \n")
    for kk in range(len(K_List)):
        f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
        f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
        f10.write("L" + str(kk) + "=" + str(L_Opt_List[kk]) + "\n")

        f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                      "${cardinality}_qhull_layer_" + str(kk) + "\n")

        f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                        "${cardinality}_opt/result_${d}D" + str(kk) + "_${K" + str(kk)
                  + "}_${L" + str(kk) + "}" + "\n")
        f10.write("temp_hash" + str(
            kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_opt_" + str(kk) + "\n")

        f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                  "} -L ${L" + str(kk) + "} -LI " + str(
            kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                  + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
            kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
        f10.write("\n")

    # append overall accuracy computation here
    f10.write("# ------------------------------------------------------------------------------ \n")
    f10.write("#     Overall-Performance \n")
    f10.write("# ------------------------------------------------------------------------------ \n")
    f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
              "${tsPath}.mip -of ${overallResult} \n \n \n")
    f10.write("sleep 2 \n")



    cur_data_type = data_type_
    cur_cardinality = cardinality_
    cur_dimension = dimension_
    # ts_data_type_ = data_type_.split('_')[0]
    f10.write("ts_datatype=" + ts_data_type_ + "\n")
    f10.write("datatype=" + cur_data_type + "\n")
    f10.write("cardinality=" + str(cur_cardinality) + "\n")
    f10.write("d=" + str(cur_dimension) + "\n")
    f10.write("qn=" + str(query_count) + "\n")
    f10.write("c0=" + str(ratio) + "\n")
    f10.write("pot=" + str(pot_) + "\n")

    temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_uni"
    overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_uni"
    sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_uni"
    f10.write("temporalResult=" + temporalResult + "\n")
    f10.write("overallResult=" + overallResult + "\n")
    f10.write("sim_angle=" + sim_angle + "\n")
    f10.write("S=" + str(sim_threshold) + "\n")
    f10.write("num_layer=" + str(len(K_List)) + "\n")
    f10.write("top_k=" + str(top_k_) + "\n")
    f10.write("# ------------------------------------------------------------------------------ \n")
    f10.write("#     Ground-Truth \n")
    f10.write("# ------------------------------------------------------------------------------ \n")
    f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
    f10.write(
        "tsPath=./result/result_${ts_datatype}_${d}D_${cardinality} # path for the ground truth \n")
    f10.write("qPath=./query/query_${d}D.txt \n")
    f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
    f10.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
              "${oFolder}.mip \n")
    f10.write("\n \n \n")
    f10.write("# ------------------------------------------------------------------------------ \n")
    f10.write("#     Layer-Performance \n")
    f10.write("# ------------------------------------------------------------------------------ \n")
    for kk in range(len(K_List)):
        f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
        f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
        f10.write("L" + str(kk) + "=" + str(L_Uni_List[kk]) + "\n")

        f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                      "${cardinality}_qhull_layer_" + str(kk) + "\n")

        f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                        "${cardinality}_uni/result_${d}D" + str(kk) + "_${K" + str(kk)
                  + "}_${L" + str(kk) + "}" + "\n")
        f10.write("temp_hash" + str(kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_uni_" + str(
            kk) + "\n")

        f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                  "} -L ${L" + str(kk) + "} -LI " + str(
            kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                  + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
            kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
        f10.write("\n")
    f10.write("# ------------------------------------------------------------------------------ \n")
    f10.write("#     Overall-Performance \n")
    f10.write("# ------------------------------------------------------------------------------ \n")
    f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
              "${tsPath}.mip -of ${overallResult} \n \n \n")
    f10.write("sleep 2 \n")
    f10.close()
    return 0
    # return file_names


####################################################################################
##########################code below working on post-opt resource allocation########
Data_Types = ['anti_correlated', 'correlated', 'random']
Data_Gen_Types = ['EW', 'ED_card', 'ED_prob']
types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]

# Data_Types = ['random']
# Data_Gen_Types = ['EW']
# types = ["log"]

# bin_count = [40, 60, 80]
# top_ks = 25

ratio = 2
red_cell = 'I3'
bin_num = 'E1'
top_k_cell = 'E2'
budget_cell = 'B4'
cardinality_cell = 'B2'
card_excel_cell = 'I2'
is_real_life_cell = 'E3'
column_dist = 17
row_dist = 6
dimensions = [50]
query_count = 1000
excel_file_dir = './'

data_list_40 = ['J6',  'J45', 'J51', 'J90']
k_ranges_40 = ['E6',  'E45', 'E51', 'E90']
l_ranges_opt_40 = ['F6', 'F45']
l_ranges_max_40 = ['G6', 'G45']
l_ranges_uni_40 = ['H6', 'H45']
hash_used_opt_cells_40 = ['I46', 'I45']
hash_used_uni_cells_40 = ['O46', 'O45']

data_list_60 = ['J6', 'J30', 'J38', 'J62', 'J69', 'J93', 'J100', 'J124', 'J131', 'J155']
k_ranges_60 = ['E6', 'E30', 'E38', 'E62', 'E69', 'E93', 'E100', 'E124', 'E131', 'E155']
l_ranges_opt_60 = ['F6', 'F30', 'F38', 'F62', 'F69', 'F93', 'F100', 'F124', 'F131', 'F155']
l_ranges_max_60 = ['G6', 'G30', 'G38', 'G62', 'G69', 'G93', 'G100', 'G124', 'G131', 'G155']
l_ranges_uni_60 = ['H6', 'H30', 'H38', 'H62', 'H69', 'H93', 'H100', 'H124', 'H131', 'H155']
hash_used_opt_cells_60 = ['I31', 'I63', 'I94', 'I125', 'I156']
hash_used_uni_cells_60 = ['O31', 'O63', 'O94', 'O125', 'O156']


data_list_80 = ['J6', 'J55', 'J63', 'J112', 'J120', 'J169', 'J177', 'J226', 'J234', 'J283']
k_ranges_80 = ['E6', 'E55', 'E63', 'E112', 'E120', 'E169', 'E177', 'E226', 'E234', 'E283']
l_ranges_opt_80 = ['F6', 'F55', 'F63', 'F112', 'F120', 'F169', 'F177', 'F226', 'F234', 'F283']
l_ranges_max_80 = ['G6', 'G55', 'G63', 'G112', 'G120', 'G169', 'G177', 'G226', 'G234', 'G283']
l_ranges_uni_80 = ['H6', 'H55', 'H63', 'H112', 'H120', 'H169', 'H177', 'H226', 'H234', 'H283']
hash_used_opt_cells_80 = ['I56', 'I113', 'I170', 'I227', 'I284']
hash_used_uni_cells_80 = ['O56', 'O113', 'O170', 'O227', 'O284']

SCRIPT_FOLDER = "../H2_ALSH/"
SCRIPT_OUTPUT_FILE = "../H2_ALSH/parameters/Mathematica_norm_bin_partition_Parameters_"
QHULL_OUTPUT_FOLDER = '/Users/sicongliu/Desktop/StreamingTopK/H2_ALSH/qhull_data/Synthetic/'
DATA_FOLDER = '/Users/sicongliu/Desktop/StreamingTopK/H2_ALSH/raw_data/Synthetic/'
QUERY_FOLDER = '/Users/sicongliu/Desktop/StreamingTopK/H2_ALSH/query/'
GROUNDTRUEH_FOLDER = "../H2_ALSH/result/"
BASE_FOLDER = "../H2_ALSH/qhull_data/Synthetic/"

repeated_run = 5
####################################################################################


bin_count = -1
top_ks = -1
hash_budget = -1
total_cardinality = -1
total_card_excel = -1
is_real_life = -1


# for each excel file
for i in range(len(dimensions)):
    cur_d = dimensions[i]
    excel_file_name = excel_file_dir + str(cur_d) + 'D_1M_after.xlsx'
    wb = load_workbook(filename=excel_file_name, data_only=True)
    wb1 = load_workbook(filename=excel_file_name)
    # wss = wb.get_sheet_names()
    wss = wb.sheetnames

    # no need to loop data_gen_type for now, iterate through all excel sheets for now
    for wwss in wss:
        print(wwss)
        # ws = wb.get_sheet_by_name(wwss)
        # ws1 = wb1.get_sheet_by_name(wwss)
        ws = wb[wwss]
        ws1 = wb1[wwss]
        bin_count = ws[bin_num].value
        top_ks = ws[top_k_cell].value
        hash_budget = ws[budget_cell].value
        total_cardinality = ws[cardinality_cell].value
        total_card_excel = ws[card_excel_cell].value
        is_real_life = ws[is_real_life_cell].value
        budget_factor = ws[red_cell].value
        if bin_count == 40:
            hash_used_opt_cells = hash_used_opt_cells_40
            hash_used_uni_cells = hash_used_uni_cells_40
            data_list = data_list_40
            k_ranges = k_ranges_40
            l_ranges_opt = l_ranges_opt_40
            l_ranges_max = l_ranges_max_40
            l_ranges_uni = l_ranges_uni_40
        elif bin_count == 60:
            hash_used_opt_cells = hash_used_opt_cells_60
            hash_used_uni_cells = hash_used_uni_cells_60
            data_list = data_list_60
            k_ranges = k_ranges_60
            l_ranges_opt = l_ranges_opt_60
            l_ranges_max = l_ranges_max_60
            l_ranges_uni = l_ranges_uni_60
        else:# 80-bins
            hash_used_opt_cells = hash_used_opt_cells_80
            hash_used_uni_cells = hash_used_uni_cells_80
            data_list = data_list_80
            k_ranges = k_ranges_80
            l_ranges_opt = l_ranges_opt_80
            l_ranges_max = l_ranges_max_80
            l_ranges_uni = l_ranges_uni_80

        # Data_Types = ['anti_correlated', 'correlated', 'random']
        for kk in range(Data_Types.__len__()):
            data_ = []
            # each round udpate data_list_start
            # hash_used_opt_cells
            # hash_used_uni_cells
            data_list_start = data_list[0]
            data_list_end = data_list[1]

            data_list_start, data_list_end = compute_data_list_start_end(kk, data_list, column_dist)
            for columns in ws[data_list_start: data_list_end]:
                for cell in columns:
                    data_.append(cell.value)

            # modify parameter for load_weight, check current sheet name
            if str(wwss).__contains__('EW'):
                data_type_file_name_ = './' + Data_Types[kk] + '_EW_' + str(cur_d) + '_' + str(total_cardinality) + '_' + str(
                    bin_count) + '_' + 'top_' + str(top_ks) + '_EW.txt'
                data_gen_type = '_EW'
            elif str(wwss).__contains__('ED_card'):
                data_type_file_name_ = './' + Data_Types[kk] + '_ED_card_' + str(cur_d) + '_' + str(total_cardinality) + '_' + str(
                    bin_count) + '_' + 'top_' + str(top_ks) + '_ED_card.txt'
                data_gen_type = '_ED_card'
            else:
                temp_weight_ = np.ones(bin_count)# set default
                data_gen_type = '_ED_prob'

            print(data_type_file_name_)
            if data_gen_type == '_ED_prob':
                print('it is _ED_prob  \n')
            else:
                temp_weight_ = load_weight(data_type_file_name_, bin_count)

            # types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
            for jj in range(types.__len__()):
                # read k and l
                type_name = types[jj]
                print("type_name: " + str(type_name))
                k_ = []
                k_ranges_temp_start, k_ranges_temp_end = compute_ranges_start_end(kk, jj, k_ranges, row_dist, column_dist)
                for columns in ws[k_ranges_temp_start: k_ranges_temp_end]:
                    for cell in columns:
                        k_.append(cell.value)

                l_opt = []
                l_ranges_opt_temp_start, l_ranges_opt_temp_end = compute_ranges_start_end(kk, jj, l_ranges_opt, row_dist, column_dist)
                for columns in ws[l_ranges_opt_temp_start: l_ranges_opt_temp_end]:
                    for cell in columns:
                        l_opt.append(cell.value)

                l_max = []

                l_ranges_max_temp_start, l_ranges_max_temp_end = compute_ranges_start_end(kk, jj, l_ranges_max, row_dist, column_dist)
                for columns in ws[l_ranges_max_temp_start: l_ranges_max_temp_end]:
                    for cell in columns:
                        l_max.append(cell.value)

                l_uni = []
                l_ranges_uni_temp_start, l_ranges_uni_temp_end = compute_ranges_start_end(kk, jj, l_ranges_uni, row_dist, column_dist)
                for columns in ws[l_ranges_uni_temp_start: l_ranges_uni_temp_end]:
                    for cell in columns:
                        l_uni.append(cell.value)

                # update hash_used_opt_cell
                # row same as l_ranges_opt_temp_end row + 1
                # column as l_ranges_opt_temp_end column + 3
                temp_row_dist = 1
                temp_column_dist = 3
                hash_used_opt_cell = compute_hash_cell(l_ranges_opt_temp_end, temp_row_dist, temp_column_dist)
                # hash_used_opt_cell = hash_used_opt_cells[jj]
                hash_used_opt = ws[hash_used_opt_cell].value

                # right now it's 0.75
                collision_probilities = compute_collision_prob(cur_d, data_)
                l_opt_revised = l_opt
                l_opt_revised = post_optimization_opt_revised(collision_probilities, temp_weight_, data_, k_, l_opt_revised, hash_used_opt, hash_budget)

                # row same as l_ranges_opt_temp_end row + 1
                # column as l_ranges_opt_temp_end column + 9
                temp_row_dist = 1
                temp_column_dist = 9
                hash_used_uni_cell = compute_hash_cell(l_ranges_opt_temp_end, temp_row_dist, temp_column_dist)
                # hash_used_uni_cell = hash_used_uni_cells[jj]
                hash_used_uni = ws[hash_used_uni_cell].value
                l_uni_revised = l_uni
                l_uni_revised = post_optimization_uni(data_, l_uni_revised, hash_used_uni, hash_budget)

                # write udpate LList back to excel file
                for ll in range(len(l_opt_revised)):
                    cur_cell_opt = column_row_index(l_ranges_opt_temp_start, ll)
                    cur_cell_uni = column_row_index(l_ranges_uni_temp_start, ll)
                    ws1[cur_cell_opt] = l_opt_revised[ll]
                    ws1[cur_cell_uni] = l_uni_revised[ll]

                # write to script here
                with_without_opt = 'without_opt'
                pot = 0
                write_script(Data_Types[kk] + data_gen_type, cur_d, bin_count, top_ks, type_name, hash_budget, total_card_excel, total_cardinality, query_count, ratio, with_without_opt, pot, k_, l_opt_revised, l_max, l_uni_revised, data_)

                # with_without_opt = 'with_opt'
                # pot = 1
                # write_script(Data_Types[kk] + data_gen_type, cur_d, bin_count, top_ks, type_name, hash_budget,
                #                                       total_card_excel, total_cardinality, query_count, ratio,
                #                                       with_without_opt, pot, k_, l_opt, l_max, l_uni, data_)

                # aggregated_file_name = SCRIPT_FOLDER + "run_bash_" + str(dimensions[0]) + "D_all.sh"
                # f1 = open(aggregated_file_name, 'a+')
                # f1.write("#!/bin/bash \n")
                # for rr in range(0, repeated_run):
                #     for files in file_names_without_opt:
                #         f1.write('sh ' + files + '\n')
                #
                #     temp_str = "aggregating for non-opt round " + str(rr)
                #     f1.write('echo \"' + temp_str + '\" \n')
                #
                #     # python ../Python_Analysis/LSH_Post_Process_300D.py without_opt 1
                #     # python ../Python_Analysis/Clean_All_300D.py
                #
                #     # with_without_opt = str(sys.argv[1])
                #     # run_index = str(sys.argv[2])
                #     # cur_dimension = int(str(sys.argv[2]))
                #     # cur_budget = int(str(sys.argv[3]))
                #     # cur_bin_count = int(str(sys.argv[4]))
                #     # cur_top_o = int(str(sys.argv[5]))
                #     # equal_type = str(sys.argv[6])
                #
                #     f1.write(
                #         'python ../Python_Analysis/Norm_Analysis_LSH_Agg_Result_To_Excel.py without_opt ' + str(rr) + ' ' + str(cur_d) + ' ' + str(hash_budget) + ' ' + str(bin_count) + ' ' + str(top_ks) + ' ' + str(data_gen_type[1:data_gen_type.__len__()]) + ' ' + str(is_real_life) + ' \n')
                #     f1.write('sleep 3' + '\n')
                #
                #     # before starting pot = 1, clean everything except hash_table
                #     f1.write('python ../Python_Analysis/Norm_Analysis_Sim_Overall_run_clean.py ' + str(cur_d) + ' \n')
                #     f1.write('sleep 5' + '\n')
                #
                #     for files in file_names_with_opt:
                #         f1.write('sh ' + files + '\n')
                #
                #     temp_str = "aggregating for opt round " + str(rr)
                #     f1.write('echo \"' + temp_str + '\" \n')
                #
                #     f1.write(
                #         'python ../Python_Analysis/Norm_Analysis_LSH_Result_To_Excel.py with_opt ' + str(
                #             rr) + ' ' + str(cur_d) + ' ' + str(hash_budget) + ' ' + str(bin_count) + ' ' + str(
                #             top_ks) + ' ' + str(data_gen_type[1:data_gen_type.__len__()]) + ' ' + str(is_real_life) + ' \n')
                #     f1.write('sleep 3' + '\n')
                #
                #     # f1.write('python ../Python_Analysis/LSH_Post_Process_' + str(cur_d) + 'D.py with_opt ' + str(
                #     #         rr) + ' \n')
                #     # f1.write('sleep 3' + '\n')
                #
                #     # before starting pot = 1, clean everything except hash_table
                #     f1.write('python ../Python_Analysis/Norm_Analysis_Clean_All.py ' + str(cur_d) + ' \n')
                #     f1.write('sleep 5' + '\n')
                #
                # f1.close()
    # wb.save(excel_file_name)
    wb1.save(excel_file_name)

    aggregated_file_name = SCRIPT_FOLDER + "run_bash_" + str(cur_d) + "D_all.sh"
    f1 = open(aggregated_file_name, 'a+')
    f1.write("#!/bin/bash \n")

    for rr in range(0, repeated_run):
        file_name = SCRIPT_FOLDER + "run_bash_set_cur_" + str(cur_d) + 'D_' + str(total_card_excel) + \
                    '_' + str('without_opt') + '.sh'

        f1.write('sh ' + file_name + '\n')
        temp_str = "aggregating for non-opt round " + str(rr)
        f1.write('echo \"' + temp_str + '\" \n')

        f1.write('source ../Python_Analysis/venv/bin/activate \n')
        f1.write(
            'python3 ../Python_Analysis/Norm_Analysis_LSH_Agg_Result_To_Excel.py without_opt ' + str(rr) + ' ' + str(
                cur_d) + ' ' + str(hash_budget) + ' ' + str(bin_count) + ' ' + str(top_ks) + ' ' + str(
                total_cardinality) + ' ' + str(is_real_life) + ' \n')
        f1.write('sleep 3' + '\n')

        # before starting pot = 1, clean everything except hash_table
        f1.write('python3 ../Python_Analysis/Norm_Analysis_Sim_Overall_run_clean.py ' + str(cur_d) + ' ' +
                 str(total_cardinality) + ' ' + str(budget_factor) + ' ' + str(top_ks) + ' \n')
        f1.write('sleep 5' + '\n')


        # before starting next round, clean all
        f1.write('python3 ../Python_Analysis/Norm_Analysis_Clean_All.py ' + str(cur_d) + ' ' +
                 str(total_cardinality) + ' ' + str(budget_factor) + ' ' + str(top_ks) + ' \n')
        f1.write('sleep 5' + '\n')



        # for data_gen_type in Data_Gen_Types:
        # f1.write('source ../Python_Analysis/venv/bin/activate')
        # f1.write(
        #     'python ../Python_Analysis/Norm_Analysis_LSH_Agg_Result_To_Excel.py without_opt ' + str(rr) + ' ' + str(
        #         cur_d) + ' ' + str(hash_budget) + ' ' + str(bin_count) + ' ' + str(top_ks) + ' ' + str(
        #         data_gen_type) + ' ' + str(is_real_life) + ' \n')
        # f1.write('sleep 3' + '\n')



        # with_without_opt = str(sys.argv[1])
        # run_index = str(sys.argv[2])
        # cur_dimension = int(str(sys.argv[3]))
        # cur_budget = int(str(sys.argv[4]))
        # cur_bin_count = int(str(sys.argv[5]))
        # cur_top_o = int(str(sys.argv[6]))
        # # equal_type = str(sys.argv[7])
        # cur_card = str(sys.argv[7])
        # is_real_life = str(sys.argv[8])

    f1.close()
    # wb.save(excel_file_name)

print("All done")