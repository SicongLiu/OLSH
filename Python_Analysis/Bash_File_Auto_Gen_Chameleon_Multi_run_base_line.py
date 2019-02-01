# script to generate .sh file for execution
# bash file includes execution of Ground Truth, LSH Scheme and Overall Performance

import math
import os
from openpyxl import load_workbook

k_ranges_anti_50 = ['E6',  'E6', 'E14', 'E14', 'E21', 'E21', 'E28', 'E28', 'E35', 'E35']
l_ranges_opt_anti_50 = ['F6', 'F6', 'F14', 'F14', 'F21', 'F21', 'F28', 'F28', 'F35', 'F35']
l_ranges_max_anti_50 = ['G6', 'G6', 'G14', 'G14', 'G21', 'G21', 'G28', 'G28', 'G35', 'G35']
l_ranges_uni_anti_50 = ['H6', 'H6', 'H14', 'H14', 'H21', 'H21', 'H28', 'H28', 'H35', 'H35']

k_ranges_corr_50 = ['V6', 'V6', 'V14', 'V14', 'V21', 'V21', 'V28', 'V28', 'V35', 'V35']
l_ranges_opt_corr_50 = ['W6', 'W6', 'W14', 'W14', 'W21', 'W21', 'W28', 'W28', 'W35', 'W35']
l_ranges_max_corr_50 = ['X6', 'X6', 'X14', 'X14', 'X21', 'X21', 'X28', 'X28', 'X35', 'X35']
l_ranges_uni_corr_50 = ['Y6', 'Y6', 'Y14', 'Y14', 'Y21', 'Y21', 'Y28', 'Y28', 'Y35', 'Y35']

k_ranges_random_50 = ['AL6', 'AL6', 'AL14', 'AL14', 'AL21', 'AL21', 'AL28', 'AL28', 'AL35', 'AL35']
l_ranges_opt_random_50 = ['AM6', 'AM6', 'AM14', 'AM14', 'AM21', 'AM21', 'AM28', 'AM28', 'AM35', 'AM35']
l_ranges_max_random_50 = ['AN6', 'AN6', 'AN14', 'AN14', 'AN21', 'AN21', 'AN28', 'AN28', 'AN35', 'AN35']
l_ranges_uni_random_50 = ['AO6', 'AO6', 'AO14', 'AO14', 'AO21', 'AO21', 'AO28', 'AO28', 'AO35', 'AO35']


BASE_FOLDER = "../H2_ALSH/qhull_data/Synthetic/"
BASH_FILE_BASE_FOLDER = "../H2_ALSH/"


def separate_string(dimension_, excel_file_, top_k_, types_, card_excel_, save_file_path_):
    wb = load_workbook(filename=excel_file_, data_only=True)
    k_ranges_anti = k_ranges_anti_50
    l_ranges_opt_anti = l_ranges_opt_anti_50
    l_ranges_max_anti = l_ranges_max_anti_50
    l_ranges_uni_anti = l_ranges_uni_anti_50

    k_ranges_corr = k_ranges_corr_50
    l_ranges_opt_corr = l_ranges_opt_corr_50
    l_ranges_max_corr = l_ranges_max_corr_50
    l_ranges_uni_corr = l_ranges_uni_corr_50

    k_ranges_random = k_ranges_random_50
    l_ranges_opt_random = l_ranges_opt_random_50
    l_ranges_max_random = l_ranges_max_random_50
    l_ranges_uni_random = l_ranges_uni_random_50

    sheetname = str(dimension_) + "D_top" + str(top_k_) + "_" + card_excel_
    if sheetname in wb.sheetnames:
        print("Sheetname: " + str(sheetname))
        ws = wb[sheetname]
        for m in range(types_.__len__()):
            type_name = types[m]
            start = 2 * m
            end = 2 * m + 1
            # read data type anti_correlated
            k_anti = []
            for row in ws[k_ranges_anti[start]: k_ranges_anti[end]]:
                for cell in row:
                    k_anti.append(cell.value)

            l_anti_opt = []
            for row in ws[l_ranges_opt_anti[start]: l_ranges_opt_anti[end]]:
                for cell in row:
                    l_anti_opt.append(cell.value)

            l_anti_max = []
            for row in ws[l_ranges_max_anti[start]: l_ranges_max_anti[end]]:
                for cell in row:
                    l_anti_max.append(cell.value)

            l_anti_uni = []
            for row in ws[l_ranges_uni_anti[start]: l_ranges_uni_anti[end]]:
                for cell in row:
                    l_anti_uni.append(cell.value)

            # read data type correlated
            k_corr = []
            for row in ws[k_ranges_corr[start]: k_ranges_corr[end]]:
                for cell in row:
                    k_corr.append(cell.value)

            l_corr_opt = []
            for row in ws[l_ranges_opt_corr[start]: l_ranges_opt_corr[end]]:
                for cell in row:
                    l_corr_opt.append(cell.value)

            l_corr_max = []
            for row in ws[l_ranges_max_corr[start]: l_ranges_max_corr[end]]:
                for cell in row:
                    l_corr_max.append(cell.value)

            l_corr_uni = []
            for row in ws[l_ranges_uni_corr[start]: l_ranges_uni_corr[end]]:
                for cell in row:
                    l_corr_uni.append(cell.value)

            # read data type random
            k_random = []
            for row in ws[k_ranges_random[start]: k_ranges_random[end]]:
                for cell in row:
                    k_random.append(cell.value)

            l_random_opt = []
            for row in ws[l_ranges_opt_random[start]: l_ranges_opt_random[end]]:
                for cell in row:
                    l_random_opt.append(cell.value)

            l_random_max = []
            for row in ws[l_ranges_max_random[start]: l_ranges_max_random[end]]:
                for cell in row:
                    l_random_max.append(cell.value)

            l_random_uni = []
            for row in ws[l_ranges_uni_random[start]: l_ranges_uni_random[end]]:
                for cell in row:
                    l_random_uni.append(cell.value)
            # save current K and L parameters to files
            # anti_opt
            # anti_uni
            save_file_dir = save_file_path_ + str(dimension_) + "D_top" + str(top_k_) + "_" + type_name + "_" + \
                            card_excel_ + "/"

            if not os.path.exists(save_file_dir):
                os.makedirs(save_file_dir)
            anti_k_name = save_file_dir + "k_anti_correlated"
            f = open(anti_k_name, 'w')
            f.write(','.join(map(str, k_anti)))
            f.close()

            anti_opt_name = save_file_dir + "l_anti_correlated_opt"
            f = open(anti_opt_name, 'w')
            f.write(','.join(map(str, l_anti_opt)))
            f.close()

            anti_max_name = save_file_dir + "l_anti_correlated_max"
            f = open(anti_max_name, 'w')
            f.write(','.join(map(str, l_anti_max)))
            f.close()

            anti_uni_name = save_file_dir + "l_anti_correlated_uni"
            f = open(anti_uni_name, 'w')
            f.write(','.join(map(str, l_anti_uni)))
            f.close()

            # corr_opt
            # corr_uni
            corr_k_name = save_file_dir + "k_correlated"
            f = open(corr_k_name, 'w')
            f.write(','.join(map(str, k_corr)))
            f.close()

            corr_opt_name = save_file_dir + "l_correlated_opt"
            f = open(corr_opt_name, 'w')
            f.write(','.join(map(str, l_corr_opt)))
            f.close()

            corr_max_name = save_file_dir + "l_correlated_max"
            f = open(corr_max_name, 'w')
            f.write(','.join(map(str, l_corr_max)))
            f.close()

            corr_uni_name = save_file_dir + "l_correlated_uni"
            f = open(corr_uni_name, 'w')
            f.write(','.join(map(str, l_corr_uni)))
            f.close()

            # random_opt
            # random_uni
            random_k_name = save_file_dir + "k_random"
            f = open(random_k_name, 'w')
            f.write(','.join(map(str, k_random)))
            f.close()

            random_opt_name = save_file_dir + "l_random_opt"
            f = open(random_opt_name, 'w')
            f.write(','.join(map(str, l_random_opt)))
            f.close()

            random_max_name = save_file_dir + "l_random_max"
            f = open(random_max_name, 'w')
            f.write(','.join(map(str, l_random_max)))
            f.close()

            random_uni_name = save_file_dir + "l_random_uni"
            f = open(random_uni_name, 'w')
            f.write(','.join(map(str, l_random_uni)))
            f.close()
    wb.close()
    print("Separate String and Save Parameter Files Done .\n")


def write_script(data_type_, dimension_, top_k_, types_, card_excel_, cardinality_, card_of_interest_, parameter_path_, with_without_opt_, pot_):
    file_names = []
    for ii in range(data_type_.__len__()):
        total_bash_file = BASH_FILE_BASE_FOLDER + "run_bash_set_cur_" + str(dimension_) + 'D_' + cardinality_[ii] + \
                          '_' + str(with_without_opt_) + '.sh'
        file_names.append(total_bash_file)
        f10 = open(total_bash_file, 'w')
        f10.write("#!/bin/bash \n")
        for m in range(types_.__len__()):
            type_name = types_[m]
            parameter_dir = parameter_path_ + str(dimension) + "D_top" + str(top_k_) + "_" + type_name + "_" + card_excel_ + "/"
            print(parameter_dir + ' ' + type_name)
            if not os.path.exists(parameter_dir):
                continue
            else:
                BASH_FILE_FOLDER = BASH_FILE_BASE_FOLDER + "bash_set_" + str(dimension) + "D_top" + str(
                    top_k_) + "_" + type_name + "_" + str(card_excel_) + "/"

                TEMPORAL_RESULT = "../H2_ALSH/temp_result_" + str(dimension) + "D_top" + str(top_k_) + "_" + type_name + \
                                  "_" + str(card_excel_) + "/"
                TEMPORAL_RESULT_FOR_BASH = "./temp_result_" + str(dimension) + "D_top" + str(top_k_) +  "_" + \
                                           type_name + "_" + str(card_excel_) + "/"

                if not os.path.exists(BASH_FILE_FOLDER):
                    os.makedirs(BASH_FILE_FOLDER)

                if not os.path.exists(TEMPORAL_RESULT):
                    os.makedirs(TEMPORAL_RESULT)

                cur_bash_set_dir = "bash_set_" + str(dimension) + "D_top" + str(top_k_) + "_" + type_name + "_" + str(cardinality_[ii]) + "/"

                K_List = []
                L_Opt_List = []
                L_Max_List = []
                L_Uni_List = []
                qhull_data_count = []

                paramK_path = parameter_dir + "k_" + data_type_[ii]
                f1 = open(paramK_path, 'r')
                K_lines = f1.readlines()
                actual_top_k = min(top_k_,  K_lines[0].split(',').__len__())
                for k_index in range(actual_top_k):
                    K_List.append(int(K_lines[0].split(',')[k_index]))
                f1.close()

                paramL_opt_path = parameter_dir + "l_" + data_type_[ii] + "_opt"
                f2 = open(paramL_opt_path, 'r')
                L_lines = f2.readlines()
                for l_index in range(0, actual_top_k):
                    L_Opt_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                f2.close()

                paramL_max_path = parameter_dir + "l_" + data_type_[ii] + "_max"
                f2 = open(paramL_max_path, 'r')
                L_lines = f2.readlines()
                for l_index in range(0, actual_top_k):
                    L_Max_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                f2.close()

                paramL_uni_path = parameter_dir + "l_" + data_type_[ii] + "_uni"
                f2 = open(paramL_uni_path, 'r')
                L_lines = f2.readlines()

                for l_index in range(0, actual_top_k):
                    L_Uni_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                f2.close()

                for mm in range(len(K_List)):
                    # qhull_file = BASE_FOLDER + data_type_[ii] + "_" + str(dimension_) + "_" + str(cardinality[ii])\
                    #              + "_" + "qhull_layer_" + str(mm)
                    qhull_file = data_type_[ii] + "_" + str(dimension_) + "_" + str(cardinality_[ii]) + '.txt'
                    f = open(qhull_file, 'r')
                    lines = f.readlines()
                    second_line = lines[1]
                    cur_data_count = int(second_line.split('\n')[0])
                    qhull_data_count.append(cur_data_count)
                    f.close()

                # opt cumsum_hashsize
                obj_cumsum = []
                hashsize_cumsum = []
                obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_opt_" + data_type_[ii] + "_" + str(
                    dimension_) + "_" + str(cardinality_[ii]) + '_' + str(with_without_opt_) + ".txt"
                f4 = open(obj_hashsize_file, 'w')
                for mm in range(len(L_Opt_List)):
                    if obj_cumsum.__len__() == 0:
                        obj_cumsum.append(qhull_data_count[mm])
                        hashsize_cumsum.append(qhull_data_count[mm] * L_Opt_List[mm])
                    else:
                        obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                        hashsize_cumsum.append(
                            hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] * L_Opt_List[mm])
                f4.write(','.join(map(repr, obj_cumsum)))
                f4.write("\n")
                f4.write(','.join(map(repr, hashsize_cumsum)))
                f4.close()

                # max cumsum_hashsize
                obj_cumsum = []
                hashsize_cumsum = []
                obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_max_" + data_type_[ii] + \
                                    "_" + str(dimension_) + "_" + str(cardinality_[ii]) + '_' + \
                                    str(with_without_opt_) + ".txt"
                f4 = open(obj_hashsize_file, 'w')
                for mm in range(len(L_Max_List)):
                    if obj_cumsum.__len__() == 0:
                        obj_cumsum.append(qhull_data_count[mm])
                        hashsize_cumsum.append(qhull_data_count[mm] * L_Max_List[mm])
                    else:
                        obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                        hashsize_cumsum.append(
                            hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] *
                            L_Max_List[mm])
                f4.write(','.join(map(repr, obj_cumsum)))
                f4.write("\n")
                f4.write(','.join(map(repr, hashsize_cumsum)))
                f4.close()

                # uni cumsum hashsize
                obj_cumsum = []
                hashsize_cumsum = []
                obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_uni_" + data_type_[ii] + "_" + str(
                    dimension_) + "_" + str(cardinality_[ii]) + '_' + str(with_without_opt_) + ".txt"
                f4 = open(obj_hashsize_file, 'w')
                for mm in range(len(L_Uni_List)):
                    if obj_cumsum.__len__() == 0:
                        obj_cumsum.append(qhull_data_count[mm])
                        hashsize_cumsum.append(qhull_data_count[mm] * L_Uni_List[mm])
                    else:
                        obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                        hashsize_cumsum.append(
                            hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] * L_Uni_List[mm])
                f4.write(','.join(map(repr, obj_cumsum)))
                f4.write("\n")
                f4.write(','.join(map(repr, hashsize_cumsum)))
                f4.close()

                cur_data_type = data_type_[ii]
                cur_cardinality = cardinality_[ii]
                cur_dimension = dimension_
                f10.write("datatype=" + cur_data_type + "\n")
                f10.write("card_of_interest=" + str(card_of_interest_) + "\n")
                f10.write("cardinality=" + str(cur_cardinality) + "\n")
                f10.write("d=" + str(cur_dimension) + "\n")
                f10.write("qn=" + str(query_count) + "\n")
                f10.write("c0=" + str(ratio) + "\n")
                f10.write("pot=" + str(pot_) + "\n")

                temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_opt"
                overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_opt"
                sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_opt"
                f10.write("temporalResult=" + temporalResult + "\n")
                f10.write("overallResult=" + overallResult + "\n")
                f10.write("sim_angle=" + sim_angle + "\n")
                f10.write("S=" + str(sim_threshold) + "\n")
                f10.write("num_layer=" + str(len(K_List)) + "\n")
                f10.write("top_k=" + str(top_k_) + "\n")
                f10.write("# ------------------------------------------------------------------------------ \n")
                f10.write("#     Ground-Truth \n")
                f10.write("# ------------------------------------------------------------------------------ \n")
                f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                f10.write(
                    "tsPath=./result/result_${datatype}_${d}D_${card_of_interest} # path for the ground truth \n")
                f10.write("qPath=./query/query_${d}D.txt \n")
                f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")

                f10.write(" # ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                          "${oFolder}.mip \n")
                f10.write(" # sleep 1 \n")

                f10.write("\n \n \n")
                f10.write("# ------------------------------------------------------------------------------ \n")
                f10.write("#     Layer-Performance \n")
                f10.write("# ------------------------------------------------------------------------------ \n")
                for kk in range(len(K_List)):
                    f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                    f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                    f10.write("L" + str(kk) + "=" + str(L_Opt_List[kk]) + "\n")

                    f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                                 "${cardinality}_qhull_layer_" + str(kk) + "\n")

                    f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                   "${cardinality}_opt/result_${d}D" + str(kk) + "_${K" + str(kk)
                             + "}_${L" + str(kk) + "}" + "\n")
                    f10.write("temp_hash" + str(
                        kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_opt_" + str(kk) + "\n")

                    f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                             "} -L ${L" + str(kk) + "} -LI " + str(
                        kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                             + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                        kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                    f10.write("\n")

                # append overall accuracy computation here
                f10.write("# ------------------------------------------------------------------------------ \n")
                f10.write("#     Overall-Performance \n")
                f10.write("# ------------------------------------------------------------------------------ \n")
                f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                         "${tsPath}.mip -of ${overallResult} \n \n \n")
                f10.write("sleep 2 \n")

                cur_data_type = data_type_[ii]
                cur_cardinality = cardinality_[ii]
                cur_dimension = dimension_
                f10.write("datatype=" + cur_data_type + "\n")
                f10.write("card_of_interest=" + str(card_of_interest_) + "\n")
                f10.write("cardinality=" + str(cur_cardinality) + "\n")
                f10.write("d=" + str(cur_dimension) + "\n")
                f10.write("qn=" + str(query_count) + "\n")
                f10.write("c0=" + str(ratio) + "\n")
                f10.write("pot=" + str(pot_) + "\n")

                temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_max"
                overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_max"
                sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_max"
                f10.write("temporalResult=" + temporalResult + "\n")
                f10.write("overallResult=" + overallResult + "\n")
                f10.write("sim_angle=" + sim_angle + "\n")
                f10.write("S=" + str(sim_threshold) + "\n")
                f10.write("num_layer=" + str(len(K_List)) + "\n")
                f10.write("top_k=" + str(top_k_) + "\n")
                f10.write(
                    "# ------------------------------------------------------------------------------ \n")
                f10.write("#     Ground-Truth \n")
                f10.write(
                    "# ------------------------------------------------------------------------------ \n")
                f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                f10.write(
                    "tsPath=./result/result_${datatype}_${d}D_${card_of_interest} # path for the ground truth \n")
                f10.write("qPath=./query/query_${d}D.txt \n")
                f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
                f10.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                         "${oFolder}.mip \n")
                f10.write("\n \n \n")
                f10.write(
                    "# ------------------------------------------------------------------------------ \n")
                f10.write("#     Layer-Performance \n")
                f10.write(
                    "# ------------------------------------------------------------------------------ \n")
                for kk in range(len(K_List)):
                    f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                    f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                    f10.write("L" + str(kk) + "=" + str(L_Max_List[kk]) + "\n")

                    f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                                 "${cardinality}_qhull_layer_" + str(kk) + "\n")

                    f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                   "${cardinality}_max/result_${d}D" + str(
                        kk) + "_${K" + str(kk)
                             + "}_${L" + str(kk) + "}" + "\n")
                    f10.write("temp_hash" + str(
                        kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_max_" + str(kk) + "\n")

                    f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                             "} -L ${L" + str(kk) + "} -LI " + str(
                        kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                             + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                        kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                    f10.write("\n")

                # append overall accuracy computation here
                f10.write(
                    "# ------------------------------------------------------------------------------ \n")
                f10.write("#     Overall-Performance \n")
                f10.write(
                    "# ------------------------------------------------------------------------------ \n")
                f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                         "${tsPath}.mip -of ${overallResult} \n \n \n")
                f10.write("sleep 2 \n")

                cur_data_type = data_type_[ii]
                cur_cardinality = cardinality_[ii]
                cur_dimension = dimension_
                f10.write("datatype=" + cur_data_type + "\n")
                f10.write("card_of_interest=" + str(card_of_interest_) + "\n")
                f10.write("cardinality=" + str(cur_cardinality) + "\n")
                f10.write("d=" + str(cur_dimension) + "\n")
                f10.write("qn=" + str(query_count) + "\n")
                f10.write("c0=" + str(ratio) + "\n")
                f10.write("pot=" + str(pot_) + "\n")

                temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_uni"
                overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_uni"
                sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_uni"
                f10.write("temporalResult=" + temporalResult + "\n")
                f10.write("overallResult=" + overallResult + "\n")
                f10.write("sim_angle=" + sim_angle + "\n")
                f10.write("S=" + str(sim_threshold) + "\n")
                f10.write("num_layer=" + str(len(K_List)) + "\n")
                f10.write("top_k=" + str(top_k_) + "\n")
                f10.write("# ------------------------------------------------------------------------------ \n")
                f10.write("#     Ground-Truth \n")
                f10.write("# ------------------------------------------------------------------------------ \n")
                f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                f10.write(
                    "tsPath=./result/result_${datatype}_${d}D_${card_of_interest} # path for the ground truth \n")
                f10.write("qPath=./query/query_${d}D.txt \n")
                f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
                f10.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                         "${oFolder}.mip \n")
                f10.write("\n \n \n")
                f10.write("# ------------------------------------------------------------------------------ \n")
                f10.write("#     Layer-Performance \n")
                f10.write("# ------------------------------------------------------------------------------ \n")
                for kk in range(len(K_List)):
                    f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                    f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                    f10.write("L" + str(kk) + "=" + str(L_Uni_List[kk]) + "\n")

                    f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                                 "${cardinality}_qhull_layer_" + str(kk) + "\n")

                    f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                   "${cardinality}_uni/result_${d}D" + str(kk) + "_${K" + str(kk)
                             + "}_${L" + str(kk) + "}" + "\n")
                    f10.write("temp_hash" + str(kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_uni_" + str(kk) + "\n")

                    f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                             "} -L ${L" + str(kk) + "} -LI " + str(
                        kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                             + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                        kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                    f10.write("\n")
                f10.write("# ------------------------------------------------------------------------------ \n")
                f10.write("#     Overall-Performance \n")
                f10.write("# ------------------------------------------------------------------------------ \n")
                f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                         "${tsPath}.mip -of ${overallResult} \n \n \n")
                f10.write("sleep 2 \n")
        f10.close()
    return file_names


################################################################
data_type = ["anti_correlated", "correlated", "random"]
card_excel = '200k'
card_of_interest = '200000'
cardinality = ['95149', '63751', '105366']
types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
sim_threshold = 0.75
dimension = 4
excel_file_before = "./4D_redundancy_3_200k_top25_Merged.xlsx"
top_ks = 25


################################################################
save_file_path_before = '../H2_ALSH/parameters_before/'
separate_string(dimension, excel_file_before, top_ks, types, card_excel, save_file_path_before)

# write and complete ground truth shell
query_count = 1000
ratio = 2
parameter_path_before = '../H2_ALSH/parameters_before/'
parameter_type = ["opt", "max", "uni"]

with_without_opt = 'without_opt'
pot = 0
file_names_without_opt = write_script(data_type, dimension, top_ks, types, card_excel, cardinality, card_of_interest, parameter_path_before, with_without_opt, pot)
repeated_run = 4

aggregated_file_name = BASH_FILE_BASE_FOLDER + "run_bash_" + str(dimension) + "D_all.sh"
f1 = open(aggregated_file_name, 'w')
f1.write("#!/bin/bash \n")
for rr in range(0, repeated_run):
    for files in file_names_without_opt:
        f1.write('sh ' + files + '\n')

    temp_str = "aggregating for non-opt round " + str(rr)
    f1.write('echo \"' + temp_str + '\" \n')

    f1.write('python ../Python_Analysis/LSH_Post_Process_' + str(dimension) + 'D_merged.py without_opt ' + str(rr) + ' \n')
    f1.write('sleep 3' + '\n')

    # before starting pot = 1, clean everything except hash_table
    f1.write('python ../Python_Analysis/Clean_All_' + str(dimension) + 'D_merged.py' + '\n')
    f1.write('sleep 5' + '\n')

f1.close()