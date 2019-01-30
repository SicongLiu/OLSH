# script to generate .sh file for execution
# bash file includes execution of Ground Truth, LSH Scheme and Overall Performance

import math
import os
from openpyxl import load_workbook

k_ranges_25 = ['E6', 'E14', 'E20', 'E28', 'E35', 'E43', 'E48', 'E56', 'E64', 'E72']
l_ranges_opt_25 = ['F6', 'F14', 'F20', 'F28', 'F35', 'F43', 'F48', 'F56', 'F64', 'F72']
l_ranges_max_25 = ['G6', 'G14', 'G20', 'G28', 'G35', 'G43', 'G48', 'G56', 'G64', 'G72']
l_ranges_uni_25 = ['H6', 'H14', 'H20', 'H28', 'H35', 'H43', 'H48', 'H56', 'H64', 'H72']

BASE_FOLDER = "../H2_ALSH/qhull_data/NBA/"
BASH_FILE_BASE_FOLDER = "../H2_ALSH/"


def separate_string(data_type_, dimensions_, excel_file_, top_ks_, types_, card_excel_, cardinality_, save_file_path_):
    wb = load_workbook(filename=excel_file_, data_only=True)

    for i in range(top_ks_.__len__()):
        top_k = top_ks[i]
        if top_k == 25:
            k_ranges_nba = k_ranges_25
            l_ranges_opt_nba = l_ranges_opt_25
            l_ranges_max_nba = l_ranges_max_25
            l_ranges_uni_nba = l_ranges_uni_25
        for k in range(dimensions_.__len__()):
            dimension = dimensions_[k]
            for cc in range(cardinality_.__len__()):
                sheetname = data_type_
                if sheetname in wb.sheetnames:
                    print("Sheetname: " + str(sheetname))
                    ws = wb[sheetname]
                    for m in range(types_.__len__()):
                        type_name = types[m]
                        start = 2 * m
                        end = 2 * m + 1
                        # read data type nba_correlated
                        k_nba = []
                        for row in ws[k_ranges_nba[start]: k_ranges_nba[end]]:
                            for cell in row:
                                k_nba.append(cell.value)

                        l_nba_opt = []
                        for row in ws[l_ranges_opt_nba[start]: l_ranges_opt_nba[end]]:
                            for cell in row:
                                l_nba_opt.append(cell.value)

                        l_nba_max = []
                        for row in ws[l_ranges_max_nba[start]: l_ranges_max_nba[end]]:
                            for cell in row:
                                l_nba_max.append(cell.value)

                        l_nba_uni = []
                        for row in ws[l_ranges_uni_nba[start]: l_ranges_uni_nba[end]]:
                            for cell in row:
                                l_nba_uni.append(cell.value)

                        save_file_dir = save_file_path_ + str(dimension) + "D_top" + str(top_k) + "_" + \
                                        type_name + "_" + card_excel_[cc] + "/"

                        if not os.path.exists(save_file_dir):
                            os.makedirs(save_file_dir)
                        nba_k_name = save_file_dir + "k_NBA"
                        f = open(nba_k_name, 'w')
                        f.write(','.join(map(str, k_nba)))
                        f.close()

                        nba_opt_name = save_file_dir + "l_nba_opt"
                        f = open(nba_opt_name, 'w')
                        f.write(','.join(map(str, l_nba_opt)))
                        f.close()

                        nba_max_name = save_file_dir + "l_nba_max"
                        f = open(nba_max_name, 'w')
                        f.write(','.join(map(str, l_nba_max)))
                        f.close()

                        nba_uni_name = save_file_dir + "l_nba_uni"
                        f = open(nba_uni_name, 'w')
                        f.write(','.join(map(str, l_nba_uni)))
                        f.close()
    wb.close()
    print("Separate String and Save Parameter Files Done .\n")


def write_script(data_type_, dimensions_, top_ks_, types_, card_excel_, cardinality_, parameter_path_, with_without_opt_, pot_):
    file_names = []
    for k in range(dimensions_.__len__()):
        dimension = dimensions_[k]
        for cc in range(cardinality_.__len__()):
            total_bash_file = BASH_FILE_BASE_FOLDER + "run_bash_set_cur_" + str(dimension) + 'D_' + card_excel_[cc] + \
                              '_' + str(with_without_opt_) + '.sh'
            file_names.append(total_bash_file)
            f10 = open(total_bash_file, 'w')
            f10.write("#!/bin/bash \n")
            for m in range(types_.__len__()):
                type_name = types_[m]
                for i in range(top_ks_.__len__()):
                    top_k = top_ks_[i]
                    parameter_dir = parameter_path_ + str(dimension) + "D_top" + str(
                        top_k) + "_" + type_name + "_" + card_excel_[cc] + "/"
                    print(parameter_dir + ' ' + type_name)
                    if not os.path.exists(parameter_dir):
                        continue
                    else:
                        BASH_FILE_FOLDER = BASH_FILE_BASE_FOLDER + "bash_set_" + str(dimension) + "D_top" + str(
                            top_k) + "_" + type_name + "_" + str(
                            cardinality_[cc]) + "/"
                        TEMPORAL_RESULT = "../H2_ALSH/temp_result_" + str(dimension) + "D_top" + str(
                            top_k) + "_" + type_name + "_" + str(cardinality_[cc]) + "/"
                        TEMPORAL_RESULT_FOR_BASH = "./temp_result_" + str(dimension) + "D_top" + str(
                            top_k) + "_" + type_name + "_" + str(cardinality_[cc]) + "/"

                        if not os.path.exists(BASH_FILE_FOLDER):
                            os.makedirs(BASH_FILE_FOLDER)

                        if not os.path.exists(TEMPORAL_RESULT):
                            os.makedirs(TEMPORAL_RESULT)


                        K_List = []
                        L_Opt_List = []
                        L_Max_List = []
                        L_Uni_List = []
                        qhull_data_count = []

                        paramK_path = parameter_dir + "k_" + data_type_ #str(cardinality[k])
                        f1 = open(paramK_path, 'r')
                        K_lines = f1.readlines()
                        temp_length = K_lines[0].split(',').__len__()
                        for k_index in range(min(temp_length, top_k)):
                            K_List.append(int(K_lines[0].split(',')[k_index]))
                        f1.close()

                        paramL_opt_path = parameter_dir + "l_" + data_type_ + "_opt" # str(cardinality[k])
                        f2 = open(paramL_opt_path, 'r')
                        L_lines = f2.readlines()
                        for l_index in range(min(temp_length, top_k)):
                            L_Opt_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                        f2.close()

                        paramL_max_path = parameter_dir + "l_" + data_type_ + "_max"  # str(cardinality[k])
                        f2 = open(paramL_max_path, 'r')
                        L_lines = f2.readlines()
                        for l_index in range(min(temp_length, top_k)):
                            L_Max_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                        f2.close()

                        paramL_uni_path = parameter_dir + "l_" + data_type_ + "_uni"  # str(cardinality[k])
                        f2 = open(paramL_uni_path, 'r')
                        L_lines = f2.readlines()

                        for l_index in range(min(temp_length, top_k)):
                            L_Uni_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                        f2.close()

                        for mm in range(len(K_List)):
                            qhull_file = BASE_FOLDER + data_type_ + "_" + str(dimensions[k]) + "_" + str(cardinality[cc])\
                                         + "_" + "qhull_layer_" + str(mm)
                            f = open(qhull_file, 'r')
                            lines = f.readlines()
                            second_line = lines[1]
                            cur_data_count = int(second_line.split('\n')[0])
                            qhull_data_count.append(cur_data_count)
                            f.close()

                        # opt cumsum_hashsize
                        obj_cumsum = []
                        hashsize_cumsum = []
                        obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_opt_" + data_type_ + "_" + str(
                            dimensions_[k]) + "_" + str(cardinality_[cc]) + '_' + str(with_without_opt_) + ".txt"
                        f4 = open(obj_hashsize_file, 'w')
                        for mm in range(len(L_Opt_List)):
                            if obj_cumsum.__len__() == 0:
                                obj_cumsum.append(qhull_data_count[mm])
                                hashsize_cumsum.append(qhull_data_count[mm] * L_Opt_List[mm])
                            else:
                                obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                hashsize_cumsum.append(
                                    hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] * L_Opt_List[mm])
                        f4.write(','.join(map(repr, obj_cumsum)))
                        f4.write("\n")
                        f4.write(','.join(map(repr, hashsize_cumsum)))
                        f4.close()

                        # max cumsum_hashsize
                        obj_cumsum = []
                        hashsize_cumsum = []
                        obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_max_" + data_type_ + \
                                            "_" + str(dimensions_[k]) + "_" + str(cardinality_[cc]) + '_' + \
                                            str(with_without_opt_) + ".txt"
                        f4 = open(obj_hashsize_file, 'w')
                        for mm in range(len(L_Max_List)):
                            if obj_cumsum.__len__() == 0:
                                obj_cumsum.append(qhull_data_count[mm])
                                hashsize_cumsum.append(qhull_data_count[mm] * L_Max_List[mm])
                            else:
                                obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                hashsize_cumsum.append(
                                    hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] *
                                    L_Max_List[mm])
                        f4.write(','.join(map(repr, obj_cumsum)))
                        f4.write("\n")
                        f4.write(','.join(map(repr, hashsize_cumsum)))
                        f4.close()

                        # uni cumsum hashsize
                        obj_cumsum = []
                        hashsize_cumsum = []
                        obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_uni_" + data_type_ + "_" + str(
                            dimensions_[k]) + "_" + str(cardinality_[cc]) + '_' + str(with_without_opt_) + ".txt"
                        f4 = open(obj_hashsize_file, 'w')
                        for mm in range(len(L_Uni_List)):
                            if obj_cumsum.__len__() == 0:
                                obj_cumsum.append(qhull_data_count[mm])
                                hashsize_cumsum.append(qhull_data_count[mm] * L_Uni_List[mm])
                            else:
                                obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                hashsize_cumsum.append(
                                    hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] * L_Uni_List[mm])
                        f4.write(','.join(map(repr, obj_cumsum)))
                        f4.write("\n")
                        f4.write(','.join(map(repr, hashsize_cumsum)))
                        f4.close()

                        cur_data_type = data_type_
                        cur_cardinality = cardinality_[cc]
                        cur_dimension = dimensions_[k]
                        f10.write("datatype=" + cur_data_type + "\n")
                        f10.write("cardinality=" + str(cur_cardinality) + "\n")
                        f10.write("d=" + str(cur_dimension) + "\n")
                        f10.write("qn=" + str(query_count) + "\n")
                        f10.write("c0=" + str(ratio) + "\n")
                        f10.write("pot=" + str(pot_) + "\n")

                        temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_opt"
                        overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_opt"
                        sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_opt"
                        f10.write("temporalResult=" + temporalResult + "\n")
                        f10.write("overallResult=" + overallResult + "\n")
                        f10.write("sim_angle=" + sim_angle + "\n")
                        f10.write("S=" + str(sim_threshold) + "\n")
                        f10.write("num_layer=" + str(len(K_List)) + "\n")
                        f10.write("top_k=" + str(top_k) + "\n")
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        f10.write("#     Ground-Truth \n")
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        f10.write("dPath=./raw_data/${datatype}/${datatype}_${d}_${cardinality}.txt \n")
                        f10.write(
                            "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                        f10.write("qPath=./query/query_${datatype}.txt \n")
                        f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")

                        f10.write(" # ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                                  "${oFolder}.mip \n")
                        f10.write(" # sleep 1 \n")

                        f10.write("\n \n \n")
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        f10.write("#     Layer-Performance \n")
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        for kk in range(len(K_List)):
                            f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                            f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                            f10.write("L" + str(kk) + "=" + str(L_Opt_List[kk]) + "\n")

                            f10.write("dPath" + str(kk) + "=./qhull_data/${datatype}/${datatype}_${d}_"
                                                         "${cardinality}_qhull_layer_" + str(kk) + "\n")

                            f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                           "${cardinality}_opt/result_${d}D" + str(kk) + "_${K" + str(kk)
                                     + "}_${L" + str(kk) + "}" + "\n")
                            f10.write("temp_hash" + str(
                                kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_opt_" + str(kk) + "\n")

                            f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                                     "} -L ${L" + str(kk) + "} -LI " + str(
                                kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                                     + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                                kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                            f10.write("\n")

                        # append overall accuracy computation here
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        f10.write("#     Overall-Performance \n")
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                                 "${tsPath}.mip -of ${overallResult} \n \n \n")
                        f10.write("sleep 2 \n")

                        cur_data_type = data_type_
                        cur_cardinality = cardinality_[cc]
                        cur_dimension = dimensions_[k]
                        f10.write("datatype=" + cur_data_type + "\n")
                        f10.write("cardinality=" + str(cur_cardinality) + "\n")
                        f10.write("d=" + str(cur_dimension) + "\n")
                        f10.write("qn=" + str(query_count) + "\n")
                        f10.write("c0=" + str(ratio) + "\n")
                        f10.write("pot=" + str(pot_) + "\n")

                        temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_max"
                        overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_max"
                        sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_max"
                        f10.write("temporalResult=" + temporalResult + "\n")
                        f10.write("overallResult=" + overallResult + "\n")
                        f10.write("sim_angle=" + sim_angle + "\n")
                        f10.write("S=" + str(sim_threshold) + "\n")
                        f10.write("num_layer=" + str(len(K_List)) + "\n")
                        f10.write("top_k=" + str(top_k) + "\n")
                        f10.write(
                            "# ------------------------------------------------------------------------------ \n")
                        f10.write("#     Ground-Truth \n")
                        f10.write(
                            "# ------------------------------------------------------------------------------ \n")
                        f10.write("dPath=./raw_data/${datatype}/${datatype}_${d}_${cardinality}.txt \n")
                        f10.write(
                            "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                        f10.write("qPath=./query/query_${datatype}.txt \n")
                        f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
                        f10.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                                 "${oFolder}.mip \n")
                        f10.write("\n \n \n")
                        f10.write(
                            "# ------------------------------------------------------------------------------ \n")
                        f10.write("#     Layer-Performance \n")
                        f10.write(
                            "# ------------------------------------------------------------------------------ \n")
                        for kk in range(len(K_List)):
                            f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                            f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                            f10.write("L" + str(kk) + "=" + str(L_Max_List[kk]) + "\n")

                            f10.write("dPath" + str(kk) + "=./qhull_data/${datatype}/${datatype}_${d}_"
                                                         "${cardinality}_qhull_layer_" + str(kk) + "\n")

                            f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                           "${cardinality}_max/result_${d}D" + str(
                                kk) + "_${K" + str(kk)
                                     + "}_${L" + str(kk) + "}" + "\n")
                            f10.write("temp_hash" + str(
                                kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_max_" + str(kk) + "\n")

                            f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                                     "} -L ${L" + str(kk) + "} -LI " + str(
                                kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                                     + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                                kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                            f10.write("\n")

                        # append overall accuracy computation here
                        f10.write(
                            "# ------------------------------------------------------------------------------ \n")
                        f10.write("#     Overall-Performance \n")
                        f10.write(
                            "# ------------------------------------------------------------------------------ \n")
                        f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                                 "${tsPath}.mip -of ${overallResult} \n \n \n")
                        f10.write("sleep 2 \n")

                        cur_data_type = data_type_
                        cur_cardinality = cardinality_[cc]
                        cur_dimension = dimensions_[k]
                        f10.write("datatype=" + cur_data_type + "\n")
                        f10.write("cardinality=" + str(cur_cardinality) + "\n")
                        f10.write("d=" + str(cur_dimension) + "\n")
                        f10.write("qn=" + str(query_count) + "\n")
                        f10.write("c0=" + str(ratio) + "\n")
                        f10.write("pot=" + str(pot_) + "\n")

                        temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_uni"
                        overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_uni"
                        sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_uni"
                        f10.write("temporalResult=" + temporalResult + "\n")
                        f10.write("overallResult=" + overallResult + "\n")
                        f10.write("sim_angle=" + sim_angle + "\n")
                        f10.write("S=" + str(sim_threshold) + "\n")
                        f10.write("num_layer=" + str(len(K_List)) + "\n")
                        f10.write("top_k=" + str(top_k) + "\n")
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        f10.write("#     Ground-Truth \n")
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        f10.write("dPath=./raw_data/${datatype}/${datatype}_${d}_${cardinality}.txt \n")
                        f10.write(
                            "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                        f10.write("qPath=./query/query_${datatype}.txt \n")
                        f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
                        f10.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                                 "${oFolder}.mip \n")
                        f10.write("\n \n \n")
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        f10.write("#     Layer-Performance \n")
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        for kk in range(len(K_List)):
                            f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                            f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                            f10.write("L" + str(kk) + "=" + str(L_Uni_List[kk]) + "\n")

                            f10.write("dPath" + str(kk) + "=./qhull_data/${datatype}/${datatype}_${d}_"
                                                         "${cardinality}_qhull_layer_" + str(kk) + "\n")

                            f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                           "${cardinality}_uni/result_${d}D" + str(kk) + "_${K" + str(kk)
                                     + "}_${L" + str(kk) + "}" + "\n")
                            f10.write("temp_hash" + str(kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_uni_" + str(kk) + "\n")

                            f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                                     "} -L ${L" + str(kk) + "} -LI " + str(
                                kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                                     + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                                kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                            f10.write("\n")
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        f10.write("#     Overall-Performance \n")
                        f10.write("# ------------------------------------------------------------------------------ \n")
                        f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                                 "${tsPath}.mip -of ${overallResult} \n \n \n")
                        f10.write("sleep 2 \n")
            f10.close()
    return file_names


################################################################
# data_type = ["nba_correlated", "correlated", "random"]
# # card_excel = ['100k', '200k', '500k', '1M', '15M', '2M']
# # cardinality = [100000, 200000, 500000, 1000000, 1500000, 2000000]
# card_excel = ['200k']
# cardinality = [200000]
#
# types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
#
# budgets = ["1M", "10M"]
# sim_threshold = 0.75
# dimensions = [4]
# excel_file_before = "./4D_075_redundancy_3_all_before.xlsx"
# excel_file_after = "./4D_075_redundancy_3_all_after.xlsx"
# top_ks = [10, 25, 50]


################################################################
data_type = "NBA"
card_excel = ['23338']
cardinality = [23338]
types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]

sim_threshold = 0.75
dimensions = [7]
excel_file_before = "./NBA_075_redundancy_2_before.xlsx"
excel_file_after = "./NBA_075_redundancy_2_after.xlsx"
top_ks = [25]


################################################################

save_file_path_before = '../H2_ALSH/parameters_before/'
save_file_path_after = '../H2_ALSH/parameters_after/'

separate_string(data_type, dimensions, excel_file_before, top_ks, types, card_excel, cardinality, save_file_path_before)
separate_string(data_type, dimensions, excel_file_after, top_ks, types, card_excel, cardinality, save_file_path_after)

# write and complete ground truth shell
query_count = 1000
ratio = 2
parameter_path_before = '../H2_ALSH/parameters_before/'
parameter_path_after = '../H2_ALSH/parameters_after/'
parameter_type = ["opt", "max", "uni"]

with_without_opt = 'without_opt'
pot = 0
file_names_without_opt = write_script(data_type, dimensions, top_ks, types, card_excel, cardinality, parameter_path_before, with_without_opt, pot)


with_without_opt = 'with_opt'
pot = 1
file_names_with_opt = write_script(data_type, dimensions, top_ks, types, card_excel, cardinality, parameter_path_after, with_without_opt, pot)

repeated_run = 5

aggregated_file_name = BASH_FILE_BASE_FOLDER + "run_bash_" + str(dimensions[0]) + "D_all.sh"
f1 = open(aggregated_file_name, 'w')
f1.write("#!/bin/bash \n")
for rr in range(0, repeated_run):
    for files in file_names_without_opt:
        f1.write('sh ' + files + '\n')

    temp_str = "aggregating for non-opt round " + str(rr)
    f1.write('echo \"' + temp_str + '\" \n')

    f1.write('python ../Python_Analysis/LSH_Post_Process_' + str(data_type) + '.py without_opt ' + str(rr) + ' \n')
    f1.write('sleep 3' + '\n')

    # before starting pot = 1, clean everything except hash_table
    f1.write('python ../Python_Analysis/Clean_Sim_Overall_run_test_' + str(data_type) + '.py' + '\n')
    f1.write('sleep 5' + '\n')

    for files in file_names_with_opt:
        f1.write('sh ' + files + '\n')

    temp_str = "aggregating for opt round " + str(rr)
    f1.write('echo \"' + temp_str + '\" \n')

    f1.write('python ../Python_Analysis/LSH_Post_Process_' + str(data_type) + '.py with_opt ' + str(rr) + ' \n')
    f1.write('sleep 3' + '\n')

    # before starting pot = 1, clean everything except hash_table
    f1.write('python ../Python_Analysis/Clean_All_' + str(data_type) + '.py' + '\n')
    f1.write('sleep 5' + '\n')

f1.close()