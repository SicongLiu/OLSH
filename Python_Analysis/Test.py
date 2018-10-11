'''
Created on Feb 20, 2018

@author: sicongliu
'''

import numpy as np
import pandas as pd
from openpyxl import load_workbook
import os

excel_file = "../Checkpoint_Result_Oct_10.xlsx"
data_type = ["anti_correlated", "correlated", "random"]
top_ks = [10]
budgets = ["500k", "1M"]
dimensions = [5]

k_ranges_anti = ['E6', 'E15', 'E21', 'E30', 'E37', 'E46', 'E51', 'E60', 'E68', 'E77']
l_ranges_opt_anti = ['F6', 'F15', 'F21', 'F30', 'F37', 'F46', 'F51', 'F60', 'F68', 'F77']
l_ranges_uni_anti = ['G6', 'G15', 'G21', 'G30', 'G37', 'G46', 'G51', 'G60', 'G68', 'G77']

k_ranges_corr = ['U6', 'U15', 'U21', 'U30', 'U37', 'U46', 'U51', 'U60', 'U68', 'U77']
l_ranges_opt_corr = ['V6', 'V15', 'V21', 'V30', 'V37', 'V46', 'V51', 'V60', 'V68', 'V77']
l_ranges_uni_corr = ['W6', 'W15', 'W21', 'W30', 'W37', 'W46', 'W51', 'W60', 'W68', 'W77']

k_ranges_random = ['AJ6', 'AJ15', 'AJ21', 'AJ30', 'AJ37', 'AJ46', 'AJ51', 'AJ60', 'AJ68', 'AJ77']
l_ranges_opt_random = ['AK6', 'AK15', 'AK21', 'AK30', 'AK37', 'AK46', 'AK51', 'AK60', 'AK68', 'AK77']
l_ranges_uni_random = ['AL6', 'AL15', 'AL21', 'AL30', 'AL37', 'AL46', 'AL51', 'AL60', 'AL68', 'AL77']

# dframe = pd.read_excel(excel_file)
wb = load_workbook(filename=excel_file, data_only=True)

types = ["opt", "log", "log_plus", "log_plus_plus", "uni"]
save_file_path = '../H2_ALSH/parameters/'

for i in range(top_ks.__len__()):
    top_k = top_ks[i]
    for j in range(budgets.__len__()):
        budget = budgets[j]
        for k in range(dimensions.__len__()):
            dimension = dimensions[k]
            sheetname = "Budget_" + str(dimension) + "D_top" + str(top_k) + "_budget_" + str(budget)
            if sheetname in wb.sheetnames:
                ws = wb[sheetname]
                for m in range(types.__len__()):
                    type_name = types[m]
                    start = 2 * m
                    end = 2 * m + 1
                    # read data type anti_correlated
                    k_anti = []
                    for row in ws[k_ranges_anti[start]: k_ranges_anti[end]]:
                        for cell in row:
                            k_anti.append(cell.value)

                    l_anti_opt = []
                    for row in ws[l_ranges_opt_anti[start]: l_ranges_opt_anti[end]]:
                        for cell in row:
                            l_anti_opt.append(cell.value)

                    l_anti_uni = []
                    for row in ws[l_ranges_uni_anti[start]: l_ranges_uni_anti[end]]:
                        for cell in row:
                            l_anti_uni.append(cell.value)

                    # read data type correlated
                    k_corr = []
                    for row in ws[k_ranges_corr[start]: k_ranges_corr[end]]:
                        for cell in row:
                            k_corr.append(cell.value)

                    l_corr_opt = []
                    for row in ws[l_ranges_opt_corr[start]: l_ranges_opt_corr[end]]:
                        for cell in row:
                            l_corr_opt.append(cell.value)

                    l_corr_uni = []
                    for row in ws[l_ranges_uni_corr[start]: l_ranges_uni_corr[end]]:
                        for cell in row:
                            l_corr_uni.append(cell.value)

                    # read data type random
                    k_random = []
                    for row in ws[k_ranges_random[start]: k_ranges_random[end]]:
                        for cell in row:
                            k_random.append(cell.value)

                    l_random_opt = []
                    for row in ws[l_ranges_opt_random[start]: l_ranges_opt_random[end]]:
                        for cell in row:
                            l_random_opt.append(cell.value)

                    l_random_uni = []
                    for row in ws[l_ranges_uni_random[start]: l_ranges_uni_random[end]]:
                        for cell in row:
                            l_random_uni.append(cell.value)
                    # save current K and L parameters to files
                    # anti_opt
                    # anti_uni
                    save_file_dir = save_file_path + str(dimension) + "D_top" + str(top_k) + "_budget_" + str(budget)\
                                    + "_" + type_name + "/"

                    if not os.path.exists(save_file_dir):
                        os.makedirs(save_file_dir)
                    anti_k_name = save_file_dir + "k_anti_correlated"
                    f = open(anti_k_name, 'w')
                    f.write(','.join(map(str, k_anti)))
                    f.close()

                    anti_opt_name = save_file_dir + "l_anti_correlated_opt"
                    f = open(anti_opt_name, 'w')
                    f.write(','.join(map(str, l_anti_opt)))
                    f.close()

                    anti_uni_name = save_file_dir + "l_anti_correlated_uni"
                    f = open(anti_uni_name, 'w')
                    f.write(','.join(map(str, l_anti_uni)))
                    f.close()

                    # corr_opt
                    # corr_uni
                    corr_k_name = save_file_dir + "k_correlated"
                    f = open(corr_k_name, 'w')
                    f.write(','.join(map(str, k_corr)))
                    f.close()

                    corr_opt_name = save_file_dir + "l_correlated_opt"
                    f = open(corr_opt_name, 'w')
                    f.write(','.join(map(str, l_corr_opt)))
                    f.close()

                    corr_uni_name = save_file_dir + "l_correlated_uni"
                    f = open(corr_uni_name, 'w')
                    f.write(','.join(map(str, l_corr_uni)))
                    f.close()
                    # random_opt
                    # random_uni
                    random_k_name = save_file_dir + "k_random"
                    f = open(random_k_name, 'w')
                    f.write(','.join(map(str, k_random)))
                    f.close()

                    random_opt_name = save_file_dir + "l_random_opt"
                    f = open(random_opt_name, 'w')
                    f.write(','.join(map(str, l_random_opt)))
                    f.close()

                    random_uni_name = save_file_dir + "l_random_uni"
                    f = open(random_uni_name, 'w')
                    f.write(','.join(map(str, l_random_uni)))
                    f.close()
wb.close()
print("Done .\n")




