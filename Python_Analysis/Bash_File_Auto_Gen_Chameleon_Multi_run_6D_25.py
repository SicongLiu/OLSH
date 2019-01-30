# script to generate .sh file for execution
# bash file includes execution of Ground Truth, LSH Scheme and Overall Performance

import math
import os
from openpyxl import load_workbook
import re


def separate_string(input_string):
    items = []
    match = re.match(r"([a-z]+)([0-9]+)", input_string, re.I)
    if match:
        items = match.groups()
    return items


def compute_index(list_start_, top_m_):
    cell_index_ = []
    for start_ in list_start_:
        cell_index_.append(start_)
        ret_index = separate_string(start_)
        ret_ = ret_index[0] + str(int(ret_index[1]) + top_m_ - 1)
        cell_index_.append(ret_)
    return cell_index_


def computer_used_resource_index(start_, top_m_, count_):
    start_index_ = separate_string(start_)
    cell_index_ = []
    cell_index_.append(start_)
    first_ = start_index_[0] + str(int(start_index_[1]) + top_m_ - 1 + 8)
    cell_index_.append(first_)

    for i in range(2, count_):
        first_index = separate_string(first_)
        first_ = first_index[0] + str(int(first_index[1]) + top_m_ - 1 + 7)
        cell_index_.append(first_)
    return cell_index_

####################################################
cur_topm = 21
data_list_start = 'J6'
k_ranges_start = 'E6'
l_ranges_opt_start = 'F6'
l_ranges_max_start = 'G6'
l_ranges_uni_astart = 'H6'
hash_used_opt_cells_start = 'I27'
hash_used_uni_cells_start = 'O27'

data_anti_list_21 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_anti_21 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_anti_21 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_anti_21 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_anti_21 = compute_index(computer_used_resource_index(l_ranges_uni_astart, cur_topm, 5), cur_topm)
hash_used_anti_opt_cells_21 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_anti_uni_cells_21 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)


cur_topm = 25
data_list_start = 'AA6'
k_ranges_start = 'V6'
l_ranges_opt_start = 'W6'
l_ranges_max_start = 'X6'
l_ranges_uni_start = 'Y6'
hash_used_opt_cells_start = 'Z31'
hash_used_uni_cells_start = 'AF31'


data_corr_list_25 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_corr_25 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_corr_25 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_corr_25 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_corr_25 = compute_index(computer_used_resource_index(l_ranges_uni_start, cur_topm, 5), cur_topm)
hash_used_corr_opt_cells_25 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_corr_uni_cells_25 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)

cur_topm = 20
data_list_start = 'AQ6'
k_ranges_random_start = 'AL6'
l_ranges_opt_random_start = 'AM6'
l_ranges_max_random_start = 'AN6'
l_ranges_uni_random_start = 'AO6'
hash_used_cells_start = 'AP26'
hash_used_uni_cells_13_start = 'AV26'

data_random_list_20 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_random_20 = compute_index(computer_used_resource_index(k_ranges_random_start, cur_topm, 5), cur_topm)
l_ranges_opt_random_20 = compute_index(computer_used_resource_index(l_ranges_opt_random_start, cur_topm, 5), cur_topm)
l_ranges_max_random_20 = compute_index(computer_used_resource_index(l_ranges_max_random_start, cur_topm, 5), cur_topm)
l_ranges_uni_random_20 = compute_index(computer_used_resource_index(l_ranges_uni_random_start, cur_topm, 5), cur_topm)
hash_used_rand_opt_cells_20 = computer_used_resource_index(hash_used_cells_start, cur_topm, 5)
hash_used_rand_uni_cells_20 = computer_used_resource_index(hash_used_uni_cells_13_start, cur_topm, 5)
####################################################


def topk_map(dimension_, cardinality_, type_, top_k_, map_, real_topk):
    temp_key_ = str(dimension_) + '_' + str(cardinality_) + '_' + str(type_) + '_' + str(top_k_)
    map_[temp_key_] = real_topk
    return map_


BASE_FOLDER = "../H2_ALSH/qhull_data/Synthetic/"
BASH_FILE_BASE_FOLDER = "../H2_ALSH/"


def get_real_topk(dimension_, cardinality_, type_, top_k_, map_):
    temp_key_ = str(dimension_) + '_' + str(cardinality_) + '_' + str(type_) + '_' + str(top_k_)
    real_topk = map_[temp_key_]
    return real_topk


def separate_string(data_type_, budgets_, dimensions_, excel_file_, top_ks_, types_, card_excel_, cardinality_, save_file_path_, map_):
    wb = load_workbook(filename=excel_file_, data_only=True)
    for i in range(top_ks_.__len__()):
        top_k = top_ks[i]
        for j in range(budgets_.__len__()):
            budget = budgets_[j]
            for k in range(dimensions_.__len__()):
                dimension = dimensions_[k]
                for cc in range(cardinality_.__len__()):
                    sheetname = "Budget_" + str(budget) + "_" + str(dimension) + "D_top" + str(top_k) + "_" + card_excel_[cc]
                    if sheetname in wb.sheetnames:
                        print("Sheetname: " + str(sheetname))
                        ws = wb[sheetname]
                        real_topk = ws['E1'].value
                        if int(real_topk) == 25:
                            k_ranges_anti = k_ranges_anti_21
                            l_ranges_opt_anti = l_ranges_opt_anti_21
                            l_ranges_max_anti = l_ranges_max_anti_21
                            l_ranges_uni_anti = l_ranges_uni_anti_21
                            map_ = topk_map(dimension, cardinality_[cc], 'anti', top_k, map_, 21)

                            k_ranges_corr = k_ranges_corr_25
                            l_ranges_opt_corr = l_ranges_opt_corr_25
                            l_ranges_max_corr = l_ranges_max_corr_25
                            l_ranges_uni_corr = l_ranges_uni_corr_25
                            map_ = topk_map(dimension, cardinality_[cc], 'corr', top_k, map_, 25)

                            k_ranges_random = k_ranges_random_20
                            l_ranges_opt_random = l_ranges_opt_random_20
                            l_ranges_max_random = l_ranges_max_random_20
                            l_ranges_uni_random = l_ranges_uni_random_20
                            map_ = topk_map(dimension, cardinality_[cc], 'random', top_k, map_, 20)

                        for m in range(types_.__len__()):
                            type_name = types[m]
                            start = 2 * m
                            end = 2 * m + 1
                            # read data type anti_correlated
                            k_anti = []
                            for row in ws[k_ranges_anti[start]: k_ranges_anti[end]]:
                                for cell in row:
                                    k_anti.append(cell.value)

                            l_anti_opt = []
                            for row in ws[l_ranges_opt_anti[start]: l_ranges_opt_anti[end]]:
                                for cell in row:
                                    l_anti_opt.append(cell.value)

                            l_anti_max = []
                            for row in ws[l_ranges_max_anti[start]: l_ranges_max_anti[end]]:
                                for cell in row:
                                    l_anti_max.append(cell.value)

                            l_anti_uni = []
                            for row in ws[l_ranges_uni_anti[start]: l_ranges_uni_anti[end]]:
                                for cell in row:
                                    l_anti_uni.append(cell.value)

                            # read data type correlated
                            k_corr = []
                            for row in ws[k_ranges_corr[start]: k_ranges_corr[end]]:
                                for cell in row:
                                    k_corr.append(cell.value)

                            l_corr_opt = []
                            for row in ws[l_ranges_opt_corr[start]: l_ranges_opt_corr[end]]:
                                for cell in row:
                                    l_corr_opt.append(cell.value)

                            l_corr_max = []
                            for row in ws[l_ranges_max_corr[start]: l_ranges_max_corr[end]]:
                                for cell in row:
                                    l_corr_max.append(cell.value)

                            l_corr_uni = []
                            for row in ws[l_ranges_uni_corr[start]: l_ranges_uni_corr[end]]:
                                for cell in row:
                                    l_corr_uni.append(cell.value)

                            # read data type random
                            k_random = []
                            for row in ws[k_ranges_random[start]: k_ranges_random[end]]:
                                for cell in row:
                                    k_random.append(cell.value)

                            l_random_opt = []
                            for row in ws[l_ranges_opt_random[start]: l_ranges_opt_random[end]]:
                                for cell in row:
                                    l_random_opt.append(cell.value)

                            l_random_max = []
                            for row in ws[l_ranges_max_random[start]: l_ranges_max_random[end]]:
                                for cell in row:
                                    l_random_max.append(cell.value)

                            l_random_uni = []
                            for row in ws[l_ranges_uni_random[start]: l_ranges_uni_random[end]]:
                                for cell in row:
                                    l_random_uni.append(cell.value)
                            # save current K and L parameters to files
                            # anti_opt
                            # anti_uni
                            save_file_dir = save_file_path_ + str(dimension) + "D_top" + str(top_k) + "_budget_" + str(budget)\
                                            + "_" + type_name + "_" + card_excel_[cc] + "/"

                            if not os.path.exists(save_file_dir):
                                os.makedirs(save_file_dir)
                            anti_k_name = save_file_dir + "k_anti_correlated"
                            f = open(anti_k_name, 'w')
                            f.write(','.join(map(str, k_anti)))
                            f.close()

                            anti_opt_name = save_file_dir + "l_anti_correlated_opt"
                            f = open(anti_opt_name, 'w')
                            f.write(','.join(map(str, l_anti_opt)))
                            f.close()

                            anti_max_name = save_file_dir + "l_anti_correlated_max"
                            f = open(anti_max_name, 'w')
                            f.write(','.join(map(str, l_anti_max)))
                            f.close()

                            anti_uni_name = save_file_dir + "l_anti_correlated_uni"
                            f = open(anti_uni_name, 'w')
                            f.write(','.join(map(str, l_anti_uni)))
                            f.close()

                            # corr_opt
                            # corr_uni
                            corr_k_name = save_file_dir + "k_correlated"
                            f = open(corr_k_name, 'w')
                            f.write(','.join(map(str, k_corr)))
                            f.close()

                            corr_opt_name = save_file_dir + "l_correlated_opt"
                            f = open(corr_opt_name, 'w')
                            f.write(','.join(map(str, l_corr_opt)))
                            f.close()

                            corr_max_name = save_file_dir + "l_correlated_max"
                            f = open(corr_max_name, 'w')
                            f.write(','.join(map(str, l_corr_max)))
                            f.close()

                            corr_uni_name = save_file_dir + "l_correlated_uni"
                            f = open(corr_uni_name, 'w')
                            f.write(','.join(map(str, l_corr_uni)))
                            f.close()

                            # random_opt
                            # random_uni
                            random_k_name = save_file_dir + "k_random"
                            f = open(random_k_name, 'w')
                            f.write(','.join(map(str, k_random)))
                            f.close()

                            random_opt_name = save_file_dir + "l_random_opt"
                            f = open(random_opt_name, 'w')
                            f.write(','.join(map(str, l_random_opt)))
                            f.close()

                            random_max_name = save_file_dir + "l_random_max"
                            f = open(random_max_name, 'w')
                            f.write(','.join(map(str, l_random_max)))
                            f.close()

                            random_uni_name = save_file_dir + "l_random_uni"
                            f = open(random_uni_name, 'w')
                            f.write(','.join(map(str, l_random_uni)))
                            f.close()
    wb.close()

    print("Separate String and Save Parameter Files Done .\n")
    return map_


def write_script(data_type_, budgets_, dimensions_, top_ks_, types_, card_excel_, cardinality_, parameter_path_, with_without_opt_, pot_, map_, data_type_for_keys_):
    file_names = []
    for k in range(dimensions_.__len__()):
        dimension = dimensions_[k]
        for cc in range(cardinality_.__len__()):
            total_bash_file = BASH_FILE_BASE_FOLDER + "run_bash_set_cur_" + str(dimension) + 'D_' + card_excel_[cc] + \
                              '_' + str(with_without_opt_) + '.sh'
            file_names.append(total_bash_file)
            f10 = open(total_bash_file, 'w')
            f10.write("#!/bin/bash \n")
            for m in range(types_.__len__()):
                type_name = types_[m]
                for j in range(budgets_.__len__()):
                    budget = budgets_[j]
                    for i in range(top_ks_.__len__()):
                        top_k = top_ks_[i]
                        parameter_dir = parameter_path_ + str(dimension) + "D_top" + str(top_k) + "_budget_" + \
                                        str(budget) + "_" + type_name + "_" + card_excel_[cc] + "/"
                        print(parameter_dir + ' ' + type_name)
                        if not os.path.exists(parameter_dir):
                            continue
                        else:


                            BASH_FILE_FOLDER = BASH_FILE_BASE_FOLDER + "bash_set_" + str(dimension) + "D_top" + str(
                                top_k) + "_budget_" + str(budget) + "_" + type_name + "_" + str(
                                cardinality_[cc]) + "/"
                            TEMPORAL_RESULT = "../H2_ALSH/temp_result_" + str(dimension) + "D_top" + str(
                                top_k) + "_budget_" + str(budget) + "_" + type_name + "_" + str(cardinality_[cc]) + "/"
                            TEMPORAL_RESULT_FOR_BASH = "./temp_result_" + str(dimension) + "D_top" + str(
                                top_k) + "_budget_" + str(budget) + "_" + type_name + "_" + str(cardinality_[cc]) + "/"

                            if not os.path.exists(BASH_FILE_FOLDER):
                                os.makedirs(BASH_FILE_FOLDER)

                            if not os.path.exists(TEMPORAL_RESULT):
                                os.makedirs(TEMPORAL_RESULT)

                            cur_bash_set_dir = "bash_set_" + str(dimension) + "D_top" + str(top_k) + \
                                               "_budget_" + str(budget) + "_" + type_name + "_" + str(
                                cardinality_[cc]) + "/"
                            for ii in range(len(data_type)):
                                K_List = []
                                L_Opt_List = []
                                L_Max_List = []
                                L_Uni_List = []
                                qhull_data_count = []

                                real_topk = get_real_topk(dimension, cardinality[cc], data_type_for_keys_[ii], top_k, map_)
                                # parameter_dir11 = parameter_path_ + str(dimension) + "D_top" + str(real_topk) + "_budget_" + \
                                #                 str(budget) + "_" + type_name + "_" + card_excel_[cc] + "/"
                                paramK_path = parameter_dir + "k_" + data_type_[ii] #str(cardinality[k])
                                f1 = open(paramK_path, 'r')
                                K_lines = f1.readlines()
                                for k_index in range(real_topk):
                                    K_List.append(int(K_lines[0].split(',')[k_index]))
                                f1.close()

                                paramL_opt_path = parameter_dir + "l_" + data_type_[ii] + "_opt" # str(cardinality[k])
                                f2 = open(paramL_opt_path, 'r')
                                L_lines = f2.readlines()
                                for l_index in range(0, real_topk):
                                    L_Opt_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                                f2.close()

                                paramL_max_path = parameter_dir + "l_" + data_type_[ii] + "_max"  # str(cardinality[k])
                                f2 = open(paramL_max_path, 'r')
                                L_lines = f2.readlines()
                                for l_index in range(0, real_topk):
                                    L_Max_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                                f2.close()

                                paramL_uni_path = parameter_dir + "l_" + data_type_[ii] + "_uni"  # str(cardinality[k])
                                f2 = open(paramL_uni_path, 'r')
                                L_lines = f2.readlines()

                                for l_index in range(0, real_topk):
                                    L_Uni_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                                f2.close()

                                for mm in range(len(K_List)):
                                    qhull_file = BASE_FOLDER + data_type_[ii] + "_" + str(dimensions[k]) + "_" + str(cardinality[cc])\
                                                 + "_" + "qhull_layer_" + str(mm)
                                    f = open(qhull_file, 'r')
                                    lines = f.readlines()
                                    second_line = lines[1]
                                    cur_data_count = int(second_line.split('\n')[0])
                                    qhull_data_count.append(cur_data_count)
                                    f.close()

                                # opt cumsum_hashsize
                                obj_cumsum = []
                                hashsize_cumsum = []
                                # obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_opt_" + data_type_[ii] + "_" + str(
                                #     dimensions_[k]) +  "_" + str(cardinality_[cc]) + ".txt"

                                obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_opt_" + data_type_[ii] + \
                                                    "_" + str(dimensions_[k]) + "_" + str(cardinality_[cc]) + '_' + \
                                                    str(with_without_opt_) + ".txt"

                                f4 = open(obj_hashsize_file, 'w')
                                for mm in range(len(L_Opt_List)):
                                    if obj_cumsum.__len__() == 0:
                                        obj_cumsum.append(qhull_data_count[mm])
                                        hashsize_cumsum.append(qhull_data_count[mm] * L_Opt_List[mm])
                                    else:
                                        obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                        hashsize_cumsum.append(
                                            hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] * L_Opt_List[mm])
                                f4.write(','.join(map(repr, obj_cumsum)))
                                f4.write("\n")
                                f4.write(','.join(map(repr, hashsize_cumsum)))
                                f4.close()

                                # max cumsum_hashsize
                                obj_cumsum = []
                                hashsize_cumsum = []
                                # obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_max_" + data_type_[ii] + \
                                #                     "_" + str(dimensions_[k]) + "_" + str(cardinality_[cc]) + ".txt"
                                obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_max_" + data_type_[ii] + \
                                                    "_" + str(dimensions_[k]) + "_" + str(cardinality_[cc]) + '_' + \
                                                    str(with_without_opt_) + ".txt"
                                f4 = open(obj_hashsize_file, 'w')
                                for mm in range(len(L_Max_List)):
                                    if obj_cumsum.__len__() == 0:
                                        obj_cumsum.append(qhull_data_count[mm])
                                        hashsize_cumsum.append(qhull_data_count[mm] * L_Max_List[mm])
                                    else:
                                        obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                        hashsize_cumsum.append(
                                            hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] *
                                            L_Max_List[mm])
                                f4.write(','.join(map(repr, obj_cumsum)))
                                f4.write("\n")
                                f4.write(','.join(map(repr, hashsize_cumsum)))
                                f4.close()

                                # uni cumsum hashsize
                                obj_cumsum = []
                                hashsize_cumsum = []
                                # obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_uni_" + data_type_[ii] + "_" + str(
                                #     dimensions_[k]) + "_" + str(cardinality_[cc]) + ".txt"
                                obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_uni_" + data_type_[ii] + \
                                                    "_" + str(dimensions_[k]) + "_" + str(cardinality_[cc]) + '_' + \
                                                    str(with_without_opt_) + ".txt"
                                f4 = open(obj_hashsize_file, 'w')
                                for mm in range(len(L_Uni_List)):
                                    if obj_cumsum.__len__() == 0:
                                        obj_cumsum.append(qhull_data_count[mm])
                                        hashsize_cumsum.append(qhull_data_count[mm] * L_Uni_List[mm])
                                    else:
                                        obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                        hashsize_cumsum.append(
                                            hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] * L_Uni_List[mm])
                                f4.write(','.join(map(repr, obj_cumsum)))
                                f4.write("\n")
                                f4.write(','.join(map(repr, hashsize_cumsum)))
                                f4.close()

                                cur_data_type = data_type_[ii]
                                cur_cardinality = cardinality_[cc]
                                cur_dimension = dimensions_[k]
                                f10.write("datatype=" + cur_data_type + "\n")
                                f10.write("cardinality=" + str(cur_cardinality) + "\n")
                                f10.write("d=" + str(cur_dimension) + "\n")
                                f10.write("qn=" + str(query_count) + "\n")
                                f10.write("c0=" + str(ratio) + "\n")
                                f10.write("pot=" + str(pot_) + "\n")

                                temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_opt"
                                overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_opt"
                                sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_opt"
                                f10.write("temporalResult=" + temporalResult + "\n")
                                f10.write("overallResult=" + overallResult + "\n")
                                f10.write("sim_angle=" + sim_angle + "\n")
                                f10.write("S=" + str(sim_threshold) + "\n")
                                f10.write("num_layer=" + str(len(K_List)) + "\n")
                                f10.write("top_k=" + str(top_k) + "\n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Ground-Truth \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                                f10.write(
                                    "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                                f10.write("qPath=./query/query_${d}D.txt \n")
                                f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")

                                f10.write(" # ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                                          "${oFolder}.mip \n")
                                f10.write(" # sleep 1 \n")

                                f10.write("\n \n \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Layer-Performance \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                for kk in range(len(K_List)):
                                    f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                                    f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                                    f10.write("L" + str(kk) + "=" + str(L_Opt_List[kk]) + "\n")

                                    f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                                                 "${cardinality}_qhull_layer_" + str(kk) + "\n")

                                    f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                                   "${cardinality}_opt/result_${d}D" + str(kk) + "_${K" + str(kk)
                                             + "}_${L" + str(kk) + "}" + "\n")
                                    f10.write("temp_hash" + str(
                                        kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_opt_" + str(kk) + "\n")

                                    f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                                             "} -L ${L" + str(kk) + "} -LI " + str(
                                        kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                                             + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                                        kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                                    f10.write("\n")

                                # append overall accuracy computation here
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Overall-Performance \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                                         "${tsPath}.mip -of ${overallResult} \n \n \n")
                                f10.write("sleep 2 \n")

                                cur_data_type = data_type_[ii]
                                cur_cardinality = cardinality_[cc]
                                cur_dimension = dimensions_[k]
                                f10.write("datatype=" + cur_data_type + "\n")
                                f10.write("cardinality=" + str(cur_cardinality) + "\n")
                                f10.write("d=" + str(cur_dimension) + "\n")
                                f10.write("qn=" + str(query_count) + "\n")
                                f10.write("c0=" + str(ratio) + "\n")
                                f10.write("pot=" + str(pot_) + "\n")

                                temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_max"
                                overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_max"
                                sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_max"
                                f10.write("temporalResult=" + temporalResult + "\n")
                                f10.write("overallResult=" + overallResult + "\n")
                                f10.write("sim_angle=" + sim_angle + "\n")
                                f10.write("S=" + str(sim_threshold) + "\n")
                                f10.write("num_layer=" + str(len(K_List)) + "\n")
                                f10.write("top_k=" + str(top_k) + "\n")
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Ground-Truth \n")
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                                f10.write(
                                    "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                                f10.write("qPath=./query/query_${d}D.txt \n")
                                f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
                                f10.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                                         "${oFolder}.mip \n")
                                f10.write("\n \n \n")
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Layer-Performance \n")
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                for kk in range(len(K_List)):
                                    f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                                    f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                                    f10.write("L" + str(kk) + "=" + str(L_Max_List[kk]) + "\n")

                                    f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                                                 "${cardinality}_qhull_layer_" + str(kk) + "\n")

                                    f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                                   "${cardinality}_max/result_${d}D" + str(
                                        kk) + "_${K" + str(kk)
                                             + "}_${L" + str(kk) + "}" + "\n")
                                    f10.write("temp_hash" + str(
                                        kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_max_" + str(kk) + "\n")

                                    f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                                             "} -L ${L" + str(kk) + "} -LI " + str(
                                        kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                                             + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                                        kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                                    f10.write("\n")

                                # append overall accuracy computation here
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Overall-Performance \n")
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                                         "${tsPath}.mip -of ${overallResult} \n \n \n")
                                f10.write("sleep 2 \n")

                                cur_data_type = data_type_[ii]
                                cur_cardinality = cardinality_[cc]
                                cur_dimension = dimensions_[k]
                                f10.write("datatype=" + cur_data_type + "\n")
                                f10.write("cardinality=" + str(cur_cardinality) + "\n")
                                f10.write("d=" + str(cur_dimension) + "\n")
                                f10.write("qn=" + str(query_count) + "\n")
                                f10.write("c0=" + str(ratio) + "\n")
                                f10.write("pot=" + str(pot_) + "\n")

                                temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_uni"
                                overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_uni"
                                sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_uni"
                                f10.write("temporalResult=" + temporalResult + "\n")
                                f10.write("overallResult=" + overallResult + "\n")
                                f10.write("sim_angle=" + sim_angle + "\n")
                                f10.write("S=" + str(sim_threshold) + "\n")
                                f10.write("num_layer=" + str(len(K_List)) + "\n")
                                f10.write("top_k=" + str(top_k) + "\n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Ground-Truth \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                                f10.write(
                                    "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                                f10.write("qPath=./query/query_${d}D.txt \n")
                                f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
                                f10.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                                         "${oFolder}.mip \n")
                                f10.write("\n \n \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Layer-Performance \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                for kk in range(len(K_List)):
                                    f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                                    f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                                    f10.write("L" + str(kk) + "=" + str(L_Uni_List[kk]) + "\n")

                                    f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                                                 "${cardinality}_qhull_layer_" + str(kk) + "\n")

                                    f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                                   "${cardinality}_uni/result_${d}D" + str(kk) + "_${K" + str(kk)
                                             + "}_${L" + str(kk) + "}" + "\n")
                                    f10.write("temp_hash" + str(kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_uni_" + str(kk) + "\n")

                                    f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                                             "} -L ${L" + str(kk) + "} -LI " + str(
                                        kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                                             + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                                        kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                                    f10.write("\n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Overall-Performance \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                                         "${tsPath}.mip -of ${overallResult} \n \n \n")
                                f10.write("sleep 2 \n")
            f10.close()
    return file_names


data_type = ["anti_correlated", "correlated", "random"]
data_type_for_keys = ['anti', 'corr', 'random']

card_excel = ['200k']
cardinality = [200000]
types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]

budgets = ["1M", "10M"]
dimensions = [6]
excel_file_before = "./6D_075_top25_200k_before.xlsx"
excel_file_after = "./6D_075_top25_200k_after.xlsx"

top_ks = [25]

save_file_path_before = '../H2_ALSH/parameters_before/'
save_file_path_after = '../H2_ALSH/parameters_after/'

map_before = {}
map_after = {}
map_before = separate_string(data_type, budgets, dimensions, excel_file_before, top_ks, types, card_excel, cardinality, save_file_path_before, map_before)
map_after = separate_string(data_type, budgets, dimensions, excel_file_after, top_ks, types, card_excel, cardinality, save_file_path_after, map_after)

# write and complete ground truth shell
query_count = 1000
ratio = 2
sim_threshold = 0.75
parameter_path_before = '../H2_ALSH/parameters_before/'
parameter_path_after = '../H2_ALSH/parameters_after/'
parameter_type = ["opt", "max", "uni"]

with_without_opt = 'without_opt'
pot = 0
file_names_without_opt = write_script(data_type, budgets, dimensions, top_ks, types, card_excel, cardinality, parameter_path_before, with_without_opt, pot, map_before, data_type_for_keys)

with_without_opt = 'with_opt'
pot = 1
file_names_with_opt = write_script(data_type, budgets, dimensions, top_ks, types, card_excel, cardinality, parameter_path_after, with_without_opt, pot, map_after, data_type_for_keys)

repeated_run = 5

aggregated_file_name = BASH_FILE_BASE_FOLDER + "run_bash_" + str(dimensions[0]) + "D_all.sh"
f1 = open(aggregated_file_name, 'w')
f1.write("#!/bin/bash \n")
for rr in range(0, repeated_run):
    for files in file_names_without_opt:
        f1.write('sh ' + files + '\n')

    temp_str = "aggregating for non-opt round " + str(rr)
    f1.write('echo \"' + temp_str + '\" \n')

    f1.write('python ../Python_Analysis/LSH_Post_Process_' + str(dimensions[0]) + 'D.py without_opt ' + str(rr) + ' \n')
    f1.write('sleep 3' + '\n')

    # before starting pot = 1, clean everything except hash_table
    f1.write('python ../Python_Analysis/Clean_Sim_Overall_run_test_' + str(dimensions[0]) + 'D.py' + '\n')
    f1.write('sleep 5' + '\n')

    for files in file_names_with_opt:
        f1.write('sh ' + files + '\n')

    temp_str = "aggregating for opt round " + str(rr)
    f1.write('echo \"' + temp_str + '\" \n')

    f1.write('python ../Python_Analysis/LSH_Post_Process_' + str(dimensions[0]) + 'D.py with_opt ' + str(rr) + ' \n')
    f1.write('sleep 3' + '\n')

    # before starting pot = 1, clean everything except hash_table
    f1.write('python ../Python_Analysis/Clean_All_' + str(dimensions[0]) + 'D.py' + '\n')
    f1.write('sleep 5' + '\n')

f1.close()