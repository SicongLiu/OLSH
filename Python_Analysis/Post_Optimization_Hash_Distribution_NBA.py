import math
import numpy as np
from openpyxl import load_workbook
import random
import re


def separate_string(input_string):
    items = []
    match = re.match(r"([a-z]+)([0-9]+)", input_string, re.I)
    if match:
        items = match.groups()
    return items


def column_row_index(input_string, column_dist):
    items = separate_string(input_string)
    column_index = items[0]
    row_index = int(items[1]) + column_dist
    return column_index + str(row_index)


def compute_weights(data_list_, top_m_):
    data_list_ = np.asarray(data_list_)
    contribution_list = []
    weight_list_ = []
    data_list_cumsum = np.cumsum(data_list_)
    for i in range(top_m_):
        c_h = data_list_[i]
        temp_contribution = 0
        for j in range(i, top_m_):
            denominator = data_list_cumsum[j]
            temp_contribution = 1.0 * temp_contribution + (1.0 * c_h/denominator)
        contribution_list.append(temp_contribution)
    sum_contribution = sum(contribution_list)
    for i in range(top_m_):
        temp_weight = (1.0 * contribution_list[i])/sum_contribution
        weight_list_.append(temp_weight)
    return weight_list_


def compute_collision_prob(dimension_, data_list_):
    prob_list_ = []
    for ii in data_list_:
        theta = (2 * math.pi/ii) ^ (1/(dimension_ - 1))
        collision = 1 - theta / math.pi
        prob_list_.append(collision)
    return prob_list_


def post_optimization_opt_revised(collision_probilities_, weight_list_, total_error_, data_list_, K_List_, L_List_, hash_used_, hash_budget_):
    smallest = min(data_list_)
    smallest_index = data_list_.index(min(data_list_))

    # total_error = 0.412502654
    total_error_gain = 0
    flag = False
    while hash_used_ + smallest <= hash_budget_:
        # the condition of improvement is to check if the total error rate drops

        # loop through, keep track of hash-relocation and error rate drop
        # find the biggest error gain, while keep hash-resources constraints
        # update hash_used, LList
        delta_error_list = []
        for i in range(len(data_list_)):
            cur_k = K_List_[i]
            cur_l = L_List_[i]
            c_old = weight_list_[i] * math.pow((1 - math.pow(collision_probilities_[i], cur_k)), cur_l)

            # each time increment hash layer by 1
            c_new = weight_list_[i] * math.pow((1 - math.pow(collision_probilities_[i], cur_k)), (cur_l + 1))
            delta_error_list.append((c_old - c_new))

        # sort and check each delta_error
        delta_error_list = np.asarray(delta_error_list)
        # sort in descending order
        sorted_index = delta_error_list.argsort()[::-1][:len(delta_error_list)]
        sorted_pivot = 0
        while sorted_pivot < len(sorted_index):
            cur_index = sorted_index[sorted_pivot]

            # each time increment hash layer by 1
            # temp_hash_used = hash_used + (LList[cur_index] + 1) * data_list[cur_index]
            temp_hash_used = hash_used_ + data_list_[cur_index]
            if temp_hash_used <= hash_budget_:
                L_List_[cur_index] = L_List_[cur_index] + 1
                hash_used_ = hash_used_ + data_list_[cur_index]
                break
            sorted_pivot = sorted_pivot + 1
    total_error = 0
    total_hash_used = 0
    for i in range(len(L_List_)):
        cur_k = K_List_[i]
        cur_l = L_List_[i]
        total_error = total_error + weight_list_[i] * math.pow((1 - math.pow(collision_probility_, cur_k)), cur_l)
        total_hash_used = total_hash_used + data_list_[i] * cur_l
    # print("Updated total error: " + str(total_error))
    # print("total hash used: " + str(total_hash_used))
    # print("Optimized approach done")
    return L_List_

def post_optimization_opt(collision_probility_, weight_list_, total_error_, data_list_, K_List_, L_List_, hash_used_, hash_budget_):
    smallest = min(data_list_)
    smallest_index = data_list_.index(min(data_list_))

    # total_error = 0.412502654
    total_error_gain = 0
    flag = False
    while hash_used_ + smallest <= hash_budget_:
        # the condition of improvement is to check if the total error rate drops

        # loop through, keep track of hash-relocation and error rate drop
        # find the biggest error gain, while keep hash-resources constraints
        # update hash_used, LList
        delta_error_list = []
        for i in range(len(data_list_)):
            cur_k = K_List_[i]
            cur_l = L_List_[i]
            c_old = weight_list_[i] * math.pow((1 - math.pow(collision_probility_, cur_k)), cur_l)

            # each time increment hash layer by 1
            c_new = weight_list_[i] * math.pow((1 - math.pow(collision_probility_, cur_k)), (cur_l+1))
            delta_error_list.append((c_old - c_new))

        # sort and check each delta_error
        delta_error_list = np.asarray(delta_error_list)
        # sort in descending order
        sorted_index = delta_error_list.argsort()[::-1][:len(delta_error_list)]
        sorted_pivot = 0
        while sorted_pivot < len(sorted_index):
            cur_index = sorted_index[sorted_pivot]

            # each time increment hash layer by 1
            # temp_hash_used = hash_used + (LList[cur_index] + 1) * data_list[cur_index]
            temp_hash_used = hash_used_ + data_list_[cur_index]
            if temp_hash_used < hash_budget_:
                L_List_[cur_index] = L_List_[cur_index] + 1
                hash_used_ = hash_used_ + data_list_[cur_index]
                break
            sorted_pivot = sorted_pivot + 1
    total_error = 0
    total_hash_used = 0
    for i in range(len(L_List_)):
        cur_k = K_List_[i]
        cur_l = L_List_[i]
        total_error = total_error + weight_list_[i] * math.pow((1 - math.pow(collision_probility_, cur_k)), cur_l)
        total_hash_used = total_hash_used + data_list_[i] * cur_l
    # print("Updated total error: " + str(total_error))
    # print("total hash used: " + str(total_hash_used))
    # print("Optimized approach done")
    return L_List_


####################################################################################
# for uniformly distributed approach
# uniformly pick one that fits the current hash budget instead of the one minimizing total error rate
def post_optimization_uni(data_list_, L_List_, hash_used_, hash_budget_):
    flag = False
    smallest = min(data_list_)
    data_list_ = np.asarray(data_list_)
    while hash_used_ + smallest <= hash_budget_:
        # sort in descending order
        sorted_index = data_list_.argsort()[::-1][:len(data_list_)]

        temp_pivot_list = []
        # first find all the allow current hash re-allocation
        for i in range(len(sorted_index)):
            temp_pivot_index = sorted_index[i]
            if hash_used_ + data_list_[temp_pivot_index] <= hash_budget_:
                temp_pivot_list.append(temp_pivot_index)
        # randomly pick one from those
        cur_pivot = random.choice(temp_pivot_list)
        # update L_Uni_List, hash_used
        L_List_[cur_pivot] = L_List_[cur_pivot] + 1
        hash_used_ = hash_used_ + data_list_[cur_pivot]
    return L_List_


####################################################################################
# types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
# Data_Types = ['anti_correlated', 'correlated', 'random']
types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
Data_Types = ['anti_correlated']
top_m_cell = 'E1'


# need to update budget cell per data type
budget_cell_anti = 'B4'
budget_cell_corr = 'S4'
budget_cell_rand = 'AI4'


cardinality_cell = 'B2'
top_m_cardinality_anti = 0
top_m_cardinality_corr = 0
top_m_cardinality_random = 0

data_anti_list_9 = ['J6',  'J14', 'J20', 'J28', 'J35', 'J43', 'J48', 'J56', 'J64', 'J72']
k_ranges_anti_9 = ['E6',  'E14', 'E20', 'E28', 'E35', 'E43', 'E48', 'E56', 'E64', 'E72']
l_ranges_opt_anti_9 = ['F6', 'F14', 'F20', 'F28', 'F35', 'F43', 'F48', 'F56', 'F64', 'F72']
l_ranges_max_anti_9 = ['G6', 'G14', 'G20', 'G28', 'G35', 'G43', 'G48', 'G56', 'G64', 'G72']
l_ranges_uni_anti_9 = ['H6', 'H14', 'H20', 'H28', 'H35', 'H43', 'H48', 'H56', 'H64', 'H72']
hash_used_anti_opt_cells_9 = ['I15', 'I29', 'I44', 'I57', 'I73']
hash_used_anti_uni_cells_9 = ['O15', 'O29', 'O44', 'O57', 'O73']

collision_probility = 0.75
total_error = 0


####################################################################################
excel_file_dir = './'

# for each excel file

excel_file_name = excel_file_dir + 'NBA_075_redundancy_4_after.xlsx'
wb = load_workbook(filename=excel_file_name, data_only=True)
wb1 = load_workbook(filename=excel_file_name)
wss = wb.get_sheet_names()

for wwss in wss:
    print(wwss)
    ws = wb.get_sheet_by_name(wwss)
    ws1 = wb1.get_sheet_by_name(wwss)
    top_m = ws[top_m_cell].value
    # ws = wb.get_sheet_by_name(wss[0])
    # ws1 = wb1.get_sheet_by_name(wss[0])
    # top_m = ws[top_m_cell].value
    # print("Sheet name: " + str(ws) + " top-m : " + str(top_m))
    hash_budget_anti = ws[budget_cell_anti].value
    hash_budget_corr = ws[budget_cell_corr].value
    hash_budget_rand = ws[budget_cell_rand].value
    total_cardinality = ws[cardinality_cell].value

    top_m_cardinality_anti_cell = 'J15'
    top_m_cardinality_anti = ws[top_m_cardinality_anti_cell].value

    hash_used_anti_opt_cells = hash_used_anti_opt_cells_9
    hash_used_anti_uni_cells = hash_used_anti_uni_cells_9
    data_anti_list = data_anti_list_9
    k_ranges_anti = k_ranges_anti_9
    l_ranges_opt_anti = l_ranges_opt_anti_9
    l_ranges_max_anti = l_ranges_max_anti_9
    l_ranges_uni_anti = l_ranges_uni_anti_9

    data_anti = []
    data_corr = []
    data_random = []
    # within each excel file, for each data type
    for j in range(len(Data_Types)):
        # read data list
        data_anti_list_start = data_anti_list[0]
        data_anti_list_end = data_anti_list[1]

        for columns in ws[data_anti_list_start: data_anti_list_end]:
            for cell in columns:
                data_anti.append(cell.value)

        weight_anti = compute_weights(data_anti, len(data_anti))

        # for each type, log, log_minus, log_plus, etc
        for jj in range(types.__len__()):
            # read k and l
            type_name = types[jj]
            start = 2 * jj
            end = 2 * jj + 1
            k_anti = []
            for columns in ws[k_ranges_anti[start]: k_ranges_anti[end]]:
                for cell in columns:
                    k_anti.append(cell.value)

            l_anti_opt = []
            for columns in ws[l_ranges_opt_anti[start]: l_ranges_opt_anti[end]]:
                for cell in columns:
                    l_anti_opt.append(cell.value)

            l_anti_max = []
            for columns in ws[l_ranges_max_anti[start]: l_ranges_max_anti[end]]:
                for cell in columns:
                    l_anti_max.append(cell.value)

            l_anti_uni = []
            for columns in ws[l_ranges_uni_anti[start]: l_ranges_uni_anti[end]]:
                for cell in columns:
                    l_anti_uni.append(cell.value)


            hash_used_anti_opt_cell = hash_used_anti_opt_cells[jj]

            hash_used_anti_opt = ws[hash_used_anti_opt_cell].value

            # update LList
            l_anti_opt = post_optimization_opt(collision_probility, weight_anti, total_error, data_anti, k_anti, l_anti_opt,
                                           hash_used_anti_opt, hash_budget_anti)

            hash_used_anti_uni_cell = hash_used_anti_uni_cells[jj]

            hash_used_anti_uni = ws[hash_used_anti_uni_cell].value

            l_anti_uni = post_optimization_uni(data_anti, l_anti_uni, hash_used_anti_uni, hash_budget_anti)

            # write udpate LList back to excel file
            for kk in range(len(l_anti_opt)):
                cur_cell_anti_opt = column_row_index(l_ranges_opt_anti[start], kk)

                cur_cell_anti_uni = column_row_index(l_ranges_uni_anti[start], kk)

                ws1[cur_cell_anti_opt] = l_anti_opt[kk]

                ws1[cur_cell_anti_uni] = l_anti_uni[kk]
    wb1.save(excel_file_name)

print("All done")