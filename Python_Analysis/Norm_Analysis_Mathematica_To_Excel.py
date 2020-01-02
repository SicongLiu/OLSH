import os
import re
from openpyxl import load_workbook
import string


def separate_string(input_string):
    items = []
    match = re.match(r"([a-z]+)([0-9]+)", input_string, re.I)
    if match:
        items = match.groups()
    return items


# input_string 'F6'
def column_row_index(input_string, column_dist):
    items = separate_string(input_string)
    column_index = items[0]
    row_index = int(items[1]) + column_dist
    return column_index + str(row_index)


def get_file_info(full_file_name_):
    file_info_ = []
    file_name_ = full_file_name_.split('.txt')[0]
    temp_file_info_ = file_name_.split('_')
    dims_ = temp_file_info_[0]
    cards_ = re.sub('[.]', '', temp_file_info_[1])
    if full_file_name_.__contains__('anti'):
        data_distributed_type = 'anti_correlated'
    elif full_file_name_.__contains__('correlated'):
        data_distributed_type = 'correlated'
    else:
        data_distributed_type = 'random'

    if full_file_name_.__contains__('EW'):
        resource_type = 'EW'
    elif full_file_name_.__contains__('card'):
        resource_type = 'ED_card'
    else:
        resource_type = 'ED_prob'
    types_= str(resource_type) + '_' + str(dims_) + '_' + str(cards_)

    # print('dim: ' + str(dims_) + ' cards: ' + str(cards_) + ' types: ' + str(types_))
    file_info_.append(data_distributed_type)
    file_info_.append(str(types_))
    return file_info_


def find_sheet(sheet_names_, ending_):
    for sheet_name_ in sheet_names_:
        if re.search(ending_, sheet_name_, re.IGNORECASE):
            return sheet_name_
    return None


def process_line(line_concate_):
    processed_line_ = []
    temp_arr = line_concate_[line_concate_.find("{") + 1:line_concate_.find("}")].split(',')
    for iii in range(len(temp_arr)):
        temp_str = temp_arr[iii].strip()
        # if temp_str[len(temp_str) - 2] == '-':
        if '^' in temp_str:
            temp_num = 0
        else:
            temp_num = round(float(temp_str))
        processed_line_.append(temp_num)
    return processed_line_


def read_file_lines(full_file_path_, read_lines_):
    f1 = open(full_file_path_, 'r')
    bash_count = 0
    file_dict_ = {}
    line_concate_ = ''
    for line in f1.readlines():
        line_concate_ = line_concate_ + line.split('\n')[0].strip()
        line_end = line.split('\n')[0].strip()
        if line_end.endswith('}') and not line_end.startswith('{'):
            file_dict_[bash_count] = process_line(line_concate_)
            bash_count = bash_count + 1
            line_concate_ = ''
    f1.close()
    return file_dict_


def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


# J, AA AQ
def compute_data_list_start_end(data_type_index_, data_list, column_dist_):
    start_items_ = separate_string(data_list[0])
    start_letter_ = start_items_[0]
    start_num_ = start_items_[1]

    end_items_ = separate_string(data_list[1])
    end_letter_ = end_items_[0]
    end_num_ = end_items_[1]

    start_letter_integer_ = int(col2num(start_letter_))
    start_letter_integer_ = start_letter_integer_ + column_dist_ * data_type_index_
    start_letter_ = colnum_string(start_letter_integer_)
    data_list_start = start_letter_ + start_num_
    data_list_end = start_letter_ + end_num_

    return data_list_start, data_list_end


# both letter and integer row index change
def compute_ranges_start_end(column_shift, row_shift, ranges, row_dist_, column_dist_):
    range_start, range_end = compute_data_list_start_end(column_shift, ranges, column_dist_)
    range_column = separate_string(range_start)[0]
    range_row_start = separate_string(range_start)[1] # 6
    range_row_end = separate_string(range_end)[1] # 45
    row_span_ = int(range_row_end) - int(range_row_start)

    target_row_start = int(range_row_end) * row_shift + row_dist_ # 51
    target_row_end = target_row_start + row_span_
    k_ranges_temp_start = range_column + str(target_row_start)
    k_ranges_temp_end = range_column + str(target_row_end)

    return k_ranges_temp_start, k_ranges_temp_end


data_list_40 = ['J6',  'J45', 'J51', 'J90']
k_ranges_40 = ['E6',  'E45', 'E51', 'E90']
l_ranges_opt_40 = ['F6', 'F45']
l_ranges_max_40 = ['G6', 'G45']
l_ranges_uni_40 = ['H6', 'H45']
hash_used_opt_cells_40 = ['I46', 'I45']
hash_used_uni_cells_40 = ['O46', 'O45']

data_list_60 = ['J6', 'J65', 'J71', 'J130']
k_ranges_60 = ['E6', 'E65']
l_ranges_opt_60 = ['F6', 'F65']
l_ranges_max_60 = ['G6', 'G65']
l_ranges_uni_60 = ['H6', 'H65']
hash_used_opt_cells_60 = ['I31', 'I63']
hash_used_uni_cells_60 = ['O31', 'O63']


data_list_80 = ['J6', 'J55']
k_ranges_80 = ['E6', 'E55']
l_ranges_opt_80 = ['F6', 'F55']
l_ranges_max_80 = ['G6', 'G55']
l_ranges_uni_80 = ['H6', 'H55']
hash_used_opt_cells_80 = ['I56', 'I113']
hash_used_uni_cells_80 = ['O56', 'O113']


####################################################################################
column_dist = 17
row_dist = 6
text_file_path = '/Users/sicongliu/'
parameter_file_path = '../H2_ALSH/parameters/'


dimension = 100
excel_file_dir = './'
collision_probility = 0.75
read_lines = -1
excel_file_name = excel_file_dir + str(dimension) + 'D_1M_60_bins_before.xlsx'

bin_count_cell = 'E1'
k_types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
Data_Types = ['anti_correlated', 'correlated', 'random']

wb1 = load_workbook(filename=excel_file_name)
# wss = wb1.get_sheet_names()
wss = wb1.sheetnames

file_dimension_str = str(dimension) + 'D'
for file in os.listdir(text_file_path):
    if file.startswith(file_dimension_str) and file.endswith(".txt"):
        full_file_path = os.path.join(text_file_path, file)
        file_info = get_file_info(file)
        data_distributed_type = file_info[0]
        file_endings = file_info[1]

        sheet_name = find_sheet(wss, file_endings)
        print(file_endings, sheet_name)
        if sheet_name is not None:
            print('sheet name: ' + str(sheet_name) + ' , data type: ' + data_distributed_type)
            opt_l_info = []

            # ws1 = wb1.get_sheet_by_name(sheet_name)
            ws1 = wb1[sheet_name]
            bin_count = ws1[bin_count_cell].value

            if bin_count == 40:
                read_lines = 3
                data_list = data_list_40
                k_ranges_ = k_ranges_40
                l_ranges_opt_ = l_ranges_opt_40
                l_ranges_max_ = l_ranges_max_40
                l_ranges_uni_ = l_ranges_uni_40

                opt_l_info = read_file_lines(full_file_path, read_lines)

            elif bin_count == 60:
                read_lines = 3
                data_list = data_list_60
                k_ranges_ = k_ranges_60
                l_ranges_opt_ = l_ranges_opt_60
                l_ranges_max_ = l_ranges_max_60
                l_ranges_uni_ = l_ranges_uni_60

                opt_l_info = read_file_lines(full_file_path, read_lines)

            else:
                read_lines = 3
                data_list = data_list_80
                k_ranges_ = k_ranges_80
                l_ranges_opt_ = l_ranges_opt_80
                l_ranges_max_ = l_ranges_max_80
                l_ranges_uni_ = l_ranges_uni_80

                opt_l_info = read_file_lines(full_file_path, read_lines)

            # Data_Types = ['anti_correlated', 'correlated', 'random']
            ii = Data_Types.index(data_distributed_type)

            # k_types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
            for kk in range(k_types.__len__()):
                # populate count of data elements
                d_list_start, d_list_end = compute_ranges_start_end(ii, kk, data_list, row_dist, column_dist)

                l_ranges_opt_temp_start, l_ranges_opt_temp_end = compute_ranges_start_end(ii, kk,
                                                                                          l_ranges_opt_,
                                                                                          row_dist,
                                                                                          column_dist)

                # populate data list
                d_index_ = 0
                d_start_items = separate_string(d_list_start)
                d_end_items = separate_string(d_list_end)
                d_start_num_ = int(d_start_items[1])
                d_end_num_ = int(d_end_items[1])
                d_temp_letter_ = d_start_items[0]

                for d_row_index_ in range(d_start_num_, d_end_num_ + 1):
                    d_cell_index_ = d_temp_letter_ + str(d_row_index_)
                    ws1[d_cell_index_] = opt_l_info[0][d_index_]
                    d_index_ = d_index_ + 1


                # populate l_values
                l_index_ = 0
                start_items = separate_string(l_ranges_opt_temp_start)
                end_items = separate_string(l_ranges_opt_temp_end)
                start_num_ = int(start_items[1])
                end_num_ = int(end_items[1])
                temp_letter_ = start_items[0]

                for row_index_ in range(start_num_, end_num_ + 1):
                    cell_index_ = temp_letter_ + str(row_index_)
                    # print('cell_index: ', cell_index_, kk, l_index_)
                    # ws1[cell_index_] = opt_l_info[kk + 1][l_index_]


                    ws1[cell_index_] = opt_l_info[kk + 1][l_index_]
                    # ws1[cell_index_] = opt_l_info[kk][l_index_] # ws1[cell_index_] = max(opt_l_info[kk][l_index_], 1)
                    l_index_ = l_index_ + 1
                wb1.save(excel_file_name)

wb1.save(excel_file_name)

print("All done")