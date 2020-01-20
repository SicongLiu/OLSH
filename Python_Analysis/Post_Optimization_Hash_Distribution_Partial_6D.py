import math
import numpy as np
from openpyxl import load_workbook
import random
import re


def separate_string(input_string):
    items = []
    match = re.match(r"([a-z]+)([0-9]+)", input_string, re.I)
    if match:
        items = match.groups()
    return items


def column_row_index(input_string, column_dist):
    items = separate_string(input_string)
    column_index = items[0]
    row_index = int(items[1]) + column_dist
    return column_index + str(row_index)


def compute_weights(data_list_, top_m_):
    data_list_ = np.asarray(data_list_)
    contribution_list = []
    weight_list_ = []
    data_list_cumsum = np.cumsum(data_list_)
    for i in range(top_m_):
        c_h = data_list_[i]
        temp_contribution = 0
        for j in range(i, top_m_):
            denominator = data_list_cumsum[j]
            temp_contribution = 1.0 * temp_contribution + (1.0 * c_h/denominator)
        contribution_list.append(temp_contribution)
    sum_contribution = sum(contribution_list)
    for i in range(top_m_):
        temp_weight = (1.0 * contribution_list[i])/sum_contribution
        weight_list_.append(temp_weight)
    return weight_list_


# def compute_collision_prob(dimension_, data_list_):
#     prob_list_ = []
#     for ii in data_list_:
#         theta = (2 * math.pi/ii) ^ (1/(dimension_ - 1))
#         collision = 1 - theta / math.pi
#         prob_list_.append(collision)
#     return prob_list_

def compute_collision_prob(dimension_, data_list_):
    prob_list_ = []
    for ii in data_list_:
        theta = pow((2 * math.pi/ii), (1/(dimension_ - 1)))
        collision = 1 - theta / math.pi
        # prob_list_.append(collision)
        prob_list_.append(0.75)
    return prob_list_


def post_optimization_opt_revised(collision_probilities_, weight_list_, total_error_, data_list_, K_List_, L_List_, hash_used_, hash_budget_):
    smallest = min(data_list_)
    smallest_index = data_list_.index(min(data_list_))

    # total_error = 0.412502654
    total_error_gain = 0
    flag = False
    while hash_used_ + smallest <= hash_budget_:
        # the condition of improvement is to check if the total error rate drops

        # loop through, keep track of hash-relocation and error rate drop
        # find the biggest error gain, while keep hash-resources constraints
        # update hash_used, LList
        delta_error_list = []
        for i in range(len(data_list_)):
            cur_k = K_List_[i]
            cur_l = L_List_[i]
            c_old = weight_list_[i] * math.pow((1 - math.pow(collision_probilities_[i], cur_k)), cur_l)

            # each time increment hash layer by 1
            c_new = weight_list_[i] * math.pow((1 - math.pow(collision_probilities_[i], cur_k)), (cur_l + 1))
            delta_error_list.append((c_old - c_new))

        # sort and check each delta_error
        delta_error_list = np.asarray(delta_error_list)
        # sort in descending order
        sorted_index = delta_error_list.argsort()[::-1][:len(delta_error_list)]
        sorted_pivot = 0
        while sorted_pivot < len(sorted_index):
            cur_index = sorted_index[sorted_pivot]

            # each time increment hash layer by 1
            # temp_hash_used = hash_used + (LList[cur_index] + 1) * data_list[cur_index]
            temp_hash_used = hash_used_ + data_list_[cur_index]
            if temp_hash_used <= hash_budget_:
                L_List_[cur_index] = L_List_[cur_index] + 1
                hash_used_ = hash_used_ + data_list_[cur_index]
                break
            sorted_pivot = sorted_pivot + 1
    total_error = 0
    total_hash_used = 0
    for i in range(len(L_List_)):
        cur_k = K_List_[i]
        cur_l = L_List_[i]
        total_error = total_error + weight_list_[i] * math.pow((1 - math.pow(collision_probilities_[i], cur_k)), cur_l)
        total_hash_used = total_hash_used + data_list_[i] * cur_l
    # print("Updated total error: " + str(total_error))
    # print("total hash used: " + str(total_hash_used))
    # print("Optimized approach done")
    return L_List_

def post_optimization_opt(collision_probility_, weight_list_, total_error_, data_list_, K_List_, L_List_, hash_used_, hash_budget_):
    smallest = min(data_list_)
    smallest_index = data_list_.index(min(data_list_))

    # total_error = 0.412502654
    total_error_gain = 0
    flag = False
    while hash_used_ + smallest <= hash_budget_:
        # the condition of improvement is to check if the total error rate drops

        # loop through, keep track of hash-relocation and error rate drop
        # find the biggest error gain, while keep hash-resources constraints
        # update hash_used, LList
        delta_error_list = []
        for i in range(len(data_list_)):
            cur_k = K_List_[i]
            cur_l = L_List_[i]
            c_old = weight_list_[i] * math.pow((1 - math.pow(collision_probility_, cur_k)), cur_l)

            # each time increment hash layer by 1
            c_new = weight_list_[i] * math.pow((1 - math.pow(collision_probility_, cur_k)), (cur_l+1))
            delta_error_list.append((c_old - c_new))

        # sort and check each delta_error
        delta_error_list = np.asarray(delta_error_list)
        # sort in descending order
        sorted_index = delta_error_list.argsort()[::-1][:len(delta_error_list)]
        sorted_pivot = 0
        while sorted_pivot < len(sorted_index):
            cur_index = sorted_index[sorted_pivot]

            # each time increment hash layer by 1
            # temp_hash_used = hash_used + (LList[cur_index] + 1) * data_list[cur_index]
            temp_hash_used = hash_used_ + data_list_[cur_index]
            if temp_hash_used <= hash_budget_:
                L_List_[cur_index] = L_List_[cur_index] + 1
                hash_used_ = hash_used_ + data_list_[cur_index]
                break
            sorted_pivot = sorted_pivot + 1
    total_error = 0
    total_hash_used = 0
    for i in range(len(L_List_)):
        cur_k = K_List_[i]
        cur_l = L_List_[i]
        total_error = total_error + weight_list_[i] * math.pow((1 - math.pow(collision_probility_, cur_k)), cur_l)
        total_hash_used = total_hash_used + data_list_[i] * cur_l
    # print("Updated total error: " + str(total_error))
    # print("total hash used: " + str(total_hash_used))
    # print("Optimized approach done")
    return L_List_


####################################################################################
# for uniformly distributed approach
# uniformly pick one that fits the current hash budget instead of the one minimizing total error rate
def post_optimization_uni(data_list_, L_List_, hash_used_, hash_budget_):
    flag = False
    smallest = min(data_list_)
    data_list_ = np.asarray(data_list_)
    while hash_used_ + smallest <= hash_budget_:
        # sort in descending order
        sorted_index = data_list_.argsort()[::-1][:len(data_list_)]

        temp_pivot_list = []
        # first find all the allow current hash re-allocation
        for i in range(len(sorted_index)):
            temp_pivot_index = sorted_index[i]
            if hash_used_ + data_list_[temp_pivot_index] <= hash_budget_:
                temp_pivot_list.append(temp_pivot_index)
        # randomly pick one from those
        cur_pivot = random.choice(temp_pivot_list)
        # update L_Uni_List, hash_used
        L_List_[cur_pivot] = L_List_[cur_pivot] + 1
        hash_used_ = hash_used_ + data_list_[cur_pivot]
    return L_List_



def compute_index(list_start_, top_m_):
    cell_index_ = []
    for start_ in list_start_:
        cell_index_.append(start_)
        ret_index = separate_string(start_)
        ret_ = ret_index[0] + str(int(ret_index[1]) + top_m_ - 1)
        cell_index_.append(ret_)
    return cell_index_


def computer_used_resource_index(start_, top_m_, count_):
    start_index_ = separate_string(start_)
    cell_index_ = []
    cell_index_.append(start_)
    first_ = start_index_[0] + str(int(start_index_[1]) + top_m_ - 1 + 8)
    cell_index_.append(first_)

    for i in range(2, count_):
        first_index = separate_string(first_)
        first_ = first_index[0] + str(int(first_index[1]) + top_m_ - 1 + 7)
        cell_index_.append(first_)
    return cell_index_


####################################################################################
# types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
Data_Types = ['anti_correlated', 'correlated', 'random']
types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
top_m_cell_anti = 'E1'
top_m_cell_corr = 'V1'
top_m_cell_rand = 'AL1'


# need to update budget cell per data type
budget_cell_anti = 'B4'
budget_cell_corr = 'S4'
budget_cell_rand = 'AI4'


cardinality_cell = 'B2'
top_m_cardinality_anti = 0
top_m_cardinality_corr = 0
top_m_cardinality_random = 0

####################################################
cur_topm = 21
data_list_start = 'J6'
k_ranges_start = 'E6'
l_ranges_opt_start = 'F6'
l_ranges_max_start = 'G6'
l_ranges_uni_astart = 'H6'
hash_used_opt_cells_start = 'I27'
hash_used_uni_cells_start = 'O27'

data_anti_list_21 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_anti_21 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_anti_21 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_anti_21 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_anti_21 = compute_index(computer_used_resource_index(l_ranges_uni_astart, cur_topm, 5), cur_topm)
hash_used_anti_opt_cells_21 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_anti_uni_cells_21 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)


cur_topm = 25
data_list_start = 'AA6'
k_ranges_start = 'V6'
l_ranges_opt_start = 'W6'
l_ranges_max_start = 'X6'
l_ranges_uni_start = 'Y6'
hash_used_opt_cells_start = 'Z31'
hash_used_uni_cells_start = 'AF31'


data_corr_list_25 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_corr_25 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_corr_25 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_corr_25 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_corr_25 = compute_index(computer_used_resource_index(l_ranges_uni_start, cur_topm, 5), cur_topm)
hash_used_corr_opt_cells_25 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_corr_uni_cells_25 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)

cur_topm = 20
data_list_start = 'AQ6'
k_ranges_random_start = 'AL6'
l_ranges_opt_random_start = 'AM6'
l_ranges_max_random_start = 'AN6'
l_ranges_uni_random_start = 'AO6'
hash_used_cells_start = 'AP26'
hash_used_uni_cells_13_start = 'AV26'

data_random_list_20 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_random_20 = compute_index(computer_used_resource_index(k_ranges_random_start, cur_topm, 5), cur_topm)
l_ranges_opt_random_20 = compute_index(computer_used_resource_index(l_ranges_opt_random_start, cur_topm, 5), cur_topm)
l_ranges_max_random_20 = compute_index(computer_used_resource_index(l_ranges_max_random_start, cur_topm, 5), cur_topm)
l_ranges_uni_random_20 = compute_index(computer_used_resource_index(l_ranges_uni_random_start, cur_topm, 5), cur_topm)
hash_used_rand_opt_cells_20 = computer_used_resource_index(hash_used_cells_start, cur_topm, 5)
hash_used_rand_uni_cells_20 = computer_used_resource_index(hash_used_uni_cells_13_start, cur_topm, 5)
####################################################

collision_probility = 0.75
total_error = 0

# dimensions = [2, 3, 4, 5, 6, 7]
dimensions = [6]
excel_file_dir = './'

# for each excel file
for i in range(len(dimensions)):
    cur_d = dimensions[i]
    # excel_file_name = excel_file_dir + str(cur_d) + 'D_075_top25_200k_after.xlsx'
    excel_file_name = excel_file_dir + str(cur_d) + 'D_075_redundancy_4_all_after.xlsx'
    wb = load_workbook(filename=excel_file_name, data_only=True)
    wb1 = load_workbook(filename=excel_file_name)
    wss = wb.get_sheet_names()

    for wwss in wss:
        print(wwss)
        ws = wb.get_sheet_by_name(wwss)
        ws1 = wb1.get_sheet_by_name(wwss)
        top_m_anti = ws[top_m_cell_anti].value
        top_m_corr = ws[top_m_cell_corr].value
        top_m_rand = ws[top_m_cell_rand].value
        # ws = wb.get_sheet_by_name(wss[0])
        # ws1 = wb1.get_sheet_by_name(wss[0])
        # top_m = ws[top_m_cell].value
        # print("Sheet name: " + str(ws) + " top-m : " + str(top_m))
        hash_budget_anti = ws[budget_cell_anti].value
        hash_budget_corr = ws[budget_cell_corr].value
        hash_budget_rand = ws[budget_cell_rand].value
        total_cardinality = ws[cardinality_cell].value
        if top_m_anti == 25:
            top_m_cardinality_anti_cell = 'J27'
            top_m_cardinality_anti = ws[top_m_cardinality_anti_cell].value
            hash_used_anti_opt_cells = hash_used_anti_opt_cells_21
            hash_used_anti_uni_cells = hash_used_anti_uni_cells_21
            data_anti_list = data_anti_list_21
            k_ranges_anti = k_ranges_anti_21
            l_ranges_opt_anti = l_ranges_opt_anti_21
            l_ranges_max_anti = l_ranges_max_anti_21
            l_ranges_uni_anti = l_ranges_uni_anti_21

            top_m_cardinality_corr_cell = 'AA31'
            top_m_cardinality_corr = ws[top_m_cardinality_corr_cell].value
            hash_used_corr_opt_cells = hash_used_corr_opt_cells_25
            hash_used_corr_uni_cells = hash_used_corr_uni_cells_25
            data_corr_list = data_corr_list_25
            k_ranges_corr = k_ranges_corr_25
            l_ranges_opt_corr = l_ranges_opt_corr_25
            l_ranges_max_corr = l_ranges_max_corr_25
            l_ranges_uni_corr = l_ranges_uni_corr_25

            top_m_cardinality_random_cell = 'AQ27'
            top_m_cardinality_random = ws[top_m_cardinality_random_cell].value
            hash_used_rand_opt_cells = hash_used_rand_opt_cells_20
            hash_used_rand_uni_cells = hash_used_rand_uni_cells_20
            data_random_list = data_random_list_20
            k_ranges_random = k_ranges_random_20
            l_ranges_opt_random = l_ranges_opt_random_20
            l_ranges_max_random = l_ranges_max_random_20
            l_ranges_uni_random = l_ranges_uni_random_20

        data_anti = []
        data_corr = []
        data_random = []
        # within each excel file, for each data type
        # for j in range(len(Data_Types)):
        # read data list
        data_anti_list_start = data_anti_list[0]
        data_anti_list_end = data_anti_list[1]

        data_corr_list_start = data_corr_list[0]
        data_corr_list_end = data_corr_list[1]

        data_random_list_start = data_random_list[0]
        data_random_list_end = data_random_list[1]

        for columns in ws[data_anti_list_start: data_anti_list_end]:
            for cell in columns:
                data_anti.append(cell.value)

        for columns in ws[data_corr_list_start: data_corr_list_end]:
            for cell in columns:
                data_corr.append(cell.value)

        for columns in ws[data_random_list_start: data_random_list_end]:
            for cell in columns:
                data_random.append(cell.value)

        weight_anti = compute_weights(data_anti, len(data_anti))
        weight_corr = compute_weights(data_corr, len(data_corr))
        weight_random = compute_weights(data_random, len(data_random))

        # for each type, log, log_minus, log_plus, etc
        for jj in range(types.__len__()):
            # read k and l
            type_name = types[jj]
            start = 2 * jj
            end = 2 * jj + 1
            k_anti = []
            for columns in ws[k_ranges_anti[start]: k_ranges_anti[end]]:
                for cell in columns:
                    k_anti.append(cell.value)

            l_anti_opt = []
            for columns in ws[l_ranges_opt_anti[start]: l_ranges_opt_anti[end]]:
                for cell in columns:
                    l_anti_opt.append(cell.value)

            l_anti_max = []
            for columns in ws[l_ranges_max_anti[start]: l_ranges_max_anti[end]]:
                for cell in columns:
                    l_anti_max.append(cell.value)

            l_anti_uni = []
            for columns in ws[l_ranges_uni_anti[start]: l_ranges_uni_anti[end]]:
                for cell in columns:
                    l_anti_uni.append(cell.value)

            # read data type correlated
            k_corr = []
            for columns in ws[k_ranges_corr[start]: k_ranges_corr[end]]:
                for cell in columns:
                    k_corr.append(cell.value)

            l_corr_opt = []
            for columns in ws[l_ranges_opt_corr[start]: l_ranges_opt_corr[end]]:
                for cell in columns:
                    l_corr_opt.append(cell.value)

            l_corr_max = []
            for columns in ws[l_ranges_max_corr[start]: l_ranges_max_corr[end]]:
                for cell in columns:
                    l_corr_max.append(cell.value)

            l_corr_uni = []
            for columns in ws[l_ranges_uni_corr[start]: l_ranges_uni_corr[end]]:
                for cell in columns:
                    l_corr_uni.append(cell.value)

            # read data type random
            k_random = []
            for columns in ws[k_ranges_random[start]: k_ranges_random[end]]:
                for cell in columns:
                    k_random.append(cell.value)

            l_random_opt = []
            for columns in ws[l_ranges_opt_random[start]: l_ranges_opt_random[end]]:
                for cell in columns:
                    l_random_opt.append(cell.value)

            l_random_max = []
            for columns in ws[l_ranges_max_random[start]: l_ranges_max_random[end]]:
                for cell in columns:
                    l_random_max.append(cell.value)

            l_random_uni = []
            for columns in ws[l_ranges_uni_random[start]: l_ranges_uni_random[end]]:
                for cell in columns:
                    l_random_uni.append(cell.value)

            hash_used_anti_opt_cell = hash_used_anti_opt_cells[jj]
            hash_used_corr_opt_cell = hash_used_corr_opt_cells[jj]
            hash_used_random_opt_cell = hash_used_rand_opt_cells[jj]

            hash_used_anti_opt = ws[hash_used_anti_opt_cell].value
            hash_used_corr_opt = ws[hash_used_corr_opt_cell].value
            hash_used_random_opt = ws[hash_used_random_opt_cell].value

            # update LList
            # l_anti_opt = post_optimization_opt(collision_probility, weight_anti, total_error, data_anti, k_anti, l_anti_opt,
            #                                hash_used_anti_opt, hash_budget_anti)
            # l_corr_opt = post_optimization_opt(collision_probility, weight_corr, total_error, data_corr, k_corr, l_corr_opt,
            #                                hash_used_corr_opt, hash_budget_corr)
            # l_random_opt = post_optimization_opt(collision_probility, weight_random, total_error, data_random, k_anti, l_random_opt,
            #                                hash_used_random_opt, hash_budget_rand)

            collision_probility_anti = compute_collision_prob(cur_d, data_anti)
            collision_probility_corr = compute_collision_prob(cur_d, data_corr)
            collision_probility_random = compute_collision_prob(cur_d, data_random)

            l_anti_opt = post_optimization_opt_revised(collision_probility_anti, weight_anti, total_error,
                                                       data_anti, k_anti,
                                                       l_anti_opt,
                                                       hash_used_anti_opt, hash_budget_anti)
            l_corr_opt = post_optimization_opt_revised(collision_probility_corr, weight_corr, total_error,
                                                       data_corr, k_corr,
                                                       l_corr_opt,
                                                       hash_used_corr_opt, hash_budget_corr)
            l_random_opt = post_optimization_opt_revised(collision_probility_random, weight_random, total_error,
                                                         data_random, k_anti,
                                                         l_random_opt,
                                                         hash_used_random_opt, hash_budget_rand)

            hash_used_anti_uni_cell = hash_used_anti_uni_cells[jj]
            hash_used_corr_uni_cell = hash_used_corr_uni_cells[jj]
            hash_used_random_uni_cell = hash_used_rand_uni_cells[jj]

            hash_used_anti_uni = ws[hash_used_anti_uni_cell].value
            hash_used_corr_uni = ws[hash_used_corr_uni_cell].value
            hash_used_random_uni = ws[hash_used_random_uni_cell].value

            l_anti_uni = post_optimization_uni(data_anti, l_anti_uni, hash_used_anti_uni, hash_budget_anti)
            l_corr_uni = post_optimization_uni(data_corr, l_corr_uni, hash_used_corr_uni, hash_budget_corr)
            l_random_uni = post_optimization_uni(data_random, l_random_uni, hash_used_random_uni, hash_budget_rand)

            # write udpate LList back to excel file
            for kk in range(len(l_anti_opt)):
                cur_cell_anti_opt = column_row_index(l_ranges_opt_anti[start], kk)
                cur_cell_anti_uni = column_row_index(l_ranges_uni_anti[start], kk)
                ws1[cur_cell_anti_opt] = l_anti_opt[kk]
                ws1[cur_cell_anti_uni] = l_anti_uni[kk]

            for kk in range(len(l_corr_opt)):
                cur_cell_corr_opt = column_row_index(l_ranges_opt_corr[start], kk)
                cur_cell_corr_uni = column_row_index(l_ranges_uni_corr[start], kk)
                ws1[cur_cell_corr_opt] = l_corr_opt[kk]
                ws1[cur_cell_corr_uni] = l_corr_uni[kk]

            for kk in range(len(l_random_opt)):
                cur_cell_random_opt = column_row_index(l_ranges_opt_random[start], kk)
                cur_cell_random_uni = column_row_index(l_ranges_uni_random[start], kk)
                ws1[cur_cell_random_uni] = l_random_uni[kk]
                ws1[cur_cell_random_opt] = l_random_opt[kk]
    wb1.save(excel_file_name)

print("All done")