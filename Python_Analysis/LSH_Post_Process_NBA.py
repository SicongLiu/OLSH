import os
import re
import sys
from openpyxl import load_workbook

result_file_dir = '../H2_ALSH/'

result_type = ['without_threshold']
result_type_excel = ['wo_threshold']

# get obj and hash count from bash_set
obj_hashsize_prefix = 'bash_set_'

# get candidate, recall and NDCG from temp_result
cand_recall_prefix = 'temp_result_'

dimensions = [7]

# optimized_tops = [25, 50]
optimized_tops = [25]

cardinality = 23338
top_ks = [1, 2, 5, 10, 25, 50]
# types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
types = ["log"]
data_type = "NBA"
comp_types = ['opt', 'max', 'uni']

over_reault = result_file_dir + 'all_aggregated_Jan_30.txt'


def separate_string(input_string):
    items = []
    match = re.match(r"([a-z]+)([0-9]+)", input_string, re.I)
    if match:
        items = match.groups()
    return items


def column_row_index(input_string, column_dist):
    items = separate_string(input_string)
    column_index = items[0]
    row_index = int(items[1]) + column_dist
    return column_index + str(row_index)


log_optimized_recall = ['B5',  'B6', 'B7', 'B8', 'B9', 'B10']
log_optimized_NDCG = ['C5',  'C6', 'C7', 'C8', 'C9', 'C10']
log_optimized_cand = ['D5',  'D6', 'D7', 'D8', 'D9', 'D10']
log_optimized_obj = ['E5',  'E6', 'E7', 'E8', 'E9', 'E10']
log_optimized_hashsize = ['F5',  'F6', 'F7', 'F8', 'F9', 'F10']


log_optimized_max_recall = ['G5',  'G6', 'G7', 'G8', 'G9', 'G10']
log_optimized_max_NDCG = ['H5',  'H6', 'H7', 'H8', 'H9', 'H10']
log_optimized_max_cand = ['I5',  'I6', 'I7', 'I8', 'I9', 'I10']
log_optimized_max_obj = ['J5',  'J6', 'J7', 'J8', 'J9', 'J10']
log_optimized_max_hashsize = ['K5',  'K6', 'K7', 'K8', 'K9', 'K10']

log_optimized_uni_recall = ['L5',  'L6', 'L7', 'L8', 'L9', 'L10']
log_optimized_uni_NDCG = ['M5',  'M6', 'M7', 'M8', 'M9', 'M10']
log_optimized_uni_cand = ['N5',  'N6', 'N7', 'N8', 'N9', 'N10']
log_optimized_uni_obj = ['O5',  'O6', 'O7', 'O8', 'O9', 'O10']
log_optimized_uni_hashsize = ['P5',  'P6', 'P7', 'P8', 'P9', 'P10']



f = open(over_reault, 'w')

with_without_opt = str(sys.argv[1])
run_index = str(sys.argv[2])


# with_without_opt = 'without_post_opt'
# run_index = '0'

for dd in range(dimensions.__len__()):
    cur_dimension = dimensions[dd]

    # new excel file here
    excel_file = "Aggregation_" + str(data_type) + "_" + with_without_opt + "_" + run_index + ".xlsx"

    excel_file_hash_hits = "Aggregation_" + str(data_type) + "_" + with_without_opt + "_" + \
                           run_index + "_hash_hits.xlsx"
    # wb = load_workbook(filename=excel_file, data_only=True)
    wb = load_workbook(filename=excel_file)
    wb_hash_hits = load_workbook(filename=excel_file_hash_hits)

    for oo in range(optimized_tops.__len__()):
        cur_top_o = optimized_tops[oo]
        for cr in range(result_type.__len__()):
            cur_cr = result_type[cr]

            # sheet name goes here
            sheet_name = data_type
            for ct in range(comp_types.__len__()):
                cur_ct = comp_types[ct]
                for tt in range(types.__len__()):
                    cur_type = types[tt]
                    column_dist = 0
                    if cur_type == 'log':
                        column_dist = 0
                    elif cur_type == 'log_minus':
                        column_dist = 10
                    elif cur_type == 'log_plus':
                        column_dist = 20
                    elif cur_type == 'log_plus_plus':
                        column_dist = 30
                    else:
                        column_dist = 40

                    start_recall = ''
                    start_NDCG = ''
                    start_cand = ''
                    start_obj = ''
                    start_hashsize = ''
                    if cur_ct == 'opt':
                        start_recall = log_optimized_recall[0]
                        start_NDCG = log_optimized_NDCG[0]
                        start_cand = log_optimized_cand[0]
                        start_obj = log_optimized_obj[0]
                        start_hashsize = log_optimized_hashsize[0]
                    elif cur_ct == 'max':
                        start_recall = log_optimized_max_recall[0]
                        start_NDCG = log_optimized_max_NDCG[0]
                        start_cand = log_optimized_max_cand[0]
                        start_obj = log_optimized_max_obj[0]
                        start_hashsize = log_optimized_max_hashsize[0]
                    else:
                        start_recall = log_optimized_uni_recall[0]
                        start_NDCG = log_optimized_uni_NDCG[0]
                        start_cand = log_optimized_uni_cand[0]
                        start_obj = log_optimized_uni_obj[0]
                        start_hashsize = log_optimized_uni_hashsize[0]

                    obj_file_dir = result_file_dir + obj_hashsize_prefix + str(cur_dimension) + 'D_top' + \
                                str(cur_top_o) + '_' + cur_type + '_' + str(cardinality) + '/'
                    if not os.path.exists(obj_file_dir):
                        continue
                    # ws = wb[sheet_name]
                    print("sheet name: " + str(sheet_name))
                    ws = wb.get_sheet_by_name(sheet_name)

                    ws_hash_hits = wb_hash_hits.get_sheet_by_name(sheet_name)

                    obj_file = obj_file_dir + 'cumsum_hashsize_obj_' + cur_ct + '_' + data_type + '_' + \
                               str(cur_dimension) + '_' + str(cardinality) + '_' + with_without_opt + '.txt'

                    f1 = open(obj_file, 'r')
                    lines = f1.readlines()
                    obj_s = lines[0].split(',')
                    # print(obj_s)
                    hash_s = lines[1].split(',')
                    # print(hash_s)

                    obj = []
                    hash = []
                    top_ks_length = -1
                    if cur_top_o == 10:
                        top_ks_length = 4
                    elif cur_top_o == 25:
                        top_ks_length = 5
                    else:
                        top_ks_length = 6
                    obj_s_length = obj_s.__len__()
                    for oh_index in range(top_ks_length):
                        temp_index_ = min(obj_s_length - 1, top_ks[oh_index] - 1)
                        obj.append(int(obj_s[temp_index_]))
                        hash.append(int(hash_s[temp_index_]))
                    f1.close()

                    temp_result_dir = result_file_dir + cand_recall_prefix + str(cur_dimension) + 'D_top' + \
                                str(cur_top_o) + '_' + cur_type + '_' + str(cardinality) + '/'

                    print("result path: " + str(temp_result_dir))
                    if not os.path.exists(temp_result_dir):
                        continue

                    hash_table_hits_list = []
                    cand_size_list = []
                    for ii in range(top_ks_length):
                        cand_result_file = temp_result_dir + 'run_test_' + data_type + '_' + str(cur_dimension) + '_' + \
                                       str(cardinality) + '_' + cur_ct + '_' + cur_cr + '_top_' + str(top_ks[ii]) + '_candidate_size.txt'
                        cand_size = 0
                        hash_table_hits = 0
                        f1 = open(cand_result_file, 'r')
                        lines = f1.readlines()
                        for jj in range(top_ks[ii]):
                            temp_index_ = min(lines.__len__() - 1, jj)
                            # cand_size += float(lines[jj].split(',')[0])
                            # hash_table_hits += float(lines[jj].split(',')[2])
                            cand_size += float(lines[temp_index_].split(',')[0])
                            hash_table_hits += float(lines[temp_index_ ].split(',')[2])
                        # cand_size = float(cand_size)/float(top_ks[ii])
                        cand_size = float(cand_size)
                        hash_table_hits = float(hash_table_hits)
                        cand_size_list.append(cand_size)
                        hash_table_hits_list.append(hash_table_hits)
                        f1.close()

                    overall_result_file = temp_result_dir + 'overall_run_test_' + data_type + '_' + str(
                        cur_dimension) + '_' + str(cardinality) + '_' + cur_ct + '_' + cur_cr + '.txt'

                    f1 = open(overall_result_file, 'r')
                    lines = f1.readlines()
                    recall = []
                    NDCG = []

                    for ll in range(top_ks_length):

                        ttt = lines[2*ll+1].split('\t')[0]

                        recall.append(float(lines[2*ll+1].split('\t')[1]))
                        NDCG.append(float(lines[2*ll+1].split('\t')[2]))

                    spec_string = 'cardinality: ' + str(cardinality) + ', dimension: ' + str(cur_dimension) +\
                                  ', top-k: ' + str(cur_top_o) + ', type: ' \
                                  + cur_type + ', data: ' + data_type + ', compute type: ' + cur_ct + \
                                  ', result type: ' + cur_cr
                    f.write(spec_string + '\n')
                    for ee in range(top_ks_length):
                        f.write(str(recall[ee]) + ', ')
                        f.write(str(NDCG[ee]) + ', ')
                        f.write(str(cand_size_list[ee])  + ', ')
                        f.write(str(obj[ee]) + ', ')
                        f.write(str(hash[ee]) + ', ')
                        f.write('\n')
                    f.write('\n')
                    f.write('\n')

                    for ee in range(top_ks_length):
                        recall_cell = column_row_index(start_recall, column_dist + ee)
                        NDCG_cell = column_row_index(start_NDCG, column_dist + ee)
                        cand_cell = column_row_index(start_cand, column_dist + ee)
                        obj_cell = column_row_index(start_obj, column_dist + ee)
                        hashsize_cell = column_row_index(start_hashsize, column_dist + ee)
                        ws[recall_cell] = float(recall[ee])
                        ws[NDCG_cell] = float(NDCG[ee])
                        ws[cand_cell] = float(cand_size_list[ee])
                        ws[obj_cell] = float(obj[ee])
                        ws[hashsize_cell] = float(hash[ee])
                    for ee in range(top_ks_length):
                        ws_hash_hits[hashsize_cell] = float(hash_table_hits_list[ee])
    wb.save(excel_file)
    wb_hash_hits.save(excel_file_hash_hits)
f.close()
print('Done')