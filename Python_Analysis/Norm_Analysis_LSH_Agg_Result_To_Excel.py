import os
import re
import sys
import math
from openpyxl import load_workbook

result_file_dir = '../H2_ALSH/'
result_type = ['without_threshold']
result_type_excel = ['wo_threshold']

# get obj and hash count from bash_set
obj_hashsize_prefix = 'bash_set_'

# get candidate, recall and NDCG from temp_result
cand_recall_prefix = 'temp_result_'

# card_file_name = ['100k', '200k', '500k', '1M', '1.5M', '2M']
# card = [100000, 200000, 500000, 1000000, 1500000, 2000000]
types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
# data_type = ["anti_correlated", "correlated", "random"]
comp_types = ['opt', 'max', 'uni']

top_ks = [1, 2, 5, 10, 25, 50]


def separate_string(input_string):
    items = []
    match = re.match(r"([a-z]+)([0-9]+)", input_string, re.I)
    if match:
        items = match.groups()
    return items


def column_row_index(input_string, column_dist):
    items = separate_string(input_string)
    column_index = items[0]
    row_index = int(items[1]) + column_dist
    return column_index + str(row_index)


def comp_card_file_name(cur_card_, is_real_life_):
    if is_real_life_ == 1:
        card_file_name_ = cur_card_
    else:
        millnames = ['', 'K', 'M']
        n = float(cur_card_)
        millidx = max(0, min(len(millnames) - 1, int(math.floor(0 if n == 0 else math.log10(abs(n)) / 3))))

        card_file_name_ = '{:.0f}{}'.format(n / 10 ** (3 * millidx), millnames[millidx])
    return card_file_name_





anti_log_optimized_recall = ['B5',  'B6', 'B7', 'B8', 'B9', 'B10']
anti_log_optimized_NDCG = ['C5',  'C6', 'C7', 'C8', 'C9', 'C10']
anti_log_optimized_cand = ['D5',  'D6', 'D7', 'D8', 'D9', 'D10']
anti_log_optimized_obj = ['E5',  'E6', 'E7', 'E8', 'E9', 'E10']
anti_log_optimized_hashsize = ['F5',  'F6', 'F7', 'F8', 'F9', 'F10']


anti_log_optimized_max_recall = ['G5',  'G6', 'G7', 'G8', 'G9', 'G10']
anti_log_optimized_max_NDCG = ['H5',  'H6', 'H7', 'H8', 'H9', 'H10']
anti_log_optimized_max_cand = ['I5',  'I6', 'I7', 'I8', 'I9', 'I10']
anti_log_optimized_max_obj = ['J5',  'J6', 'J7', 'J8', 'J9', 'J10']
anti_log_optimized_max_hashsize = ['K5',  'K6', 'K7', 'K8', 'K9', 'K10']

anti_log_optimized_uni_recall = ['L5',  'L6', 'L7', 'L8', 'L9', 'L10']
anti_log_optimized_uni_NDCG = ['M5',  'M6', 'M7', 'M8', 'M9', 'M10']
anti_log_optimized_uni_cand = ['N5',  'N6', 'N7', 'N8', 'N9', 'N10']
anti_log_optimized_uni_obj = ['O5',  'O6', 'O7', 'O8', 'O9', 'O10']
anti_log_optimized_uni_hashsize = ['P5',  'P6', 'P7', 'P8', 'P9', 'P10']

corr_log_optimized_recall = ['Q5',  'Q6', 'Q7', 'Q8', 'Q9', 'Q10']
corr_log_optimized_NDCG = ['R5',  'R6', 'R7', 'R8', 'R9', 'R10']
corr_log_optimized_cand = ['S5',  'S6', 'S7', 'S8', 'S9', 'S10']
corr_log_optimized_obj = ['T5',  'T6', 'T7', 'T8', 'T9', 'T10']
corr_log_optimized_hashsize = ['U5',  'U6', 'U7', 'U8', 'U9', 'U10']

corr_log_optimized_max_recall = ['V5',  'V6', 'V7', 'V8', 'V9', 'V10']
corr_log_optimized_max_NDCG = ['W5',  'W6', 'W7', 'W8', 'W9', 'W10']
corr_log_optimized_max_cand = ['X5',  'X6', 'X7', 'X8', 'X9', 'X10']
corr_log_optimized_max_obj = ['Y5',  'Y6', 'Y7', 'Y8', 'Y9', 'Y10']
corr_log_optimized_max_hashsize = ['Z5',  'Z6', 'Z7', 'Z8', 'Z9', 'Z10']

corr_log_optimized_uni_recall = ['AA5',  'AA6', 'AA7', 'AA8', 'AA9', 'AA10']
corr_log_optimized_uni_NDCG = ['AB5',  'AB6', 'AB7', 'AB8', 'AB9', 'AB10']
corr_log_optimized_uni_cand = ['AC5',  'AC6', 'AC7', 'AC8', 'AC9', 'AC10']
corr_log_optimized_uni_obj = ['AD5',  'AD6', 'AD7', 'AD8', 'AD9', 'AD10']
corr_log_optimized_uni_hashsize = ['AE5',  'AE6', 'AE7', 'AE8', 'AE9', 'AE10']

rand_log_optimized_recall = ['AF5',  'AF6', 'AF7', 'AF8', 'AF9', 'AF10']
rand_log_optimized_NDCG = ['AG5',  'AG6', 'AG7', 'AG8', 'AG9', 'AG10']
rand_log_optimized_cand = ['AH5',  'AH6', 'AH7', 'AH8', 'AH9', 'AH10']
rand_log_optimized_obj = ['AI5',  'AI6', 'AI7', 'AI8', 'AI9', 'AI10']
rand_log_optimized_hashsize = ['AJ5',  'AJ6', 'AJ7', 'AJ8', 'AJ9', 'AJ10']

rand_log_optimized_max_recall = ['AK5',  'AK6', 'AK7', 'AK8', 'AK9', 'AK10']
rand_log_optimized_max_NDCG = ['AL5',  'AL6', 'AL7', 'AL8', 'AL9', 'AL10']
rand_log_optimized_max_cand = ['AM5',  'AM6', 'AM7', 'AM8', 'AM9', 'AM10']
rand_log_optimized_max_obj = ['AN5',  'AN6', 'AN7', 'AN8', 'AN9', 'AN10']
rand_log_optimized_max_hashsize = ['AO5',  'AO6', 'AO7', 'AO8', 'AO9', 'AO10']

rand_log_optimized_uni_recall = ['AP5',  'AP6', 'AP7', 'AP8', 'AP9', 'AP10']
rand_log_optimized_uni_NDCG = ['AQ5',  'AQ6', 'AQ7', 'AQ8', 'AQ9', 'AQ10']
rand_log_optimized_uni_cand = ['AR5',  'AR6', 'AR7', 'AR8', 'AR9', 'AR10']
rand_log_optimized_uni_obj = ['AS5',  'AS6', 'AS7', 'AS8', 'AS9', 'AS10']
rand_log_optimized_uni_hashsize = ['AT5',  'AT6', 'AT7', 'AT8', 'AT9', 'AT10']


with_without_opt = str(sys.argv[1])
run_index = str(sys.argv[2])
cur_dimension = int(str(sys.argv[2]))
cur_budget = int(str(sys.argv[3]))
cur_bin_count = int(str(sys.argv[4]))
cur_top_o = int(str(sys.argv[5]))
equal_type = str(sys.argv[6])
cur_card = int(str(sys.argv[7]))
cur_dt = str(sys.argv[8])
is_real_life = str(sys.argv[9])

card_file_name = comp_card_file_name(cur_card, is_real_life)
# new excel file here
excel_file = "Aggregation_" + str(cur_dimension) + "D_" + with_without_opt + "_" + run_index + ".xlsx"

excel_file_hash_hits = "Aggregation_" + str(cur_dimension) + "D_" + with_without_opt + "_" + \
                       run_index + "_hash_hits.xlsx"

print("Excel file name: " + str(excel_file))
print("Excel hash hit file name: " + str(excel_file_hash_hits))
# wb = load_workbook(filename=excel_file, data_only=True)
wb = load_workbook(filename=excel_file)
wb_hash_hits = load_workbook(filename=excel_file_hash_hits)

for cr in range(result_type.__len__()):
    cur_cr = result_type[cr]

    # sheet name goes here
    sheet_name = str(cur_dimension) + "D_" + str(card_file_name) + "_" + \
                 result_type_excel[cr] + "_top_" + str(cur_top_o) + '_' + str(cur_bin_count) + '_' + equal_type
    # data_type = ["anti_correlated", "correlated", "random"]
    # comp_types = ['opt', 'max', 'uni']
    for ct in range(comp_types.__len__()):
        cur_ct = comp_types[ct]
        # types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
        for tt in range(types.__len__()):
            cur_type = types[tt]
            column_dist = 0
            if cur_type == 'log':
                column_dist = 0
            elif cur_type == 'log_minus':
                column_dist = 10
            elif cur_type == 'log_plus':
                column_dist = 20
            elif cur_type == 'log_plus_plus':
                column_dist = 30
            else:
                column_dist = 40
            # anti_correlated, correlated, random
            # log, log minus, log plus, log plus plus, uni
            # threshold, or threshold
            start_recall = ''
            start_NDCG = ''
            start_cand = ''
            start_obj = ''
            start_hashsize = ''
            if cur_dt == 'anti_correlated':
                if cur_ct == 'opt':
                    start_recall = anti_log_optimized_recall[0]
                    start_NDCG = anti_log_optimized_NDCG[0]
                    start_cand = anti_log_optimized_cand[0]
                    start_obj = anti_log_optimized_obj[0]
                    start_hashsize = anti_log_optimized_hashsize[0]
                elif cur_ct == 'max':
                    start_recall = anti_log_optimized_max_recall[0]
                    start_NDCG = anti_log_optimized_max_NDCG[0]
                    start_cand = anti_log_optimized_max_cand[0]
                    start_obj = anti_log_optimized_max_obj[0]
                    start_hashsize = anti_log_optimized_max_hashsize[0]
                else:
                    start_recall = anti_log_optimized_uni_recall[0]
                    start_NDCG = anti_log_optimized_uni_NDCG[0]
                    start_cand = anti_log_optimized_uni_cand[0]
                    start_obj = anti_log_optimized_uni_obj[0]
                    start_hashsize = anti_log_optimized_uni_hashsize[0]
            elif cur_dt == 'correlated':
                if cur_ct == 'opt':
                    start_recall = corr_log_optimized_recall[0]
                    start_NDCG = corr_log_optimized_NDCG[0]
                    start_cand = corr_log_optimized_cand[0]
                    start_obj = corr_log_optimized_obj[0]
                    start_hashsize = corr_log_optimized_hashsize[0]
                elif cur_ct == 'max':
                    start_recall = corr_log_optimized_max_recall[0]
                    start_NDCG = corr_log_optimized_max_NDCG[0]
                    start_cand = corr_log_optimized_max_cand[0]
                    start_obj = corr_log_optimized_max_obj[0]
                    start_hashsize = corr_log_optimized_max_hashsize[0]
                else:
                    start_recall = corr_log_optimized_uni_recall[0]
                    start_NDCG = corr_log_optimized_uni_NDCG[0]
                    start_cand = corr_log_optimized_uni_cand[0]
                    start_obj = corr_log_optimized_uni_obj[0]
                    start_hashsize = corr_log_optimized_uni_hashsize[0]
            else:
                if cur_ct == 'opt':
                    start_recall = rand_log_optimized_recall[0]
                    start_NDCG = rand_log_optimized_NDCG[0]
                    start_cand = rand_log_optimized_cand[0]
                    start_obj = rand_log_optimized_obj[0]
                    start_hashsize = rand_log_optimized_hashsize[0]
                elif cur_ct == 'max':
                    start_recall = rand_log_optimized_max_recall[0]
                    start_NDCG = rand_log_optimized_max_NDCG[0]
                    start_cand = rand_log_optimized_max_cand[0]
                    start_obj = rand_log_optimized_max_obj[0]
                    start_hashsize = rand_log_optimized_max_hashsize[0]
                else:
                    start_recall = rand_log_optimized_uni_recall[0]
                    start_NDCG = rand_log_optimized_uni_NDCG[0]
                    start_cand = rand_log_optimized_uni_cand[0]
                    start_obj = rand_log_optimized_uni_obj[0]
                    start_hashsize = rand_log_optimized_uni_hashsize[0]
            obj_file_dir = result_file_dir + obj_hashsize_prefix + str(cur_dimension) + 'D_top' + \
                        str(cur_top_o) + '_budget_' + cur_budget + '_' + cur_type + '_' + str(cur_card) + '/'
            if not os.path.exists(obj_file_dir):
                continue
            # ws = wb[sheet_name]
            print("sheet name: " + str(sheet_name))
            ws = wb.get_sheet_by_name(sheet_name)

            ws_hash_hits = wb_hash_hits.get_sheet_by_name(sheet_name)

            obj_file = obj_file_dir + 'cumsum_hashsize_obj_' + cur_ct + '_' + cur_dt + '_' + \
                       str(cur_dimension) + '_' + str(cur_card) + '_' + with_without_opt + '.txt'

            f1 = open(obj_file, 'r')
            lines = f1.readlines()
            obj_s = lines[0].split(',')
            # print(obj_s)
            hash_s = lines[1].split(',')
            # print(hash_s)

            obj = []
            hash = []
            top_ks_length = -1
            if cur_top_o == 10:
                top_ks_length = 4
            elif cur_top_o == 25:
                top_ks_length = 5
            else:
                top_ks_length = 6
            for oh_index in range(top_ks_length):
                obj.append(int(obj_s[top_ks[oh_index] - 1]))
                hash.append(int(hash_s[top_ks[oh_index] - 1]))
            f1.close()
            temp_result_dir = result_file_dir + cand_recall_prefix + str(cur_dimension) + 'D_top' + \
                        str(cur_top_o) + '_budget_' + cur_budget + '_' + cur_type + '_' + str(cur_card) + '/'

            print("result path: " + str(temp_result_dir))
            if not os.path.exists(temp_result_dir):
                continue

            hash_table_hits_list = []
            cand_size_list = []
            for ii in range(top_ks_length):
                cand_result_file = temp_result_dir + 'run_test_' + cur_dt + '_' + str(cur_dimension) + '_' + \
                               str(cur_card) + '_' + cur_ct + '_' + cur_cr + '_top_' + str(top_ks[ii]) + '_candidate_size.txt'
                cand_size = 0
                hash_table_hits = 0
                f1 = open(cand_result_file, 'r')
                lines = f1.readlines()
                for jj in range(min(top_ks[ii], lines.__len__())):
                    if lines[jj].split(',').__len__() < 1:
                        cand_size += 0
                        hash_table_hits += 0
                    else:
                        cand_size += float(lines[jj].split(',')[0])
                        hash_table_hits += float(lines[jj].split(',')[2])
                # cand_size = float(cand_size)/float(top_ks[ii])
                cand_size = float(cand_size)
                hash_table_hits = float(hash_table_hits)
                cand_size_list.append(cand_size)
                hash_table_hits_list.append(hash_table_hits)
                f1.close()

            overall_result_file = temp_result_dir + 'overall_run_test_' + cur_dt + '_' + str(
                cur_dimension) + '_' + str(cur_card) + '_' + cur_ct + '_' + cur_cr + '.txt'
            # print(overall_result_file)
            f1 = open(overall_result_file, 'r')
            lines = f1.readlines()
            recall = []
            NDCG = []
            # for ll in range(top_ks.__len__()):
            for ll in range(top_ks_length):
                ttt = lines[2*ll+1].split('\t')[0]
                recall.append(float(lines[2*ll+1].split('\t')[1]))
                NDCG.append(float(lines[2*ll+1].split('\t')[2]))

            # cardinality, dimension, budget, top-10, log, anti_correlated, opt
            spec_string = 'cardinality: ' + str(cur_card) + ', dimension: ' + str(cur_dimension) +\
                          ',  budget: ' + str(cur_budget) + ', top-k: ' + str(cur_top_o) + ', type: ' \
                          + cur_type + ', data: ' + cur_dt + ', compute type: ' + cur_ct + \
                          ', result type: ' + cur_cr
            for ee in range(top_ks_length):
                recall_cell = column_row_index(start_recall, column_dist + ee)
                NDCG_cell = column_row_index(start_NDCG, column_dist + ee)
                cand_cell = column_row_index(start_cand, column_dist + ee)
                obj_cell = column_row_index(start_obj, column_dist + ee)
                hashsize_cell = column_row_index(start_hashsize, column_dist + ee)
                ws[recall_cell] = float(recall[ee])
                ws[NDCG_cell] = float(NDCG[ee])
                ws[cand_cell] = float(cand_size_list[ee])
                ws[obj_cell] = float(obj[ee])
                ws[hashsize_cell] = float(hash[ee])
            for ee in range(top_ks_length):
                ws_hash_hits[hashsize_cell] = float(hash_table_hits_list[ee])
wb.save(excel_file)
wb_hash_hits.save(excel_file_hash_hits)
print('Done')