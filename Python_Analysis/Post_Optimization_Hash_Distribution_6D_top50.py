import math
import numpy as np
from openpyxl import load_workbook
import random
import re


def separate_string(input_string):
    items = []
    match = re.match(r"([a-z]+)([0-9]+)", input_string, re.I)
    if match:
        items = match.groups()
    return items


def column_row_index(input_string, column_dist):
    items = separate_string(input_string)
    column_index = items[0]
    row_index = int(items[1]) + column_dist
    return column_index + str(row_index)


def compute_weights(data_list_, top_m_):
    data_list_ = np.asarray(data_list_)
    contribution_list = []
    weight_list_ = []
    data_list_cumsum = np.cumsum(data_list_)
    for i in range(top_m_):
        c_h = data_list_[i]
        temp_contribution = 0
        for j in range(i, top_m_):
            denominator = data_list_cumsum[j]
            temp_contribution = 1.0 * temp_contribution + (1.0 * c_h/denominator)
        contribution_list.append(temp_contribution)
    sum_contribution = sum(contribution_list)
    for i in range(top_m_):
        temp_weight = (1.0 * contribution_list[i])/sum_contribution
        weight_list_.append(temp_weight)
    return weight_list_


def post_optimization_opt(collision_probility_, weight_list_, total_error_, data_list_, K_List_, L_List_, hash_used_, hash_budget_):
    smallest = min(data_list_)
    smallest_index = data_list_.index(min(data_list_))

    # total_error = 0.412502654
    total_error_gain = 0
    flag = False
    while hash_used_ + smallest <= hash_budget_:
        # the condition of improvement is to check if the total error rate drops

        # loop through, keep track of hash-relocation and error rate drop
        # find the biggest error gain, while keep hash-resources constraints
        # update hash_used, LList
        delta_error_list = []
        for i in range(len(data_list_)):
            cur_k = K_List_[i]
            cur_l = L_List_[i]
            c_old = weight_list_[i] * math.pow((1 - math.pow(collision_probility_, cur_k)), cur_l)

            # each time increment hash layer by 1
            c_new = weight_list_[i] * math.pow((1 - math.pow(collision_probility_, cur_k)), (cur_l+1))
            delta_error_list.append((c_old - c_new))

        # sort and check each delta_error
        delta_error_list = np.asarray(delta_error_list)
        # sort in descending order
        sorted_index = delta_error_list.argsort()[::-1][:len(delta_error_list)]
        sorted_pivot = 0
        while sorted_pivot < len(sorted_index):
            cur_index = sorted_index[sorted_pivot]

            # each time increment hash layer by 1
            # temp_hash_used = hash_used + (LList[cur_index] + 1) * data_list[cur_index]
            temp_hash_used = hash_used_ + data_list_[cur_index]
            if temp_hash_used <= hash_budget_:
                L_List_[cur_index] = L_List_[cur_index] + 1
                hash_used_ = hash_used_ + data_list_[cur_index]
                break
            sorted_pivot = sorted_pivot + 1
    total_error = 0
    total_hash_used = 0
    for i in range(len(L_List_)):
        cur_k = K_List_[i]
        cur_l = L_List_[i]
        total_error = total_error + weight_list_[i] * math.pow((1 - math.pow(collision_probility_, cur_k)), cur_l)
        total_hash_used = total_hash_used + data_list_[i] * cur_l
    # print("Updated total error: " + str(total_error))
    # print("total hash used: " + str(total_hash_used))
    # print("Optimized approach done")
    return L_List_


####################################################################################
# for uniformly distributed approach
# uniformly pick one that fits the current hash budget instead of the one minimizing total error rate
def post_optimization_uni(data_list_, L_List_, hash_used_, hash_budget_):
    flag = False
    smallest = min(data_list_)
    data_list_ = np.asarray(data_list_)
    while hash_used_ + smallest <= hash_budget_:
        # sort in descending order
        sorted_index = data_list_.argsort()[::-1][:len(data_list_)]

        temp_pivot_list = []
        # first find all the allow current hash re-allocation
        for i in range(len(sorted_index)):
            temp_pivot_index = sorted_index[i]
            if hash_used_ + data_list_[temp_pivot_index] <= hash_budget_:
                temp_pivot_list.append(temp_pivot_index)
        # randomly pick one from those
        cur_pivot = random.choice(temp_pivot_list)
        # update L_Uni_List, hash_used
        L_List_[cur_pivot] = L_List_[cur_pivot] + 1
        hash_used_ = hash_used_ + data_list_[cur_pivot]
    return L_List_



def compute_index(list_start_, top_m_):
    cell_index_ = []
    for start_ in list_start_:
        cell_index_.append(start_)
        ret_index = separate_string(start_)
        ret_ = ret_index[0] + str(int(ret_index[1]) + top_m_ - 1)
        cell_index_.append(ret_)
    return cell_index_


def computer_used_resource_index(start_, top_m_, count_):
    start_index_ = separate_string(start_)
    cell_index_ = []
    cell_index_.append(start_)
    first_ = start_index_[0] + str(int(start_index_[1]) + top_m_ - 1 + 8)
    cell_index_.append(first_)

    for i in range(2, count_):
        first_index = separate_string(first_)
        first_ = first_index[0] + str(int(first_index[1]) + top_m_ - 1 + 7)
        cell_index_.append(first_)
    return cell_index_


####################################################################################
# types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
Data_Types = ['anti_correlated', 'correlated', 'random']
types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
# Data_Types = ['anti_correlated']
top_m_cell_anti = 'E1'
top_m_cell_corr = 'V1'
top_m_cell_rand = 'AL1'


# need to update budget cell per data type
budget_cell_anti = 'B4'
budget_cell_corr = 'S4'
budget_cell_rand = 'AI4'


cardinality_cell = 'B2'
top_m_cardinality_anti = 0
top_m_cardinality_corr = 0
top_m_cardinality_random = 0

####################################################
cur_topm = 28
data_list_start = 'J6'
k_ranges_start = 'E6'
l_ranges_opt_start = 'F6'
l_ranges_max_start = 'G6'
l_ranges_uni_start = 'H6'
hash_used_opt_cells_start = 'I34'
hash_used_uni_cells_start = 'O34'

data_anti_list_28 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_anti_28 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_anti_28 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_anti_28 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_anti_28 = compute_index(computer_used_resource_index(l_ranges_uni_start, cur_topm, 5), cur_topm)
hash_used_anti_opt_cells_28 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_anti_uni_cells_28 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)


cur_topm = 33
data_list_start = 'AA6'
k_ranges_start = 'V6'
l_ranges_opt_start = 'W6'
l_ranges_max_start = 'X6'
l_ranges_uni_start = 'Y6'
hash_used_opt_cells_start = 'Z39'
hash_used_uni_cells_start = 'AF39'

data_corr_list_33 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_corr_33 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_corr_33 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_corr_33 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_corr_33 = compute_index(computer_used_resource_index(l_ranges_uni_start, cur_topm, 5), cur_topm)
hash_used_corr_opt_cells_33 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_corr_uni_cells_33 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)


cur_topm = 27
data_list_start = 'AQ6'
k_ranges_random_start = 'AL6'
l_ranges_opt_random_start = 'AM6'
l_ranges_max_random_start = 'AN6'
l_ranges_uni_random_start = 'AO6'
hash_used_cells_start = 'AP33'
hash_used_uni_cells_13_start = 'AV33'

data_random_list_27 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_random_27 = compute_index(computer_used_resource_index(k_ranges_random_start, cur_topm, 5), cur_topm)
l_ranges_opt_random_27 = compute_index(computer_used_resource_index(l_ranges_opt_random_start, cur_topm, 5), cur_topm)
l_ranges_max_random_27 = compute_index(computer_used_resource_index(l_ranges_max_random_start, cur_topm, 5), cur_topm)
l_ranges_uni_random_27 = compute_index(computer_used_resource_index(l_ranges_uni_random_start, cur_topm, 5), cur_topm)
hash_used_rand_opt_cells_27 = computer_used_resource_index(hash_used_cells_start, cur_topm, 5)
hash_used_rand_uni_cells_27 = computer_used_resource_index(hash_used_uni_cells_13_start, cur_topm, 5)

####################################################
cur_topm = 34
data_list_start = 'J6'
k_ranges_start = 'E6'
l_ranges_opt_start = 'F6'
l_ranges_max_start = 'G6'
l_ranges_uni_astart = 'H6'
hash_used_opt_cells_start = 'I40'
hash_used_uni_cells_start = 'O40'

data_anti_list_34 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_anti_34 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_anti_34 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_anti_34 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_anti_34 = compute_index(computer_used_resource_index(l_ranges_uni_astart, cur_topm, 5), cur_topm)
hash_used_anti_opt_cells_34 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_anti_uni_cells_34 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)


cur_topm = 40
data_list_start = 'AA6'
k_ranges_start = 'V6'
l_ranges_opt_start = 'W6'
l_ranges_max_start = 'X6'
l_ranges_uni_start = 'Y6'
hash_used_opt_cells_start = 'Z46'
hash_used_uni_cells_start = 'AF46'


data_corr_list_40 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_corr_40 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_corr_40 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_corr_40 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_corr_40 = compute_index(computer_used_resource_index(l_ranges_uni_start, cur_topm, 5), cur_topm)
hash_used_corr_opt_cells_40 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_corr_uni_cells_40 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)

cur_topm = 32
data_list_start = 'AQ6'
k_ranges_random_start = 'AL6'
l_ranges_opt_random_start = 'AM6'
l_ranges_max_random_start = 'AN6'
l_ranges_uni_random_start = 'AO6'
hash_used_cells_start = 'AP38'
hash_used_uni_cells_13_start = 'AV38'

data_random_list_32 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_random_32 = compute_index(computer_used_resource_index(k_ranges_random_start, cur_topm, 5), cur_topm)
l_ranges_opt_random_32 = compute_index(computer_used_resource_index(l_ranges_opt_random_start, cur_topm, 5), cur_topm)
l_ranges_max_random_32 = compute_index(computer_used_resource_index(l_ranges_max_random_start, cur_topm, 5), cur_topm)
l_ranges_uni_random_32 = compute_index(computer_used_resource_index(l_ranges_uni_random_start, cur_topm, 5), cur_topm)
hash_used_rand_opt_cells_32 = computer_used_resource_index(hash_used_cells_start, cur_topm, 5)
hash_used_rand_uni_cells_32 = computer_used_resource_index(hash_used_uni_cells_13_start, cur_topm, 5)
####################################################

cur_topm = 41
data_list_start = 'J6'
k_ranges_start = 'E6'
l_ranges_opt_start = 'F6'
l_ranges_max_start = 'G6'
l_ranges_uni_astart = 'H6'
hash_used_opt_cells_start = 'I47'
hash_used_uni_cells_start = 'O47'

data_anti_list_41 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_anti_41 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_anti_41 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_anti_41 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_anti_41 = compute_index(computer_used_resource_index(l_ranges_uni_astart, cur_topm, 5), cur_topm)
hash_used_anti_opt_cells_41 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_anti_uni_cells_41 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)


cur_topm = 49
data_list_start = 'AA6'
k_ranges_start = 'V6'
l_ranges_opt_start = 'W6'
l_ranges_max_start = 'X6'
l_ranges_uni_start = 'Y6'
hash_used_opt_cells_start = 'Z55'
hash_used_uni_cells_start = 'AF55'


data_corr_list_49 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_corr_49 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_corr_49 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_corr_49 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_corr_49 = compute_index(computer_used_resource_index(l_ranges_uni_start, cur_topm, 5), cur_topm)
hash_used_corr_opt_cells_49 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_corr_uni_cells_49 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)


cur_topm = 40
data_list_start = 'AQ6'
k_ranges_random_start = 'AL6'
l_ranges_opt_random_start = 'AM6'
l_ranges_max_random_start = 'AN6'
l_ranges_uni_random_start = 'AO6'
hash_used_cells_start = 'AP46'
hash_used_uni_cells_13_start = 'AV46'

data_random_list_40 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_random_40 = compute_index(computer_used_resource_index(k_ranges_random_start, cur_topm, 5), cur_topm)
l_ranges_opt_random_40 = compute_index(computer_used_resource_index(l_ranges_opt_random_start, cur_topm, 5), cur_topm)
l_ranges_max_random_40 = compute_index(computer_used_resource_index(l_ranges_max_random_start, cur_topm, 5), cur_topm)
l_ranges_uni_random_40 = compute_index(computer_used_resource_index(l_ranges_uni_random_start, cur_topm, 5), cur_topm)
hash_used_rand_opt_cells_40 = computer_used_resource_index(hash_used_cells_start, cur_topm, 5)
hash_used_rand_uni_cells_40 = computer_used_resource_index(hash_used_uni_cells_13_start, cur_topm, 5)
####################################################

cur_topm = 38
data_list_start = 'J6'
k_ranges_start = 'E6'
l_ranges_opt_start = 'F6'
l_ranges_max_start = 'G6'
l_ranges_uni_astart = 'H6'
hash_used_opt_cells_start = 'I44'
hash_used_uni_cells_start = 'O44'

data_anti_list_38 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_anti_38 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_anti_38 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_anti_38 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_anti_38 = compute_index(computer_used_resource_index(l_ranges_uni_astart, cur_topm, 5), cur_topm)
hash_used_anti_opt_cells_38 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_anti_uni_cells_38 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)


cur_topm = 45
data_list_start = 'AA6'
k_ranges_start = 'V6'
l_ranges_opt_start = 'W6'
l_ranges_max_start = 'X6'
l_ranges_uni_start = 'Y6'
hash_used_opt_cells_start = 'Z51'
hash_used_uni_cells_start = 'AF51'


data_corr_list_45 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_corr_45 = compute_index(computer_used_resource_index(k_ranges_start, cur_topm, 5), cur_topm)
l_ranges_opt_corr_45 = compute_index(computer_used_resource_index(l_ranges_opt_start, cur_topm, 5), cur_topm)
l_ranges_max_corr_45 = compute_index(computer_used_resource_index(l_ranges_max_start, cur_topm, 5), cur_topm)
l_ranges_uni_corr_45 = compute_index(computer_used_resource_index(l_ranges_uni_start, cur_topm, 5), cur_topm)
hash_used_corr_opt_cells_45 = computer_used_resource_index(hash_used_opt_cells_start, cur_topm, 5)
hash_used_corr_uni_cells_45 = computer_used_resource_index(hash_used_uni_cells_start, cur_topm, 5)


cur_topm = 37
data_list_start = 'AQ6'
k_ranges_random_start = 'AL6'
l_ranges_opt_random_start = 'AM6'
l_ranges_max_random_start = 'AN6'
l_ranges_uni_random_start = 'AO6'
hash_used_cells_start = 'AP43'
hash_used_uni_cells_13_start = 'AV43'

data_random_list_37 = compute_index(computer_used_resource_index(data_list_start, cur_topm, 5), cur_topm)
k_ranges_random_37 = compute_index(computer_used_resource_index(k_ranges_random_start, cur_topm, 5), cur_topm)
l_ranges_opt_random_37 = compute_index(computer_used_resource_index(l_ranges_opt_random_start, cur_topm, 5), cur_topm)
l_ranges_max_random_37 = compute_index(computer_used_resource_index(l_ranges_max_random_start, cur_topm, 5), cur_topm)
l_ranges_uni_random_37 = compute_index(computer_used_resource_index(l_ranges_uni_random_start, cur_topm, 5), cur_topm)
hash_used_rand_opt_cells_37 = computer_used_resource_index(hash_used_cells_start, cur_topm, 5)
hash_used_rand_uni_cells_37 = computer_used_resource_index(hash_used_uni_cells_13_start, cur_topm, 5)
####################################################

collision_probility = 0.75
total_error = 0
# K_log_list_start_cell = ['E6', 'V6', 'AL6']
# K_log_minus_list_start_cell = ['']
# K_log_plus_list_start_cell = []
# K_log_plus_plus_list_start_cell = []
# K_log_uni_list_start_cell = []

####################################################################################

# dimensions = [2, 3, 4, 5, 6, 7]
dimensions = [6]
excel_file_dir = './'

# for each excel file
for i in range(len(dimensions)):
    cur_d = dimensions[i]
    # excel_file_name = excel_file_dir + 'Checkpoint_Result_Nov_26_' + str(cur_d) + 'D_test.xlsx'
    excel_file_name = excel_file_dir + str(cur_d) + 'D_top50_075_all_before.xlsx'
    wb = load_workbook(filename=excel_file_name, data_only=True)
    wb1 = load_workbook(filename=excel_file_name)
    wss = wb.get_sheet_names()

    for wwss in wss:
        print(wwss)
        ws = wb.get_sheet_by_name(wwss)
        ws1 = wb1.get_sheet_by_name(wwss)
        top_m_anti = ws[top_m_cell_anti].value
        top_m_corr = ws[top_m_cell_corr].value
        top_m_rand = ws[top_m_cell_rand].value
        # ws = wb.get_sheet_by_name(wss[0])
        # ws1 = wb1.get_sheet_by_name(wss[0])
        # top_m = ws[top_m_cell].value
        # print("Sheet name: " + str(ws) + " top-m : " + str(top_m))
        hash_budget_anti = ws[budget_cell_anti].value
        hash_budget_corr = ws[budget_cell_corr].value
        hash_budget_rand = ws[budget_cell_rand].value
        total_cardinality = ws[cardinality_cell].value
        if top_m_anti == 28:
            top_m_cardinality_anti_cell = 'J34'
            top_m_cardinality_anti = ws[top_m_cardinality_anti_cell].value
            hash_used_anti_opt_cells = hash_used_anti_opt_cells_28
            hash_used_anti_uni_cells = hash_used_anti_uni_cells_28
            data_anti_list = data_anti_list_28
            k_ranges_anti = k_ranges_anti_28
            l_ranges_opt_anti = l_ranges_opt_anti_28
            l_ranges_max_anti = l_ranges_max_anti_28
            l_ranges_uni_anti = l_ranges_uni_anti_28

            top_m_cardinality_corr_cell = 'AA39'
            top_m_cardinality_corr = ws[top_m_cardinality_corr_cell].value
            hash_used_corr_opt_cells = hash_used_corr_opt_cells_33
            hash_used_corr_uni_cells = hash_used_corr_uni_cells_33
            data_corr_list = data_corr_list_33
            k_ranges_corr = k_ranges_corr_33
            l_ranges_opt_corr = l_ranges_opt_corr_33
            l_ranges_max_corr = l_ranges_max_corr_33
            l_ranges_uni_corr = l_ranges_uni_corr_33

            top_m_cardinality_random_cell = 'AQ33'
            top_m_cardinality_random = ws[top_m_cardinality_random_cell].value
            hash_used_rand_opt_cells = hash_used_rand_opt_cells_27
            hash_used_rand_uni_cells = hash_used_rand_uni_cells_27
            data_random_list = data_random_list_27
            k_ranges_random = k_ranges_random_27
            l_ranges_opt_random = l_ranges_opt_random_27
            l_ranges_max_random = l_ranges_max_random_27
            l_ranges_uni_random = l_ranges_uni_random_27
        elif top_m_anti == 34:
            top_m_cardinality_anti_cell = 'J40'
            top_m_cardinality_anti = ws[top_m_cardinality_anti_cell].value
            hash_used_anti_opt_cells = hash_used_anti_opt_cells_34
            hash_used_anti_uni_cells = hash_used_anti_uni_cells_34
            data_anti_list = data_anti_list_34
            k_ranges_anti = k_ranges_anti_34
            l_ranges_opt_anti = l_ranges_opt_anti_34
            l_ranges_max_anti = l_ranges_max_anti_34
            l_ranges_uni_anti = l_ranges_uni_anti_34

            top_m_cardinality_corr_cell = 'AA46'
            top_m_cardinality_corr = ws[top_m_cardinality_corr_cell].value
            hash_used_corr_opt_cells = hash_used_corr_opt_cells_40
            hash_used_corr_uni_cells = hash_used_corr_uni_cells_40
            data_corr_list = data_corr_list_40
            k_ranges_corr = k_ranges_corr_40
            l_ranges_opt_corr = l_ranges_opt_corr_40
            l_ranges_max_corr = l_ranges_max_corr_40
            l_ranges_uni_corr = l_ranges_uni_corr_40

            top_m_cardinality_random_cell = 'AQ38'
            top_m_cardinality_random = ws[top_m_cardinality_random_cell].value
            hash_used_rand_opt_cells = hash_used_rand_opt_cells_32
            hash_used_rand_uni_cells = hash_used_rand_uni_cells_32
            data_random_list = data_random_list_32
            k_ranges_random = k_ranges_random_32
            l_ranges_opt_random = l_ranges_opt_random_32
            l_ranges_max_random = l_ranges_max_random_32
            l_ranges_uni_random = l_ranges_uni_random_32
        elif top_m_anti == 41:
            top_m_cardinality_anti_cell = 'J44'
            top_m_cardinality_anti = ws[top_m_cardinality_anti_cell].value
            hash_used_anti_opt_cells = hash_used_anti_opt_cells_41
            hash_used_anti_uni_cells = hash_used_anti_uni_cells_41
            data_anti_list = data_anti_list_41
            k_ranges_anti = k_ranges_anti_41
            l_ranges_opt_anti = l_ranges_opt_anti_41
            l_ranges_max_anti = l_ranges_max_anti_41
            l_ranges_uni_anti = l_ranges_uni_anti_41

            top_m_cardinality_corr_cell = 'AA51'
            top_m_cardinality_corr = ws[top_m_cardinality_corr_cell].value
            hash_used_corr_opt_cells = hash_used_corr_opt_cells_49
            hash_used_corr_uni_cells = hash_used_corr_uni_cells_49
            data_corr_list = data_corr_list_49
            k_ranges_corr = k_ranges_corr_49
            l_ranges_opt_corr = l_ranges_opt_corr_49
            l_ranges_max_corr = l_ranges_max_corr_49
            l_ranges_uni_corr = l_ranges_uni_corr_49

            top_m_cardinality_random_cell = 'AQ43'
            top_m_cardinality_random = ws[top_m_cardinality_random_cell].value
            hash_used_rand_opt_cells = hash_used_rand_opt_cells_40
            hash_used_rand_uni_cells = hash_used_rand_uni_cells_40
            data_random_list = data_random_list_40
            k_ranges_random = k_ranges_random_40
            l_ranges_opt_random = l_ranges_opt_random_40
            l_ranges_max_random = l_ranges_max_random_40
            l_ranges_uni_random = l_ranges_uni_random_40
        else: # top_m_anti == 38:
            top_m_cardinality_anti_cell = 'J47'
            top_m_cardinality_anti = ws[top_m_cardinality_anti_cell].value
            hash_used_anti_opt_cells = hash_used_anti_opt_cells_38
            hash_used_anti_uni_cells = hash_used_anti_uni_cells_38
            data_anti_list = data_anti_list_38
            k_ranges_anti = k_ranges_anti_38
            l_ranges_opt_anti = l_ranges_opt_anti_38
            l_ranges_max_anti = l_ranges_max_anti_38
            l_ranges_uni_anti = l_ranges_uni_anti_38

            top_m_cardinality_corr_cell = 'AA55'
            top_m_cardinality_corr = ws[top_m_cardinality_corr_cell].value
            hash_used_corr_opt_cells = hash_used_corr_opt_cells_45
            hash_used_corr_uni_cells = hash_used_corr_uni_cells_45
            data_corr_list = data_corr_list_45
            k_ranges_corr = k_ranges_corr_45
            l_ranges_opt_corr = l_ranges_opt_corr_45
            l_ranges_max_corr = l_ranges_max_corr_45
            l_ranges_uni_corr = l_ranges_uni_corr_45

            top_m_cardinality_random_cell = 'AQ46'
            top_m_cardinality_random = ws[top_m_cardinality_random_cell].value
            hash_used_rand_opt_cells = hash_used_rand_opt_cells_37
            hash_used_rand_uni_cells = hash_used_rand_uni_cells_37
            data_random_list = data_random_list_37
            k_ranges_random = k_ranges_random_37
            l_ranges_opt_random = l_ranges_opt_random_37
            l_ranges_max_random = l_ranges_max_random_37
            l_ranges_uni_random = l_ranges_uni_random_37

        data_anti = []
        data_corr = []
        data_random = []
        # within each excel file, for each data type
        # for j in range(len(Data_Types)):
        # read data list
        data_anti_list_start = data_anti_list[0]
        data_anti_list_end = data_anti_list[1]

        data_corr_list_start = data_corr_list[0]
        data_corr_list_end = data_corr_list[1]

        data_random_list_start = data_random_list[0]
        data_random_list_end = data_random_list[1]

        for columns in ws[data_anti_list_start: data_anti_list_end]:
            for cell in columns:
                data_anti.append(cell.value)

        for columns in ws[data_corr_list_start: data_corr_list_end]:
            for cell in columns:
                data_corr.append(cell.value)

        for columns in ws[data_random_list_start: data_random_list_end]:
            for cell in columns:
                data_random.append(cell.value)

        weight_anti = compute_weights(data_anti, len(data_anti))
        weight_corr = compute_weights(data_corr, len(data_corr))
        weight_random = compute_weights(data_random, len(data_random))

        # for each type, log, log_minus, log_plus, etc
        for jj in range(types.__len__()):
            # read k and l
            type_name = types[jj]
            # print(Data_Types[j] + ' ' + type_name)
            start = 2 * jj
            end = 2 * jj + 1
            k_anti = []
            for columns in ws[k_ranges_anti[start]: k_ranges_anti[end]]:
                for cell in columns:
                    k_anti.append(cell.value)

            l_anti_opt = []
            for columns in ws[l_ranges_opt_anti[start]: l_ranges_opt_anti[end]]:
                for cell in columns:
                    l_anti_opt.append(cell.value)

            l_anti_max = []
            for columns in ws[l_ranges_max_anti[start]: l_ranges_max_anti[end]]:
                for cell in columns:
                    l_anti_max.append(cell.value)

            l_anti_uni = []
            for columns in ws[l_ranges_uni_anti[start]: l_ranges_uni_anti[end]]:
                for cell in columns:
                    l_anti_uni.append(cell.value)

            # read data type correlated
            k_corr = []
            for columns in ws[k_ranges_corr[start]: k_ranges_corr[end]]:
                for cell in columns:
                    k_corr.append(cell.value)

            l_corr_opt = []
            for columns in ws[l_ranges_opt_corr[start]: l_ranges_opt_corr[end]]:
                for cell in columns:
                    l_corr_opt.append(cell.value)

            l_corr_max = []
            for columns in ws[l_ranges_max_corr[start]: l_ranges_max_corr[end]]:
                for cell in columns:
                    l_corr_max.append(cell.value)

            l_corr_uni = []
            for columns in ws[l_ranges_uni_corr[start]: l_ranges_uni_corr[end]]:
                for cell in columns:
                    l_corr_uni.append(cell.value)

            # read data type random
            k_random = []
            for columns in ws[k_ranges_random[start]: k_ranges_random[end]]:
                for cell in columns:
                    k_random.append(cell.value)

            l_random_opt = []
            for columns in ws[l_ranges_opt_random[start]: l_ranges_opt_random[end]]:
                for cell in columns:
                    l_random_opt.append(cell.value)

            l_random_max = []
            for columns in ws[l_ranges_max_random[start]: l_ranges_max_random[end]]:
                for cell in columns:
                    l_random_max.append(cell.value)

            l_random_uni = []
            for columns in ws[l_ranges_uni_random[start]: l_ranges_uni_random[end]]:
                for cell in columns:
                    l_random_uni.append(cell.value)

            hash_used_anti_opt_cell = hash_used_anti_opt_cells[jj]
            hash_used_corr_opt_cell = hash_used_corr_opt_cells[jj]
            hash_used_random_opt_cell = hash_used_rand_opt_cells[jj]

            hash_used_anti_opt = ws[hash_used_anti_opt_cell].value
            hash_used_corr_opt = ws[hash_used_corr_opt_cell].value
            hash_used_random_opt = ws[hash_used_random_opt_cell].value

            # update LList
            l_anti_opt = post_optimization_opt(collision_probility, weight_anti, total_error, data_anti, k_anti, l_anti_opt,
                                           hash_used_anti_opt, hash_budget_anti)
            l_corr_opt = post_optimization_opt(collision_probility, weight_corr, total_error, data_corr, k_corr, l_corr_opt,
                                           hash_used_corr_opt, hash_budget_corr)
            l_random_opt = post_optimization_opt(collision_probility, weight_random, total_error, data_random, k_anti, l_random_opt,
                                           hash_used_random_opt, hash_budget_rand)

            hash_used_anti_uni_cell = hash_used_anti_uni_cells[jj]
            hash_used_corr_uni_cell = hash_used_corr_uni_cells[jj]
            hash_used_random_uni_cell = hash_used_rand_uni_cells[jj]

            hash_used_anti_uni = ws[hash_used_anti_uni_cell].value
            hash_used_corr_uni = ws[hash_used_corr_uni_cell].value
            hash_used_random_uni = ws[hash_used_random_uni_cell].value

            l_anti_uni = post_optimization_uni(data_anti, l_anti_uni, hash_used_anti_uni, hash_budget_anti)
            l_corr_uni = post_optimization_uni(data_corr, l_corr_uni, hash_used_corr_uni, hash_budget_corr)
            l_random_uni = post_optimization_uni(data_random, l_random_uni, hash_used_random_uni, hash_budget_rand)

            # write udpate LList back to excel file
            for kk in range(len(l_anti_opt)):
                cur_cell_anti_opt = column_row_index(l_ranges_opt_anti[start], kk)
                cur_cell_anti_uni = column_row_index(l_ranges_uni_anti[start], kk)
                ws1[cur_cell_anti_opt] = l_anti_opt[kk]
                ws1[cur_cell_anti_uni] = l_anti_uni[kk]

            for kk in range(len(l_corr_opt)):
                cur_cell_corr_opt = column_row_index(l_ranges_opt_corr[start], kk)
                cur_cell_corr_uni = column_row_index(l_ranges_uni_corr[start], kk)
                ws1[cur_cell_corr_opt] = l_corr_opt[kk]
                ws1[cur_cell_corr_uni] = l_corr_uni[kk]

            for kk in range(len(l_random_opt)):
                cur_cell_random_opt = column_row_index(l_ranges_opt_random[start], kk)
                cur_cell_random_uni = column_row_index(l_ranges_uni_random[start], kk)
                ws1[cur_cell_random_uni] = l_random_uni[kk]
                ws1[cur_cell_random_opt] = l_random_opt[kk]
    wb1.save(excel_file_name)

print("All done")