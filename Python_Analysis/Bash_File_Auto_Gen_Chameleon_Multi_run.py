# script to generate .sh file for execution
# bash file includes execution of Ground Truth, LSH Scheme and Overall Performance

import math
import os
from openpyxl import load_workbook

k_ranges_anti_10 = ['E6',  'E15', 'E21', 'E30', 'E37', 'E46', 'E51', 'E60', 'E68', 'E77']
l_ranges_opt_anti_10 = ['F6', 'F15', 'F21', 'F30', 'F37', 'F46', 'F51', 'F60', 'F68', 'F77']
l_ranges_max_anti_10 = ['G6', 'G15', 'G21', 'G30', 'G37', 'G46', 'G51', 'G60', 'G68', 'G77']
l_ranges_uni_anti_10 = ['H6', 'H15', 'H21', 'H30', 'H37', 'H46', 'H51', 'H60', 'H68', 'H77']

k_ranges_corr_10 = ['V6', 'V15', 'V21', 'V30', 'V37', 'V46', 'V51', 'V60', 'V68', 'V77']
l_ranges_opt_corr_10 = ['W6', 'W15', 'W21', 'W30', 'W37', 'W46', 'W51', 'W60', 'W68', 'W77']
l_ranges_max_corr_10 = ['X6', 'X15', 'X21', 'X30', 'X37', 'X46', 'X51', 'X60', 'X68', 'X77']
l_ranges_uni_corr_10 = ['Y6', 'Y15', 'Y21', 'Y30', 'Y37', 'Y46', 'Y51', 'Y60', 'Y68', 'Y77']

k_ranges_random_10 = ['AL6', 'AL15', 'AL21', 'AL30', 'AL37', 'AL46', 'AL51', 'AL60', 'AL68', 'AL77']
l_ranges_opt_random_10 = ['AM6', 'AM15', 'AM21', 'AM30', 'AM37', 'AM46', 'AM51', 'AM60', 'AM68', 'AM77']
l_ranges_max_random_10 = ['AN6', 'AN15', 'AN21', 'AN30', 'AN37', 'AN46', 'AN51', 'AN60', 'AN68', 'AN77']
l_ranges_uni_random_10 = ['AO6', 'AO15', 'AO21', 'AO30', 'AO37', 'AO46', 'AO51', 'AO60', 'AO68', 'AO77']

k_ranges_anti_25 = ['E6', 'E30', 'E38', 'E62', 'E69', 'E93', 'E100', 'E124', 'E131', 'E155']
l_ranges_opt_anti_25 = ['F6', 'F30', 'F38', 'F62', 'F69', 'F93', 'F100', 'F124', 'F131', 'F155']
l_ranges_max_anti_25 = ['G6', 'G30', 'G38', 'G62', 'G69', 'G93', 'G100', 'G124', 'G131', 'G155']
l_ranges_uni_anti_25 = ['H6', 'H30', 'H38', 'H62', 'H69', 'H93', 'H100', 'H124', 'H131', 'H155']

k_ranges_corr_25 = ['V6', 'V30', 'V38', 'V62', 'V69', 'V93', 'V100', 'V124', 'V131', 'V155']
l_ranges_opt_corr_25 = ['W6', 'W30', 'W38', 'W62', 'W69', 'W93', 'W100', 'W124', 'W131', 'W155']
l_ranges_max_corr_25 = ['X6', 'X30', 'X38', 'X62', 'X69', 'X93', 'X100', 'X124', 'X131', 'X155']
l_ranges_uni_corr_25 = ['Y6', 'Y30', 'Y38', 'Y62', 'Y69', 'Y93', 'Y100', 'Y124', 'Y131', 'Y155']

k_ranges_random_25 = ['AL6', 'AL30', 'AL38', 'AL62', 'AL69', 'AL93', 'AL100', 'AL124', 'AL131', 'AL155']
l_ranges_opt_random_25 = ['AM6', 'AM30', 'AM38', 'AM62', 'AM69', 'AM93', 'AM100', 'AM124', 'AM131', 'AM155']
l_ranges_max_random_25 = ['AN6', 'AN30', 'AN38', 'AN62', 'AN69', 'AN93', 'AN100', 'AN124', 'AN131', 'AN155']
l_ranges_uni_random_25 = ['AO6', 'AO30', 'AO38', 'AO62', 'AO69', 'AO93', 'AO100', 'AO124', 'AO131', 'AO155']


k_ranges_anti_50 = ['E6', 'E55', 'E63', 'E112', 'E120', 'E169', 'E177', 'E226', 'E234', 'E283']
l_ranges_opt_anti_50 = ['F6', 'F55', 'F63', 'F112', 'F120', 'F169', 'F177', 'F226', 'F234', 'F283']
l_ranges_max_anti_50 = ['G6', 'G55', 'G63', 'G112', 'G120', 'G169', 'G177', 'G226', 'G234', 'G283']
l_ranges_uni_anti_50 = ['H6', 'H55', 'H63', 'H112', 'H120', 'H169', 'H177', 'H226', 'H234', 'H283']

k_ranges_corr_50 = ['V6', 'V55', 'V63', 'V112', 'V120', 'V169', 'V177', 'V226', 'V234', 'V283']
l_ranges_opt_corr_50 = ['W6', 'W55', 'W63', 'W112', 'W120', 'W169', 'W177', 'W226', 'W234', 'W283']
l_ranges_max_corr_50 = ['X6', 'X55', 'X63', 'X112', 'X120', 'X169', 'X177', 'X226', 'X234', 'X283']
l_ranges_uni_corr_50 = ['Y6', 'Y55', 'Y63', 'Y112', 'Y120', 'Y169', 'Y177', 'Y226', 'Y234', 'Y283']

k_ranges_random_50 = ['AL6', 'AL55', 'AL63', 'AL112', 'AL120', 'AL169', 'AL177', 'AL226', 'AL234', 'AL283']
l_ranges_opt_random_50 = ['AM6', 'AM55', 'AM63', 'AM112', 'AM120', 'AM169', 'AM177', 'AM226', 'AM234', 'AM283']
l_ranges_max_random_50 = ['AN6', 'AN55', 'AN63', 'AN112', 'AN120', 'AN169', 'AN177', 'AN226', 'AN234', 'AN283']
l_ranges_uni_random_50 = ['AO6', 'AO55', 'AO63', 'AO112', 'AO120', 'AO169', 'AO177', 'AO226', 'AO234', 'AO283']

BASE_FOLDER = "../H2_ALSH/qhull_data/Synthetic/"
BASH_FILE_BASE_FOLDER = "../H2_ALSH/"


def separate_string(data_type_, budgets_, dimensions_, excel_file_, top_ks_, types_, card_excel_, cardinality_, save_file_path_):
    wb = load_workbook(filename=excel_file_, data_only=True)

    for i in range(top_ks_.__len__()):
        top_k = top_ks[i]
        if top_k == 10:
            k_ranges_anti = k_ranges_anti_10
            l_ranges_opt_anti = l_ranges_opt_anti_10
            l_ranges_max_anti = l_ranges_max_anti_10
            l_ranges_uni_anti = l_ranges_uni_anti_10

            k_ranges_corr = k_ranges_corr_10
            l_ranges_opt_corr = l_ranges_opt_corr_10
            l_ranges_max_corr = l_ranges_max_corr_10
            l_ranges_uni_corr = l_ranges_uni_corr_10

            k_ranges_random = k_ranges_random_10
            l_ranges_opt_random = l_ranges_opt_random_10
            l_ranges_max_random = l_ranges_max_random_10
            l_ranges_uni_random = l_ranges_uni_random_10
        elif top_k == 25:
            k_ranges_anti = k_ranges_anti_25
            l_ranges_opt_anti = l_ranges_opt_anti_25
            l_ranges_max_anti = l_ranges_max_anti_25
            l_ranges_uni_anti = l_ranges_uni_anti_25

            k_ranges_corr = k_ranges_corr_25
            l_ranges_opt_corr = l_ranges_opt_corr_25
            l_ranges_max_corr = l_ranges_max_corr_25
            l_ranges_uni_corr = l_ranges_uni_corr_25

            k_ranges_random = k_ranges_random_25
            l_ranges_opt_random = l_ranges_opt_random_25
            l_ranges_max_random = l_ranges_max_random_25
            l_ranges_uni_random = l_ranges_uni_random_25
        else:
            k_ranges_anti = k_ranges_anti_50
            l_ranges_opt_anti = l_ranges_opt_anti_50
            l_ranges_max_anti = l_ranges_max_anti_50
            l_ranges_uni_anti = l_ranges_uni_anti_50

            k_ranges_corr = k_ranges_corr_50
            l_ranges_opt_corr = l_ranges_opt_corr_50
            l_ranges_max_corr = l_ranges_max_corr_50
            l_ranges_uni_corr = l_ranges_uni_corr_50

            k_ranges_random = k_ranges_random_50
            l_ranges_opt_random = l_ranges_opt_random_50
            l_ranges_max_random = l_ranges_max_random_50
            l_ranges_uni_random = l_ranges_uni_random_50

        for j in range(budgets_.__len__()):
            budget = budgets_[j]
            for k in range(dimensions_.__len__()):
                dimension = dimensions_[k]
                for cc in range(cardinality_.__len__()):
                    sheetname = "Budget_" + str(budget) + "_" + str(dimension) + "D_top" + str(top_k) + "_" + card_excel_[cc]
                    if sheetname in wb.sheetnames:
                        print("Sheetname: " + str(sheetname))
                        ws = wb[sheetname]
                        for m in range(types_.__len__()):
                            type_name = types[m]
                            start = 2 * m
                            end = 2 * m + 1
                            # read data type anti_correlated
                            k_anti = []
                            for row in ws[k_ranges_anti[start]: k_ranges_anti[end]]:
                                for cell in row:
                                    k_anti.append(cell.value)

                            l_anti_opt = []
                            for row in ws[l_ranges_opt_anti[start]: l_ranges_opt_anti[end]]:
                                for cell in row:
                                    l_anti_opt.append(cell.value)

                            l_anti_max = []
                            for row in ws[l_ranges_max_anti[start]: l_ranges_max_anti[end]]:
                                for cell in row:
                                    l_anti_max.append(cell.value)

                            l_anti_uni = []
                            for row in ws[l_ranges_uni_anti[start]: l_ranges_uni_anti[end]]:
                                for cell in row:
                                    l_anti_uni.append(cell.value)

                            # read data type correlated
                            k_corr = []
                            for row in ws[k_ranges_corr[start]: k_ranges_corr[end]]:
                                for cell in row:
                                    k_corr.append(cell.value)

                            l_corr_opt = []
                            for row in ws[l_ranges_opt_corr[start]: l_ranges_opt_corr[end]]:
                                for cell in row:
                                    l_corr_opt.append(cell.value)

                            l_corr_max = []
                            for row in ws[l_ranges_max_corr[start]: l_ranges_max_corr[end]]:
                                for cell in row:
                                    l_corr_max.append(cell.value)

                            l_corr_uni = []
                            for row in ws[l_ranges_uni_corr[start]: l_ranges_uni_corr[end]]:
                                for cell in row:
                                    l_corr_uni.append(cell.value)

                            # read data type random
                            k_random = []
                            for row in ws[k_ranges_random[start]: k_ranges_random[end]]:
                                for cell in row:
                                    k_random.append(cell.value)

                            l_random_opt = []
                            for row in ws[l_ranges_opt_random[start]: l_ranges_opt_random[end]]:
                                for cell in row:
                                    l_random_opt.append(cell.value)

                            l_random_max = []
                            for row in ws[l_ranges_max_random[start]: l_ranges_max_random[end]]:
                                for cell in row:
                                    l_random_max.append(cell.value)

                            l_random_uni = []
                            for row in ws[l_ranges_uni_random[start]: l_ranges_uni_random[end]]:
                                for cell in row:
                                    l_random_uni.append(cell.value)
                            # save current K and L parameters to files
                            # anti_opt
                            # anti_uni
                            save_file_dir = save_file_path_ + str(dimension) + "D_top" + str(top_k) + "_budget_" + str(budget)\
                                            + "_" + type_name + "_" + card_excel_[cc] + "/"

                            if not os.path.exists(save_file_dir):
                                os.makedirs(save_file_dir)
                            anti_k_name = save_file_dir + "k_anti_correlated"
                            f = open(anti_k_name, 'w')
                            f.write(','.join(map(str, k_anti)))
                            f.close()

                            anti_opt_name = save_file_dir + "l_anti_correlated_opt"
                            f = open(anti_opt_name, 'w')
                            f.write(','.join(map(str, l_anti_opt)))
                            f.close()

                            anti_max_name = save_file_dir + "l_anti_correlated_max"
                            f = open(anti_max_name, 'w')
                            f.write(','.join(map(str, l_anti_max)))
                            f.close()

                            anti_uni_name = save_file_dir + "l_anti_correlated_uni"
                            f = open(anti_uni_name, 'w')
                            f.write(','.join(map(str, l_anti_uni)))
                            f.close()

                            # corr_opt
                            # corr_uni
                            corr_k_name = save_file_dir + "k_correlated"
                            f = open(corr_k_name, 'w')
                            f.write(','.join(map(str, k_corr)))
                            f.close()

                            corr_opt_name = save_file_dir + "l_correlated_opt"
                            f = open(corr_opt_name, 'w')
                            f.write(','.join(map(str, l_corr_opt)))
                            f.close()

                            corr_max_name = save_file_dir + "l_correlated_max"
                            f = open(corr_max_name, 'w')
                            f.write(','.join(map(str, l_corr_max)))
                            f.close()

                            corr_uni_name = save_file_dir + "l_correlated_uni"
                            f = open(corr_uni_name, 'w')
                            f.write(','.join(map(str, l_corr_uni)))
                            f.close()

                            # random_opt
                            # random_uni
                            random_k_name = save_file_dir + "k_random"
                            f = open(random_k_name, 'w')
                            f.write(','.join(map(str, k_random)))
                            f.close()

                            random_opt_name = save_file_dir + "l_random_opt"
                            f = open(random_opt_name, 'w')
                            f.write(','.join(map(str, l_random_opt)))
                            f.close()

                            random_max_name = save_file_dir + "l_random_max"
                            f = open(random_max_name, 'w')
                            f.write(','.join(map(str, l_random_max)))
                            f.close()

                            random_uni_name = save_file_dir + "l_random_uni"
                            f = open(random_uni_name, 'w')
                            f.write(','.join(map(str, l_random_uni)))
                            f.close()
    wb.close()
    print("Separate String and Save Parameter Files Done .\n")


def write_script(data_type_, budgets_, dimensions_, top_ks_, types_, card_excel_, cardinality_, parameter_path_, with_without_opt_, pot_):
    file_names = []
    for k in range(dimensions_.__len__()):
        dimension = dimensions_[k]
        for cc in range(cardinality_.__len__()):
            total_bash_file = BASH_FILE_BASE_FOLDER + "run_bash_set_cur_" + str(dimension) + 'D_' + card_excel_[cc] + \
                              '_' + str(with_without_opt_) + '.sh'
            file_names.append(total_bash_file)
            f10 = open(total_bash_file, 'w')
            f10.write("#!/bin/bash \n")
            for m in range(types_.__len__()):
                type_name = types_[m]
                for j in range(budgets_.__len__()):
                    budget = budgets_[j]
                    for i in range(top_ks_.__len__()):
                        top_k = top_ks_[i]
                        parameter_dir = parameter_path_ + str(dimension) + "D_top" + str(
                            top_k) + "_budget_" + str(budget) \
                                        + "_" + type_name + "_" + card_excel_[cc] + "/"
                        print(parameter_dir + ' ' + type_name)
                        if not os.path.exists(parameter_dir):
                            continue
                        else:
                            BASH_FILE_FOLDER = BASH_FILE_BASE_FOLDER + "bash_set_" + str(dimension) + "D_top" + str(
                                top_k) + "_budget_" + str(budget) + "_" + type_name + "_" + str(
                                cardinality_[cc]) + "/"
                            TEMPORAL_RESULT = "../H2_ALSH/temp_result_" + str(dimension) + "D_top" + str(
                                top_k) + "_budget_" + str(budget) + "_" + type_name + "_" + str(cardinality_[cc]) + "/"
                            TEMPORAL_RESULT_FOR_BASH = "./temp_result_" + str(dimension) + "D_top" + str(
                                top_k) + "_budget_" + str(budget) + "_" + type_name + "_" + str(cardinality_[cc]) + "/"

                            if not os.path.exists(BASH_FILE_FOLDER):
                                os.makedirs(BASH_FILE_FOLDER)

                            if not os.path.exists(TEMPORAL_RESULT):
                                os.makedirs(TEMPORAL_RESULT)

                            cur_bash_set_dir = "bash_set_" + str(dimension) + "D_top" + str(top_k) + \
                                               "_budget_" + str(budget) + "_" + type_name + "_" + str(
                                cardinality_[cc]) + "/"
                            for ii in range(len(data_type)):
                                K_List = []
                                L_Opt_List = []
                                L_Max_List = []
                                L_Uni_List = []
                                qhull_data_count = []

                                paramK_path = parameter_dir + "k_" + data_type_[ii] #str(cardinality[k])
                                f1 = open(paramK_path, 'r')
                                K_lines = f1.readlines()
                                for k_index in range(top_k):
                                    K_List.append(int(K_lines[0].split(',')[k_index]))
                                f1.close()

                                paramL_opt_path = parameter_dir + "l_" + data_type_[ii] + "_opt" # str(cardinality[k])
                                f2 = open(paramL_opt_path, 'r')
                                L_lines = f2.readlines()
                                for l_index in range(0, top_k):
                                    L_Opt_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                                f2.close()

                                paramL_max_path = parameter_dir + "l_" + data_type_[ii] + "_max"  # str(cardinality[k])
                                f2 = open(paramL_max_path, 'r')
                                L_lines = f2.readlines()
                                for l_index in range(0, top_k):
                                    L_Max_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                                f2.close()

                                paramL_uni_path = parameter_dir + "l_" + data_type_[ii] + "_uni"  # str(cardinality[k])
                                f2 = open(paramL_uni_path, 'r')
                                L_lines = f2.readlines()

                                for l_index in range(0, top_k):
                                    L_Uni_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                                f2.close()

                                for mm in range(len(K_List)):
                                    qhull_file = BASE_FOLDER + data_type_[ii] + "_" + str(dimensions[k]) + "_" + str(cardinality[cc])\
                                                 + "_" + "qhull_layer_" + str(mm)
                                    f = open(qhull_file, 'r')
                                    lines = f.readlines()
                                    second_line = lines[1]
                                    cur_data_count = int(second_line.split('\n')[0])
                                    qhull_data_count.append(cur_data_count)
                                    f.close()

                                # opt cumsum_hashsize
                                obj_cumsum = []
                                hashsize_cumsum = []
                                obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_opt_" + data_type_[ii] + "_" + str(
                                    dimensions_[k]) + "_" + str(cardinality_[cc]) + '_' + str(with_without_opt_) + ".txt"
                                f4 = open(obj_hashsize_file, 'w')
                                for mm in range(len(L_Opt_List)):
                                    if obj_cumsum.__len__() == 0:
                                        obj_cumsum.append(qhull_data_count[mm])
                                        hashsize_cumsum.append(qhull_data_count[mm] * L_Opt_List[mm])
                                    else:
                                        obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                        hashsize_cumsum.append(
                                            hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] * L_Opt_List[mm])
                                f4.write(','.join(map(repr, obj_cumsum)))
                                f4.write("\n")
                                f4.write(','.join(map(repr, hashsize_cumsum)))
                                f4.close()

                                # max cumsum_hashsize
                                obj_cumsum = []
                                hashsize_cumsum = []
                                obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_max_" + data_type_[ii] + \
                                                    "_" + str(dimensions_[k]) + "_" + str(cardinality_[cc]) + '_' + \
                                                    str(with_without_opt_) + ".txt"
                                f4 = open(obj_hashsize_file, 'w')
                                for mm in range(len(L_Max_List)):
                                    if obj_cumsum.__len__() == 0:
                                        obj_cumsum.append(qhull_data_count[mm])
                                        hashsize_cumsum.append(qhull_data_count[mm] * L_Max_List[mm])
                                    else:
                                        obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                        hashsize_cumsum.append(
                                            hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] *
                                            L_Max_List[mm])
                                f4.write(','.join(map(repr, obj_cumsum)))
                                f4.write("\n")
                                f4.write(','.join(map(repr, hashsize_cumsum)))
                                f4.close()

                                # uni cumsum hashsize
                                obj_cumsum = []
                                hashsize_cumsum = []
                                obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_uni_" + data_type_[ii] + "_" + str(
                                    dimensions_[k]) + "_" + str(cardinality_[cc]) + '_' + str(with_without_opt_) + ".txt"
                                f4 = open(obj_hashsize_file, 'w')
                                for mm in range(len(L_Uni_List)):
                                    if obj_cumsum.__len__() == 0:
                                        obj_cumsum.append(qhull_data_count[mm])
                                        hashsize_cumsum.append(qhull_data_count[mm] * L_Uni_List[mm])
                                    else:
                                        obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                        hashsize_cumsum.append(
                                            hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] * L_Uni_List[mm])
                                f4.write(','.join(map(repr, obj_cumsum)))
                                f4.write("\n")
                                f4.write(','.join(map(repr, hashsize_cumsum)))
                                f4.close()

                                cur_data_type = data_type_[ii]
                                cur_cardinality = cardinality_[cc]
                                cur_dimension = dimensions_[k]
                                f10.write("datatype=" + cur_data_type + "\n")
                                f10.write("cardinality=" + str(cur_cardinality) + "\n")
                                f10.write("d=" + str(cur_dimension) + "\n")
                                f10.write("qn=" + str(query_count) + "\n")
                                f10.write("c0=" + str(ratio) + "\n")
                                f10.write("pot=" + str(pot_) + "\n")

                                temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_opt"
                                overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_opt"
                                sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_opt"
                                f10.write("temporalResult=" + temporalResult + "\n")
                                f10.write("overallResult=" + overallResult + "\n")
                                f10.write("sim_angle=" + sim_angle + "\n")
                                f10.write("S=" + str(sim_threshold) + "\n")
                                f10.write("num_layer=" + str(len(K_List)) + "\n")
                                f10.write("top_k=" + str(top_k) + "\n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Ground-Truth \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                                f10.write(
                                    "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                                f10.write("qPath=./query/query_${d}D.txt \n")
                                f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")

                                f10.write(" # ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                                          "${oFolder}.mip \n")
                                f10.write(" # sleep 1 \n")

                                f10.write("\n \n \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Layer-Performance \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                for kk in range(len(K_List)):
                                    f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                                    f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                                    f10.write("L" + str(kk) + "=" + str(L_Opt_List[kk]) + "\n")

                                    f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                                                 "${cardinality}_qhull_layer_" + str(kk) + "\n")

                                    f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                                   "${cardinality}_opt/result_${d}D" + str(kk) + "_${K" + str(kk)
                                             + "}_${L" + str(kk) + "}" + "\n")
                                    f10.write("temp_hash" + str(
                                        kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_opt_" + str(kk) + "\n")

                                    f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                                             "} -L ${L" + str(kk) + "} -LI " + str(
                                        kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                                             + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                                        kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                                    f10.write("\n")

                                # append overall accuracy computation here
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Overall-Performance \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                                         "${tsPath}.mip -of ${overallResult} \n \n \n")
                                f10.write("sleep 2 \n")

                                cur_data_type = data_type_[ii]
                                cur_cardinality = cardinality_[cc]
                                cur_dimension = dimensions_[k]
                                f10.write("datatype=" + cur_data_type + "\n")
                                f10.write("cardinality=" + str(cur_cardinality) + "\n")
                                f10.write("d=" + str(cur_dimension) + "\n")
                                f10.write("qn=" + str(query_count) + "\n")
                                f10.write("c0=" + str(ratio) + "\n")
                                f10.write("pot=" + str(pot_) + "\n")

                                temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_max"
                                overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_max"
                                sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_max"
                                f10.write("temporalResult=" + temporalResult + "\n")
                                f10.write("overallResult=" + overallResult + "\n")
                                f10.write("sim_angle=" + sim_angle + "\n")
                                f10.write("S=" + str(sim_threshold) + "\n")
                                f10.write("num_layer=" + str(len(K_List)) + "\n")
                                f10.write("top_k=" + str(top_k) + "\n")
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Ground-Truth \n")
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                                f10.write(
                                    "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                                f10.write("qPath=./query/query_${d}D.txt \n")
                                f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
                                f10.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                                         "${oFolder}.mip \n")
                                f10.write("\n \n \n")
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Layer-Performance \n")
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                for kk in range(len(K_List)):
                                    f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                                    f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                                    f10.write("L" + str(kk) + "=" + str(L_Max_List[kk]) + "\n")

                                    f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                                                 "${cardinality}_qhull_layer_" + str(kk) + "\n")

                                    f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                                   "${cardinality}_max/result_${d}D" + str(
                                        kk) + "_${K" + str(kk)
                                             + "}_${L" + str(kk) + "}" + "\n")
                                    f10.write("temp_hash" + str(
                                        kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_max_" + str(kk) + "\n")

                                    f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                                             "} -L ${L" + str(kk) + "} -LI " + str(
                                        kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                                             + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                                        kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                                    f10.write("\n")

                                # append overall accuracy computation here
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Overall-Performance \n")
                                f10.write(
                                    "# ------------------------------------------------------------------------------ \n")
                                f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                                         "${tsPath}.mip -of ${overallResult} \n \n \n")
                                f10.write("sleep 2 \n")

                                cur_data_type = data_type_[ii]
                                cur_cardinality = cardinality_[cc]
                                cur_dimension = dimensions_[k]
                                f10.write("datatype=" + cur_data_type + "\n")
                                f10.write("cardinality=" + str(cur_cardinality) + "\n")
                                f10.write("d=" + str(cur_dimension) + "\n")
                                f10.write("qn=" + str(query_count) + "\n")
                                f10.write("c0=" + str(ratio) + "\n")
                                f10.write("pot=" + str(pot_) + "\n")

                                temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_uni"
                                overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_uni"
                                sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_uni"
                                f10.write("temporalResult=" + temporalResult + "\n")
                                f10.write("overallResult=" + overallResult + "\n")
                                f10.write("sim_angle=" + sim_angle + "\n")
                                f10.write("S=" + str(sim_threshold) + "\n")
                                f10.write("num_layer=" + str(len(K_List)) + "\n")
                                f10.write("top_k=" + str(top_k) + "\n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Ground-Truth \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                                f10.write(
                                    "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                                f10.write("qPath=./query/query_${d}D.txt \n")
                                f10.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
                                f10.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                                         "${oFolder}.mip \n")
                                f10.write("\n \n \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Layer-Performance \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                for kk in range(len(K_List)):
                                    f10.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                                    f10.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                                    f10.write("L" + str(kk) + "=" + str(L_Uni_List[kk]) + "\n")

                                    f10.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                                                 "${cardinality}_qhull_layer_" + str(kk) + "\n")

                                    f10.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                                   "${cardinality}_uni/result_${d}D" + str(kk) + "_${K" + str(kk)
                                             + "}_${L" + str(kk) + "}" + "\n")
                                    f10.write("temp_hash" + str(kk) + "=" + TEMPORAL_RESULT_FOR_BASH + "hash_proj_${datatype}_uni_" + str(kk) + "\n")

                                    f10.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                                             "} -L ${L" + str(kk) + "} -LI " + str(
                                        kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                                             + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                                        kk) + "}.simple_LSH -hr ${temp_hash" + str(kk) + "} -pot ${pot} \n")
                                    f10.write("\n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("#     Overall-Performance \n")
                                f10.write("# ------------------------------------------------------------------------------ \n")
                                f10.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                                         "${tsPath}.mip -of ${overallResult} \n \n \n")
                                f10.write("sleep 2 \n")
            f10.close()
    return file_names
                # summarize output for pot = 0 situation + parameter with without_post_opt
                # f10.write('cd ../Python_Analysis/LSH_Post_Process_' + str(dimension) + 'D.py without_post_opt ' + str(
                #         rr) + ' \n')
                # f10.write('python ../Python_Analysis/LSH_Post_Process_' + str(dimension) + 'D.py without_post_opt ' + str(rr) + ' \n')
                # f10.write('sleep 3' + '\n')
                #
                # # before starting pot = 1, clean everything except hash_table
                # f10.write('python ../Python_Analysis/Clean_Sim_Overall_run_test_' + str(dimension) + 'D.py' + '\n')
                # f10.write('sleep 5' + '\n')


################################################################
# data_type = ["anti_correlated", "correlated", "random"]
# # card_excel = ['100k', '200k', '500k', '1M', '15M', '2M']
# # cardinality = [100000, 200000, 500000, 1000000, 1500000, 2000000]
# card_excel = ['200k']
# cardinality = [200000]
#
# types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
#
# budgets = ["1M", "10M"]
# sim_threshold = 0.75
# dimensions = [4]
# excel_file_before = "./4D_075_redundancy_3_all_before.xlsx"
# excel_file_after = "./4D_075_redundancy_3_all_after.xlsx"
# top_ks = [10, 25, 50]


################################################################
data_type = ["anti_correlated", "correlated", "random"]
card_excel = ['200k']
cardinality = [200000]
types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]
budgets = ["1M", "10M"]
sim_threshold = 0.75
dimensions = [4]
excel_file_before = "./4D_075_redundancy_4_before.xlsx"
excel_file_after = "./4D_075_redundancy_4_after.xlsx"
top_ks = [10, 25]


################################################################

save_file_path_before = '../H2_ALSH/parameters_before/'
save_file_path_after = '../H2_ALSH/parameters_after/'

separate_string(data_type, budgets, dimensions, excel_file_before, top_ks, types, card_excel, cardinality, save_file_path_before)
separate_string(data_type, budgets, dimensions, excel_file_after, top_ks, types, card_excel, cardinality, save_file_path_after)

# write and complete ground truth shell
query_count = 1000
ratio = 2
parameter_path_before = '../H2_ALSH/parameters_before/'
parameter_path_after = '../H2_ALSH/parameters_after/'
parameter_type = ["opt", "max", "uni"]

with_without_opt = 'without_opt'
pot = 0
file_names_without_opt = write_script(data_type, budgets, dimensions, top_ks, types, card_excel, cardinality, parameter_path_before, with_without_opt, pot)


with_without_opt = 'with_opt'
pot = 1
file_names_with_opt = write_script(data_type, budgets, dimensions, top_ks, types, card_excel, cardinality, parameter_path_after, with_without_opt, pot)

repeated_run = 5

aggregated_file_name = BASH_FILE_BASE_FOLDER + "run_bash_" + str(dimensions[0]) + "D_all.sh"
f1 = open(aggregated_file_name, 'w')
f1.write("#!/bin/bash \n")
for rr in range(0, repeated_run):
    for files in file_names_without_opt:
        f1.write('sh ' + files + '\n')

    temp_str = "aggregating for non-opt round " + str(rr)
    f1.write('echo \"' + temp_str + '\" \n')

    f1.write('python ../Python_Analysis/LSH_Post_Process_' + str(dimensions[0]) + 'D.py without_opt ' + str(rr) + ' \n')
    f1.write('sleep 3' + '\n')

    # before starting pot = 1, clean everything except hash_table
    f1.write('python ../Python_Analysis/Clean_Sim_Overall_run_test_' + str(dimensions[0]) + 'D.py' + '\n')
    f1.write('sleep 5' + '\n')

    for files in file_names_with_opt:
        f1.write('sh ' + files + '\n')

    temp_str = "aggregating for opt round " + str(rr)
    f1.write('echo \"' + temp_str + '\" \n')

    f1.write('python ../Python_Analysis/LSH_Post_Process_' + str(dimensions[0]) + 'D.py with_opt ' + str(rr) + ' \n')
    f1.write('sleep 3' + '\n')

    # before starting pot = 1, clean everything except hash_table
    f1.write('python ../Python_Analysis/Clean_All_' + str(dimensions[0]) + 'D.py' + '\n')
    f1.write('sleep 5' + '\n')

f1.close()