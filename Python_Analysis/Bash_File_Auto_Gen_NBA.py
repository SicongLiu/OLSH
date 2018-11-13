# script to generate .sh file for execution
# bash file includes execution of Ground Truth, LSH Scheme and Overall Performance

import math
import os
from openpyxl import load_workbook

saguaro_script_string = '#SBATCH -p serial \t \t # Send this job to the serial partition \n' \
                        '#SBATCH -n 4 # number of cores \n' \
                        '#SBATCH --mem=32000                 # 32 GB \n' \
                        '#SBATCH -t 0-15:00                  # wall time (D-HH:MM) \n' \
                        '#SBATCH -o slurm.%j.out             # STDOUT (%j = JobId) \n' \
                        '#SBATCH -e slurm.%j.err             # STDERR (%j = JobId) \n' \
                        '#SBATCH --mail-type=END,FAIL        # notifications for job done & fail \n' \
                        '#SBATCH --mail-user=sliu104@asu.edu # send-to address \n' \
                        'module load gcc/4.9.2 \n'
data_type = ["NBA"]
budgets = ["300k", "500k", "1M"]
dimensions = [7]
max_layer = 9
top_ks = [10, 25, 50]
types = ["log", "log_minus", "log_plus", "log_plus_plus", "uni"]

card_excel = ['23338']
cardinality = [23338]

k_ranges_NBA_10 = ['E6',  'E14', 'E20', 'E28', 'E35', 'E43', 'E48', 'E56', 'E64', 'E72']
l_ranges_opt_NBA_10 = ['F6', 'F14', 'F20', 'F28', 'F35', 'F43', 'F48', 'F56', 'F64', 'F72']
l_ranges_max_NBA_10 = ['G6', 'G14', 'G20', 'G28', 'G35', 'G43', 'G48', 'G56', 'G64', 'G72']
l_ranges_uni_NBA_10 = ['H6', 'H14', 'H20', 'H28', 'H35', 'H43', 'H48', 'H56', 'H64', 'H72']

k_ranges_NBA_25 = ['E6',  'E14', 'E20', 'E28', 'E35', 'E43', 'E48', 'E56', 'E64', 'E72']
l_ranges_opt_NBA_25 = ['F6', 'F14', 'F20', 'F28', 'F35', 'F43', 'F48', 'F56', 'F64', 'F72']
l_ranges_max_NBA_25 = ['G6', 'G14', 'G20', 'G28', 'G35', 'G43', 'G48', 'G56', 'G64', 'G72']
l_ranges_uni_NBA_25 = ['H6', 'H14', 'H20', 'H28', 'H35', 'H43', 'H48', 'H56', 'H64', 'H72']

k_ranges_NBA_50 = ['E6',  'E14', 'E20', 'E28', 'E35', 'E43', 'E48', 'E56', 'E64', 'E72']
l_ranges_opt_NBA_50 = ['F6', 'F14', 'F20', 'F28', 'F35', 'F43', 'F48', 'F56', 'F64', 'F72']
l_ranges_max_NBA_50 = ['G6', 'G14', 'G20', 'G28', 'G35', 'G43', 'G48', 'G56', 'G64', 'G72']
l_ranges_uni_NBA_50 = ['H6', 'H14', 'H20', 'H28', 'H35', 'H43', 'H48', 'H56', 'H64', 'H72']


excel_files = ['../Checkpoint_Result_Oct_23_NBA.xlsx']

for excel_file in excel_files:
    wb = load_workbook(filename=excel_file, data_only=True)

    save_file_path = '../H2_ALSH/parameters/NBA/'

    for i in range(top_ks.__len__()):
        top_k = top_ks[i]
        if top_k == 10:
            k_ranges_NBA = k_ranges_NBA_10
            l_ranges_opt_NBA = l_ranges_opt_NBA_10
            l_ranges_max_NBA = l_ranges_max_NBA_10
            l_ranges_uni_NBA = l_ranges_uni_NBA_10
        elif top_k == 25:
            k_ranges_NBA = k_ranges_NBA_25
            l_ranges_opt_NBA = l_ranges_opt_NBA_25
            l_ranges_max_NBA = l_ranges_max_NBA_25
            l_ranges_uni_NBA = l_ranges_uni_NBA_25
        else:
            k_ranges_NBA = k_ranges_NBA_50
            l_ranges_opt_NBA = l_ranges_opt_NBA_50
            l_ranges_max_NBA = l_ranges_max_NBA_50
            l_ranges_uni_NBA = l_ranges_uni_NBA_50

        for j in range(budgets.__len__()):
            budget = budgets[j]
            for k in range(dimensions.__len__()):
                dimension = dimensions[k]
                for cc in range(cardinality.__len__()):
                    # sheetname = "Budget_" + str(budget) + "_" + str(dimension) + "D_top" + str(top_k) + "_" + card_excel[cc]
                    sheetname = "Budget_" + str(budget)
                    if sheetname in wb.sheetnames:
                        ws = wb[sheetname]
                        for m in range(types.__len__()):
                            type_name = types[m]
                            start = 2 * m
                            end = 2 * m + 1
                            # read data type NBA_correlated
                            k_NBA = []
                            for row in ws[k_ranges_NBA[start]: k_ranges_NBA[end]]:
                                for cell in row:
                                    k_NBA.append(cell.value)

                            l_NBA_opt = []
                            for row in ws[l_ranges_opt_NBA[start]: l_ranges_opt_NBA[end]]:
                                for cell in row:
                                    l_NBA_opt.append(cell.value)

                            l_NBA_max = []
                            for row in ws[l_ranges_max_NBA[start]: l_ranges_max_NBA[end]]:
                                for cell in row:
                                    l_NBA_max.append(cell.value)

                            l_NBA_uni = []
                            for row in ws[l_ranges_uni_NBA[start]: l_ranges_uni_NBA[end]]:
                                for cell in row:
                                    l_NBA_uni.append(cell.value)

                            # save current K and L parameters to files
                            # NBA_opt
                            # NBA_uni
                            save_file_dir = save_file_path + str(dimension) + "D_top" + str(top_k) + "_budget_" + str(budget)\
                                            + "_" + type_name + "_" + card_excel[cc] + "/"

                            if not os.path.exists(save_file_dir):
                                os.makedirs(save_file_dir)
                            NBA_k_name = save_file_dir + "k_NBA"
                            f = open(NBA_k_name, 'w')
                            f.write(','.join(map(str, k_NBA)))
                            f.close()

                            NBA_opt_name = save_file_dir + "l_NBA_opt"
                            f = open(NBA_opt_name, 'w')
                            f.write(','.join(map(str, l_NBA_opt)))
                            f.close()

                            NBA_max_name = save_file_dir + "l_NBA_max"
                            f = open(NBA_max_name, 'w')
                            f.write(','.join(map(str, l_NBA_max)))
                            f.close()

                            NBA_uni_name = save_file_dir + "l_NBA_uni"
                            f = open(NBA_uni_name, 'w')
                            f.write(','.join(map(str, l_NBA_uni)))
                            f.close()
    wb.close()
print("Done .\n")

query_count = 1000
ratio = 2
sim_threshold = 0.9

parameter_path = '../H2_ALSH/parameters/NBA/'
parameter_type = ["opt", "uni"]
BASE_FOLDER = "../H2_ALSH/qhull_data/NBA/"
BASH_FILE_BASE_FOLDER = "../H2_ALSH/"

dir_items = os.listdir(parameter_path)
ground_truth_path = BASH_FILE_BASE_FOLDER + "all_ground_truth_NBA.sh"
f33 = open(ground_truth_path, 'w')
f33.write("#!/bin/bash \n")
f33.write(saguaro_script_string)

for k in range(dimensions.__len__()):
    dimension = dimensions[k]
    for cc in range(cardinality.__len__()):
        cur_cardinality = cardinality[cc]
        # if any(x.startswith('runner') for x in os.listdir('/path/to/runners')):
        for ii in range(len(data_type)):
            # write ground truth to bash file
            cur_data_type = data_type[ii]

            cur_dimension = dimensions[k]
            f33.write("datatype=" + cur_data_type + "\n")
            f33.write("cardinality=" + str(cur_cardinality) + "\n")
            f33.write("d=" + str(cur_dimension) + "\n")
            f33.write("qn=" + str(query_count) + "\n")
            f33.write(
                "# ------------------------------------------------------------------------------ \n")
            f33.write("#     Ground-Truth \n")
            f33.write(
                "# ------------------------------------------------------------------------------ \n")
            f33.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
            f33.write(
                "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
            f33.write("qPath=./query/query_${d}D.txt \n")
            f33.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")

            f33.write(
                "  ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                "${oFolder}.mip sleep 2\n")
f33.close()


for k in range(dimensions.__len__()):
    dimension = dimensions[k]
    for cc in range(cardinality.__len__()):
        total_bash_file = BASH_FILE_BASE_FOLDER + "run_bash_set_cur_" + str(dimension) + 'D_' + card_excel[cc] + '.sh'
        f10 = open(total_bash_file, 'w')
        f10.write("#!/bin/bash \n")

        for m in range(types.__len__()):
            type_name = types[m]
            for j in range(budgets.__len__()):
                budget = budgets[j]
                for i in range(top_ks.__len__()):
                    top_k = top_ks[i]
                    parameter_dir = parameter_path + str(dimension) + "D_top" + str(top_k) + "_budget_" + str(budget)\
                                        + "_" + type_name + "_" + card_excel[cc] + "/"
                    print(parameter_dir)
                    if not os.path.exists(parameter_dir):
                        continue
                    else:
                        BASH_FILE_FOLDER = BASH_FILE_BASE_FOLDER + "bash_set_" + str(dimension) + "D_top" + str(top_k) + \
                                           "_budget_" + str(budget) + "_" + type_name + "_" + str(cardinality[cc]) + "/"
                        TEMPORAL_RESULT = "../H2_ALSH/temp_result_" + str(dimension) + "D_top" + str(top_k) + "_budget_" + \
                                          str(budget) + "_" + type_name + "_" + str(cardinality[cc]) + "/"
                        TEMPORAL_RESULT_FOR_BASH = "./temp_result_" + str(dimension) + "D_top" + str(top_k) + "_budget_" + \
                                                   str(budget) + "_" + type_name + "_" + str(cardinality[cc]) + "/"

                        if not os.path.exists(BASH_FILE_FOLDER):
                            os.makedirs(BASH_FILE_FOLDER)

                        if not os.path.exists(TEMPORAL_RESULT):
                            os.makedirs(TEMPORAL_RESULT)

                        cur_bash_set_dir = "bash_set_" + str(dimension) + "D_top" + str(top_k) + \
                                           "_budget_" + str(budget) + "_" + type_name + "_" + str(cardinality[cc]) + "/"
                        # ground_truth_path = BASH_FILE_FOLDER + "ground_truth.sh"
                        # f10.write('sbatch ./' + cur_bash_set_dir + 'ground_truth.sh sleep 50 \n')
                        for ii in range(len(data_type)):
                            K_List = []
                            L_Opt_List = []
                            L_Max_List = []
                            L_Uni_List = []
                            qhull_data_count = []
                            bash_file_opt = BASH_FILE_FOLDER + "run_test_" + str(data_type[ii]) + "_" + str(dimensions[k]) \
                                            + "_" + str(cardinality[cc]) + "_opt.sh"

                            bash_file_max = BASH_FILE_FOLDER + "run_test_" + str(data_type[ii]) + "_" + str(
                                dimensions[k]) \
                                            + "_" + str(cardinality[cc]) + "_max.sh"

                            bash_file_uni = BASH_FILE_FOLDER + "run_test_" + str(data_type[ii]) + "_" + str(dimensions[k]) \
                                            + "_" + str(cardinality[cc]) + "_uni.sh"

                            f10.write('sbatch ./' + cur_bash_set_dir + "run_test_" + str(data_type[ii]) + "_" + str(dimensions[k]) + "_" + str(cardinality[cc]) + "_opt.sh sleep 2\n")
                            #f10.write('sbatch ./' + cur_bash_set_dir + "run_test_" + str(data_type[ii]) + "_" + str(dimensions[k]) + "_" + str(cardinality[cc]) + "_max.sh sleep 2\n")
                            f10.write('sbatch ./' + cur_bash_set_dir + "run_test_" + str(data_type[ii]) + "_" + str(dimensions[k]) + "_" + str(cardinality[cc]) + "_uni.sh sleep 2\n")
                            paramK_path = parameter_dir + "k_" + data_type[ii] #str(cardinality[k])
                            f1 = open(paramK_path, 'r')
                            K_lines = f1.readlines()
                            for k_index in range(max_layer):
                                K_List.append(int(K_lines[0].split(',')[k_index]))
                            f1.close()

                            paramL_opt_path = parameter_dir + "l_" + data_type[ii] + "_opt" # str(cardinality[k])
                            f2 = open(paramL_opt_path, 'r')
                            L_lines = f2.readlines()
                            for l_index in range(0, max_layer):
                                L_Opt_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                            f2.close()
                            # print("L opt: " + str(L_Opt_List))

                            paramL_max_path = parameter_dir + "l_" + data_type[ii] + "_max"  # str(cardinality[k])
                            f2 = open(paramL_max_path, 'r')
                            L_lines = f2.readlines()
                            for l_index in range(0, max_layer):
                                L_Max_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                            f2.close()
                            # print("L max: " + str(L_Max_List))

                            paramL_uni_path = parameter_dir + "l_" + data_type[ii] + "_uni"  # str(cardinality[k])
                            f2 = open(paramL_uni_path, 'r')
                            L_lines = f2.readlines()

                            for l_index in range(0, max_layer):
                                L_Uni_List.append(int(math.floor(float(L_lines[0].split(',')[l_index]))))
                            f2.close()
                            # print("L uni: " + str(L_Uni_List))

                            for mm in range(len(K_List)):
                                qhull_file = BASE_FOLDER + data_type[ii] + "_" + str(dimensions[k]) + "_" + str(cardinality[cc])\
                                             + "_" + "qhull_layer_" + str(mm)
                                f = open(qhull_file, 'r')
                                lines = f.readlines()
                                second_line = lines[1]
                                cur_data_count = int(second_line.split('\n')[0])
                                qhull_data_count.append(cur_data_count)
                                f.close()

                            # opt cumsum_hashsize
                            obj_cumsum = []
                            hashsize_cumsum = []
                            obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_opt_" + data_type[ii] + "_" + str(
                                dimensions[k]) +  "_" + str(cardinality[cc]) + ".txt"
                            f4 = open(obj_hashsize_file, 'w')
                            for mm in range(len(L_Opt_List)):
                                if obj_cumsum.__len__() == 0:
                                    obj_cumsum.append(qhull_data_count[mm])
                                    hashsize_cumsum.append(qhull_data_count[mm] * L_Opt_List[mm])
                                else:
                                    obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                    hashsize_cumsum.append(
                                        hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] * L_Opt_List[mm])
                            f4.write(','.join(map(repr, obj_cumsum)))
                            f4.write("\n")
                            f4.write(','.join(map(repr, hashsize_cumsum)))
                            f4.close()

                            # max cumsum_hashsize
                            obj_cumsum = []
                            hashsize_cumsum = []
                            obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_max_" + data_type[
                                ii] + "_" + str(
                                dimensions[k]) + "_" + str(cardinality[cc]) + ".txt"
                            f4 = open(obj_hashsize_file, 'w')
                            for mm in range(len(L_Max_List)):
                                if obj_cumsum.__len__() == 0:
                                    obj_cumsum.append(qhull_data_count[mm])
                                    hashsize_cumsum.append(qhull_data_count[mm] * L_Max_List[mm])
                                else:
                                    obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                    hashsize_cumsum.append(
                                        hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] *
                                        L_Max_List[mm])
                            f4.write(','.join(map(repr, obj_cumsum)))
                            f4.write("\n")
                            f4.write(','.join(map(repr, hashsize_cumsum)))
                            f4.close()

                            # uni cumsum hashsize
                            obj_cumsum = []
                            hashsize_cumsum = []
                            obj_hashsize_file = BASH_FILE_FOLDER + "cumsum_hashsize_obj_uni_" + data_type[ii] + "_" + str(
                                dimensions[k]) + "_" + str(cardinality[cc]) + ".txt"
                            f4 = open(obj_hashsize_file, 'w')
                            for mm in range(len(L_Uni_List)):
                                if obj_cumsum.__len__() == 0:
                                    obj_cumsum.append(qhull_data_count[mm])
                                    hashsize_cumsum.append(qhull_data_count[mm] * L_Uni_List[mm])
                                else:
                                    obj_cumsum.append(obj_cumsum[obj_cumsum.__len__() - 1] + qhull_data_count[mm])
                                    hashsize_cumsum.append(
                                        hashsize_cumsum[hashsize_cumsum.__len__() - 1] + qhull_data_count[mm] * L_Uni_List[mm])
                            f4.write(','.join(map(repr, obj_cumsum)))
                            f4.write("\n")
                            f4.write(','.join(map(repr, hashsize_cumsum)))
                            f4.close()



                            # write to .sh file at save path, opt
                            f3 = open(bash_file_opt, 'w')
                            f3.write("#!/bin/bash \n")
                            f3.write("# make \n")
                            f3.write(saguaro_script_string)
                            f3.write("rm *.o \n")
                            cur_data_type = data_type[ii]
                            cur_cardinality = cardinality[cc]
                            cur_dimension = dimensions[k]
                            f3.write("datatype=" + cur_data_type + "\n")
                            f3.write("cardinality=" + str(cur_cardinality) + "\n")
                            f3.write("d=" + str(cur_dimension) + "\n")
                            f3.write("qn=" + str(query_count) + "\n")
                            f3.write("c0=" + str(ratio) + "\n")

                            temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_opt"
                            # overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_opt.txt"
                            overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_opt"
                            sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_opt"
                            f3.write("temporalResult=" + temporalResult + "\n")
                            f3.write("overallResult=" + overallResult + "\n")
                            f3.write("sim_angle=" + sim_angle + "\n")
                            f3.write("S=" + str(sim_threshold) + "\n")
                            f3.write("num_layer=" + str(len(K_List)) + "\n")
                            f3.write("top_k=" + str(top_k) + "\n")
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            f3.write("#     Ground-Truth \n")
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            f3.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                            f3.write(
                                "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                            f3.write("qPath=./query/query_${d}D.txt \n")
                            f3.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")

                            f3.write(" # ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                                      "${oFolder}.mip \n")
                            f3.write(" # sleep 1 \n")

                            f3.write("\n \n \n")
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            f3.write("#     Layer-Performance \n")
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            for kk in range(len(K_List)):
                                f3.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                                f3.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                                f3.write("L" + str(kk) + "=" + str(L_Opt_List[kk]) + "\n")

                                f3.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                                             "${cardinality}_qhull_layer_" + str(kk) + "\n")

                                f3.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                               "${cardinality}_opt/result_${d}D" + str(kk) + "_${K" + str(kk)
                                         + "}_${L" + str(kk) + "}" + "\n")

                                f3.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                                         "} -L ${L" + str(kk) + "} -LI " + str(
                                    kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                                         + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                                    kk) + "}.simple_LSH \n")

                                f3.write("\n")
                            # append overall accuracy computation here
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            f3.write("#     Overall-Performance \n")
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            f3.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                                     "${tsPath}.mip -of ${overallResult} \n")
                            f3.close()

                            # # write to .sh file at save path, max
                            # f3 = open(bash_file_max, 'w')
                            # f3.write("#!/bin/bash \n")
                            # f3.write("# make \n")
                            # f3.write(saguaro_script_string)
                            # f3.write("rm *.o \n")
                            # cur_data_type = data_type[ii]
                            # cur_cardinality = cardinality[cc]
                            # cur_dimension = dimensions[k]
                            # f3.write("datatype=" + cur_data_type + "\n")
                            # f3.write("cardinality=" + str(cur_cardinality) + "\n")
                            # f3.write("d=" + str(cur_dimension) + "\n")
                            # f3.write("qn=" + str(query_count) + "\n")
                            # f3.write("c0=" + str(ratio) + "\n")
                            #
                            # temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_max"
                            # # overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_max.txt"
                            # overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_max"
                            # sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_max"
                            # f3.write("temporalResult=" + temporalResult + "\n")
                            # f3.write("overallResult=" + overallResult + "\n")
                            # f3.write("sim_angle=" + sim_angle + "\n")
                            # f3.write("S=" + str(sim_threshold) + "\n")
                            # f3.write("num_layer=" + str(len(K_List)) + "\n")
                            # f3.write("top_k=" + str(top_k) + "\n")
                            # f3.write(
                            #     "# ------------------------------------------------------------------------------ \n")
                            # f3.write("#     Ground-Truth \n")
                            # f3.write(
                            #     "# ------------------------------------------------------------------------------ \n")
                            # f3.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                            # f3.write(
                            #     "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                            # f3.write("qPath=./query/query_${d}D.txt \n")
                            # f3.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
                            # f3.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                            #          "${oFolder}.mip \n")
                            # f3.write("\n \n \n")
                            # f3.write(
                            #     "# ------------------------------------------------------------------------------ \n")
                            # f3.write("#     Layer-Performance \n")
                            # f3.write(
                            #     "# ------------------------------------------------------------------------------ \n")
                            # for kk in range(len(K_List)):
                            #     f3.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                            #     f3.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                            #     f3.write("L" + str(kk) + "=" + str(L_Max_List[kk]) + "\n")
                            #
                            #     f3.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                            #                                  "${cardinality}_qhull_layer_" + str(kk) + "\n")
                            #
                            #     f3.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                            #                                    "${cardinality}_max/result_${d}D" + str(
                            #         kk) + "_${K" + str(kk)
                            #              + "}_${L" + str(kk) + "}" + "\n")
                            #
                            #     f3.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                            #              "} -L ${L" + str(kk) + "} -LI " + str(
                            #         kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                            #              + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                            #         kk) + "}.simple_LSH \n")
                            #
                            #     f3.write("\n")
                            # # append overall accuracy computation here
                            # f3.write(
                            #     "# ------------------------------------------------------------------------------ \n")
                            # f3.write("#     Overall-Performance \n")
                            # f3.write(
                            #     "# ------------------------------------------------------------------------------ \n")
                            # f3.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                            #          "${tsPath}.mip -of ${overallResult} \n")
                            # f3.close()


                            # write to .sh file at save path, uni
                            f3 = open(bash_file_uni, 'w')
                            f3.write("#!/bin/bash \n")
                            f3.write("# make \n")
                            f3.write(saguaro_script_string)
                            f3.write("rm *.o \n")
                            cur_data_type = data_type[ii]
                            cur_cardinality = cardinality[cc]
                            cur_dimension = dimensions[k]
                            f3.write("datatype=" + cur_data_type + "\n")
                            f3.write("cardinality=" + str(cur_cardinality) + "\n")
                            f3.write("d=" + str(cur_dimension) + "\n")
                            f3.write("qn=" + str(query_count) + "\n")
                            f3.write("c0=" + str(ratio) + "\n")

                            temporalResult = TEMPORAL_RESULT_FOR_BASH + "run_test_${datatype}_${d}_${cardinality}_uni"
                            # overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_uni.txt"
                            overallResult = TEMPORAL_RESULT_FOR_BASH + "overall_run_test_${datatype}_${d}_${cardinality}_uni"
                            sim_angle = TEMPORAL_RESULT_FOR_BASH + "sim_angle_${datatype}_${d}_${cardinality}_uni"
                            f3.write("temporalResult=" + temporalResult + "\n")
                            f3.write("overallResult=" + overallResult + "\n")
                            f3.write("sim_angle=" + sim_angle + "\n")
                            f3.write("S=" + str(sim_threshold) + "\n")
                            f3.write("num_layer=" + str(len(K_List)) + "\n")
                            f3.write("top_k=" + str(top_k) + "\n")
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            f3.write("#     Ground-Truth \n")
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            f3.write("dPath=./raw_data/Synthetic/${datatype}_${d}_${cardinality}.txt \n")
                            f3.write(
                                "tsPath=./result/result_${datatype}_${d}D_${cardinality} # path for the ground truth \n")
                            f3.write("qPath=./query/query_${d}D.txt \n")
                            f3.write("oFolder=./result/result_${datatype}_${d}D_${cardinality} \n")
                            f3.write("# ./alsh -alg 0 -n ${cardinality} -qn ${qn} -d ${d} -ds ${dPath} -qs ${qPath} -ts "
                                     "${oFolder}.mip \n")
                            f3.write("\n \n \n")
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            f3.write("#     Layer-Performance \n")
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            for kk in range(len(K_List)):
                                f3.write("n" + str(kk) + "=" + str(qhull_data_count[kk]) + "\n")
                                f3.write("K" + str(kk) + "=" + str(K_List[kk]) + "\n")
                                f3.write("L" + str(kk) + "=" + str(L_Uni_List[kk]) + "\n")

                                f3.write("dPath" + str(kk) + "=./qhull_data/Synthetic/${datatype}_${d}_"
                                                             "${cardinality}_qhull_layer_" + str(kk) + "\n")

                                f3.write("oFolder" + str(kk) + "=./result/${datatype}/Dimension_${d}_Cardinality_"
                                                               "${cardinality}_uni/result_${d}D" + str(kk) + "_${K" + str(kk)
                                         + "}_${L" + str(kk) + "}" + "\n")

                                f3.write("./alsh -alg 10 -n ${n" + str(kk) + "} -qn ${qn} -d ${d} -K ${K" + str(kk) +
                                         "} -L ${L" + str(kk) + "} -LI " + str(
                                    kk + 1) + " -tk ${top_k}" + " -S ${S} -c0 ${c0} -ds ${dPath" + str(kk)
                                         + "} -qs ${qPath} -ts ${tsPath}.mip -it ${temporalResult} -sa ${sim_angle} -of ${oFolder" + str(
                                    kk) + "}.simple_LSH \n")

                                f3.write("\n")
                            # append overall accuracy computation here
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            f3.write("#     Overall-Performance \n")
                            f3.write("# ------------------------------------------------------------------------------ \n")
                            f3.write("./alsh -alg 12 -d ${d} -qn ${qn} -L1 ${num_layer} -tk ${top_k} -it ${temporalResult} -ts "
                                     "${tsPath}.mip -of ${overallResult} \n")
                            f3.close()
f10.close()

print("Done .\n")